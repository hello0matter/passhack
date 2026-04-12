import base64
import concurrent.futures
import contextlib
import csv
import hashlib
import html
import json
import os
import queue
import re
import shutil
import socket
import subprocess
import threading
import tempfile
import time
import warnings
import webbrowser
from collections import Counter
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, scrolledtext, ttk
import tkinter as tk
from urllib.parse import urljoin, urlparse

import requests
import urllib3
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from PIL import Image, ImageTk

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.styles.stylesheet")
warnings.filterwarnings("ignore", message="Workbook contains no default style, apply openpyxl's default")
os.environ.setdefault("WDM_LOG_LEVEL", "0")
os.environ.setdefault("TF_CPP_MIN_LOG_LEVEL", "3")
os.environ.setdefault("NO_COLOR", "1")

LOGIN_KEYWORDS = [
    "login",
    "signin",
    "sign in",
    "auth",
    "passport",
    "admin",
    "登录",
    "登 录",
    "用户登录",
    "统一认证",
    "身份认证",
    "后台",
    "管理平台",
]

CAPTCHA_KEYWORDS = [
    "验证码",
    "captcha",
    "verifycode",
    "checkcode",
    "图形码",
    "校验码",
]

MFA_KEYWORDS = [
    "otp",
    "totp",
    "2fa",
    "mfa",
    "双因素",
    "动态口令",
    "短信验证码",
    "邮箱验证码",
]

LOCKOUT_KEYWORDS = [
    "锁定",
    "冻结",
    "连续失败",
    "多次失败",
    "错误次数",
    "锁住",
    "account locked",
]

DEFAULT_HINT_KEYWORDS = [
    "默认账号",
    "默认密码",
    "初始密码",
    "admin",
    "administrator",
    "root",
    "guest",
    "superadmin",
]

INPUT_HINTS = [
    "用户名",
    "账号",
    "账户",
    "user",
    "account",
    "phone",
    "手机号",
]

DEFAULT_BRUTE_DICT_MODE = "默认中文常用字典"
CUSTOM_BRUTE_DICT_MODE = "自定义账号/密码字典"
BRUTE_FORCE_SUCCESS_PREFIX = "命中弱口令:"

DEFAULT_CN_USERNAMES = [
    "admin",
    "administrator",
    "root",
    "system",
    "sysadmin",
    "manager",
    "operator",
    "test",
    "user",
    "guest",
    "sa",
    "webadmin",
    "superadmin",
]

DEFAULT_CN_PASSWORDS = [
    "123456",
    "12345678",
    "123123",
    "123qwe",
    "1qaz2wsx",
    "111111",
    "000000",
    "666666",
    "888888",
    "admin",
    "admin123",
    "admin@123",
    "system",
    "system123",
    "qwe123",
    "qwe123456",
    "Aa123456",
    "password",
    "woaini1314",
]

DICT_FILE_ENCODINGS = ("utf-8-sig", "utf-8", "gb18030", "gbk", "utf-16")

USER_FIELD_KEYWORDS = [
    "user",
    "username",
    "userid",
    "login",
    "loginname",
    "account",
    "name",
    "email",
    "mail",
    "phone",
    "mobile",
    "tel",
    "member",
    "uid",
    "账号",
    "账户",
    "用户名",
    "登录名",
    "手机号",
    "手机",
    "邮箱",
    "工号",
    "学号",
]

PASSWORD_FIELD_KEYWORDS = ["password", "passwd", "pass", "pwd", "口令", "密码"]
CAPTCHA_FIELD_KEYWORDS = [
    "captcha",
    "verify",
    "verifycode",
    "checkcode",
    "randcode",
    "vcode",
    "imgcode",
    "yzm",
    "验证码",
    "校验码",
    "图形码",
]

SKIP_USERNAME_FIELD_KEYWORDS = [
    "captcha",
    "checkcode",
    "token",
    "otp",
    "mfa",
    "2fa",
    "sms",
    "验证码",
    "校验码",
    "动态码",
    "短信",
]

LOGIN_SUCCESS_KEYWORDS = [
    "登录成功",
    "welcome",
    "dashboard",
    "logout",
    "sign out",
    "退出登录",
    "退出",
    "注销",
    "管理首页",
    "系统首页",
    "控制台",
    "用户中心",
]

LOGIN_FAILURE_KEYWORDS = [
    "密码错误",
    "用户名或密码错误",
    "账号或密码错误",
    "登录失败",
    "invalid password",
    "invalid credentials",
    "authentication failed",
    "captcha error",
    "验证码错误",
]

BROWSER_RENDER_HINTS = [
    "vue",
    "react",
    "angular",
    "webpack",
    "chunk-vendors",
    "__next",
    "__nuxt",
    "id=\"app\"",
    "id='app'",
    "id=\"root\"",
    "id='root'",
    "layui",
    "element-ui",
    "ant-design",
]

LOGIN_BUTTON_KEYWORDS = [
    "登录",
    "登 录",
    "signin",
    "sign in",
    "submit",
    "提交",
    "进入系统",
    "进入平台",
]

COMMON_LOGIN_ROUTE_SUFFIXES = [
    "/#/login",
    "/#/pages/login/login",
    "/#/pages/index/login",
    "/#/user/login",
    "/#/auth/login",
    "/#/admin/login",
    "/#/home/login",
    "/#/index/login",
    "/login",
    "/signin",
    "/user/login",
    "/admin/login",
    "/auth/login",
]

DEFAULT_OCR_ENDPOINT = "http://127.0.0.1:8888/reg"
PROXY_MODE_SINGLE = "单代理/Clash"
PROXY_MODE_POOL = "代理池轮换"
CAPTURE_POLICY_HIT = "命中项"
CAPTURE_POLICY_HIGH = "仅高风险"
CAPTURE_POLICY_ALL = "全部截图"
CAPTURE_POLICY_LOGIN = "发现登录框截图"
CAPTURE_POLICY_OPTIONS = [
    CAPTURE_POLICY_HIT,
    CAPTURE_POLICY_HIGH,
    CAPTURE_POLICY_ALL,
    CAPTURE_POLICY_LOGIN,
]


def compact_exception_message(exc: Exception) -> str:
    text = " ".join(str(exc).split())
    return text or exc.__class__.__name__


def describe_request_exception(exc: Exception) -> tuple[str, str]:
    detail = compact_exception_message(exc)
    lower_detail = detail.lower()
    error_text = f"{exc.__class__.__name__}: {detail}"

    if isinstance(exc, requests.exceptions.ConnectTimeout):
        return "连接超时，目标可能已下线或端口不通", error_text
    if isinstance(exc, requests.exceptions.ReadTimeout):
        return "读取超时，目标响应过慢或服务异常", error_text
    if isinstance(exc, requests.exceptions.SSLError):
        return "SSL握手失败，HTTPS 配置异常", error_text
    if isinstance(exc, requests.exceptions.TooManyRedirects):
        return "重定向过多，站点可能存在循环跳转", error_text
    if isinstance(exc, requests.exceptions.InvalidURL):
        return "目标地址格式无效", error_text
    if isinstance(exc, requests.exceptions.ConnectionError):
        if any(token in lower_detail for token in ["getaddrinfo failed", "name or service not known", "no such host", "failed to resolve", "nodename nor servname"]):
            return "DNS解析失败，域名可能已失效", error_text
        if any(token in lower_detail for token in ["connection refused", "actively refused", "10061"]):
            return "连接被拒绝，端口未开放或服务未启动", error_text
        if any(token in lower_detail for token in ["network is unreachable", "no route to host"]):
            return "网络不可达，目标可能已下线", error_text
        if any(token in lower_detail for token in ["remote end closed connection", "connection aborted", "connection reset", "reset by peer"]):
            return "连接被中断，目标服务提前断开", error_text
        return "连接失败，目标可能已失效或暂时不可达", error_text
    if isinstance(exc, requests.exceptions.RequestException):
        return "请求异常，目标可能暂时不可用", error_text
    return "分析异常", error_text


RETRYABLE_FAILURE_PREFIXES = (
    "连接超时，目标可能已下线或端口不通",
    "读取超时，目标响应过慢或服务异常",
    "DNS解析失败，域名可能已失效",
    "连接被拒绝，端口未开放或服务未启动",
    "网络不可达，目标可能已下线",
    "连接被中断，目标服务提前断开",
    "连接失败，目标可能已失效或暂时不可达",
    "请求异常，目标可能暂时不可用",
)


def is_retryable_failure_result(risk_level: str, result: str) -> bool:
    if risk_level != "失败":
        return False
    summary = (result or "").strip()
    return any(summary.startswith(prefix) for prefix in RETRYABLE_FAILURE_PREFIXES)


@dataclass
class AuditRecord:
    record_id: int
    target: str
    final_url: str = ""
    proxy_used: str = ""
    ocr_endpoint_used: str = ""
    ocr_route_rule: str = ""
    locator_rule_used: str = ""
    status: str = "等待中"
    title: str = ""
    risk_level: str = "-"
    result: str = "-"
    login_score: int = 0
    login_form: bool = False
    password_field_count: int = 0
    captcha_present: bool = False
    mfa_present: bool = False
    lockout_hint: bool = False
    default_hint: bool = False
    form_action: str = ""
    form_method: str = ""
    field_summary: str = ""
    screenshot_path: str = ""
    error: str = ""

class BruteForceHandler:
    def __init__(
        self,
        session,
        log_queue,
        driver_factory=None,
        render_wait=2.5,
        captcha_ocr_enabled=False,
        ocr_endpoint=DEFAULT_OCR_ENDPOINT,
        ocr_endpoint_resolver=None,
        locator_rule_resolver=None,
        captcha_lock=None,
    ):
        self.session = session
        self.log_queue = log_queue
        self.default_user = DEFAULT_CN_USERNAMES.copy()
        self.default_pass = DEFAULT_CN_PASSWORDS.copy()
        self.driver_factory = driver_factory
        self.render_wait = max(0.8, float(render_wait or 2.5))
        self.captcha_ocr_enabled = bool(captcha_ocr_enabled)
        self.ocr_endpoint = (ocr_endpoint or DEFAULT_OCR_ENDPOINT).strip() or DEFAULT_OCR_ENDPOINT
        self.ocr_endpoint_resolver = ocr_endpoint_resolver
        self.locator_rule_resolver = locator_rule_resolver
        self.captcha_lock = captcha_lock

    def run(
        self,
        record,
        dict_mode=DEFAULT_BRUTE_DICT_MODE,
        user_dict_path="",
        pass_dict_path="",
        soup=None,
    ):
        if not record.login_form:
            return "跳过(非登录页)"
        if not self.has_actionable_login_fields(record):
            return "跳过(未识别到可用登录框)"

        try:
            users, passwords, dict_label = self.load_dicts(dict_mode, user_dict_path, pass_dict_path)
        except ValueError as exc:
            return f"失败({exc})"

        browser_result = ""
        if self.should_use_browser_login(record):
            browser_result = self.run_browser_login(record, users, passwords, dict_label)
            if not browser_result.startswith("回退静态模式"):
                return browser_result

        form = self.find_login_form(soup)
        container = form if form is not None else soup
        if container is None:
            return browser_result.replace("回退静态模式", "失败", 1) if browser_result else "失败(未识别登录表单)"

        form_fields = self.extract_form_fields(container)
        user_key, pass_key = self.identify_login_fields(form_fields)
        if not user_key or not pass_key:
            user_key, pass_key = self.identify_login_fields_from_summary(record.field_summary)
        if not user_key or not pass_key:
            return browser_result.replace("回退静态模式", "失败", 1) if browser_result else "失败(未识别用户名/密码字段)"
        if (
            not form_fields
            and ("browser_user" in (record.field_summary or "") or "browser_password" in (record.field_summary or ""))
            and (not self.driver_factory or "浏览器驱动不可用" in browser_result)
        ):
            return "跳过(前端渲染登录框缺少静态字段名，当前环境无法进行浏览器自动化提交)"

        action_value = ""
        method = ""
        if form is not None:
            action_value = (form.get("action") or "").strip()
            method = (form.get("method") or "").strip().upper()
        action_url = self.get_action_url(record, action_value)
        if not method:
            method = "POST"
        if method not in {"GET", "POST"}:
            method = "POST"

        base_payload = self.build_base_payload(form_fields, user_key, pass_key)
        referer = record.final_url or record.target
        attempt_count = 0
        for u in users:
            for p in passwords:
                try:
                    attempt_count += 1
                    payload = dict(base_payload)
                    payload[user_key] = u
                    payload[pass_key] = p
                    resp = self.submit_login(action_url, method, payload, referer)
                    if self.is_successful_login(resp, action_url):
                        return f"{BRUTE_FORCE_SUCCESS_PREFIX} {u}/{p}"
                except requests.RequestException:
                    continue
        return f"未命中弱口令(已尝试 {attempt_count} 组, {dict_label})"

    def log(self, message: str):
        if self.log_queue is not None:
            self.log_queue.put(message)

    def has_actionable_login_fields(self, record) -> bool:
        summary = (getattr(record, "field_summary", "") or "").lower()
        return bool(getattr(record, "password_field_count", 0) > 0 or "password:" in summary)

    def should_use_browser_login(self, record) -> bool:
        if not self.driver_factory:
            return False
        result_text = record.result or ""
        return (record.captcha_present and self.has_actionable_login_fields(record)) or "浏览器渲染补扫命中" in result_text

    def wait_for_render(self, driver):
        deadline = time.time() + self.render_wait
        while time.time() < deadline:
            try:
                if driver.execute_script("return document.readyState") == "complete":
                    break
            except Exception:
                break
            time.sleep(0.2)
        remaining = max(0.2, deadline - time.time())
        time.sleep(remaining)

    def dismiss_browser_obstructions(self, driver):
        try:
            driver.switch_to.default_content()
        except Exception:
            pass
        try:
            driver.execute_script(
                """
                const candidates = Array.from(document.querySelectorAll('*')).filter((node) => {
                  const style = window.getComputedStyle(node);
                  if (!style) return false;
                  const fixed = style.position === 'fixed' || style.position === 'sticky';
                  const visible = style.display !== 'none' && style.visibility !== 'hidden' && style.opacity !== '0';
                  const zIndex = Number.parseInt(style.zIndex || '0', 10);
                  const rect = node.getBoundingClientRect();
                  const largeEnough = rect.width > 120 && rect.height > 36;
                  return fixed && visible && zIndex >= 1000 && largeEnough;
                });
                for (const node of candidates.slice(0, 12)) {
                  node.style.display = 'none';
                }
                """
            )
        except Exception:
            pass
        try:
            from selenium.webdriver.common.keys import Keys
            body = driver.find_element("tag name", "body")
            body.send_keys(Keys.ESCAPE)
        except Exception:
            pass

    def run_browser_login(self, record, users, passwords, dict_label):
        needs_captcha = bool(record.captcha_present)
        if needs_captcha and not self.captcha_ocr_enabled:
            return "跳过(存在图形验证码且未启用本地OCR)"

        if needs_captcha and self.captcha_lock:
            self.log(f"[*] {record.target} 存在验证码，进入串行弱口令队列。")
            self.captcha_lock.acquire()

        try:
            driver = self.driver_factory() if self.driver_factory else None
            if not driver:
                if needs_captcha:
                    return "失败(浏览器驱动不可用，无法处理验证码)"
                return "回退静态模式(浏览器驱动不可用)"

            try:
                login_url = record.final_url or record.target
                attempt_count = 0
                self._active_record = record
                for username in users:
                    for password in passwords:
                        attempt_count += 1
                        try:
                            driver.get(login_url)
                            self.wait_for_render(driver)
                            self.dismiss_browser_obstructions(driver)
                            dom = self.locate_login_dom(driver)
                            if not dom["user"] or not dom["pass"]:
                                if needs_captcha:
                                    return "失败(浏览器模式未识别用户名/密码输入框)"
                                return "回退静态模式(浏览器模式未识别用户名/密码输入框)"

                            captcha_text = ""
                            if needs_captcha or dom["captcha"]:
                                captcha_text = self.solve_captcha_from_dom(driver, dom, record)
                                if not captcha_text:
                                    if needs_captcha:
                                        continue
                                    return "回退静态模式(浏览器模式未识别验证码)"

                            self.fill_element(dom["user"], username)
                            self.fill_element(dom["pass"], password)
                            if dom["captcha"] and captcha_text:
                                self.fill_element(dom["captcha"], captcha_text)

                            submitted = self.submit_dom_form(driver, dom["submit"], dom["pass"])
                            if not submitted:
                                if needs_captcha:
                                    return "失败(浏览器模式无法提交登录表单)"
                                return "回退静态模式(浏览器模式无法提交登录表单)"

                            time.sleep(1.0)
                            if self.is_successful_browser_login(driver, login_url):
                                return f"{BRUTE_FORCE_SUCCESS_PREFIX} {username}/{password}"
                        except Exception:
                            continue
                return f"未命中弱口令(已尝试 {attempt_count} 组, {dict_label}, 浏览器模式)"
            finally:
                self._active_record = None
                try:
                    driver.quit()
                except Exception:
                    pass
        finally:
            if needs_captcha and self.captcha_lock:
                self.captcha_lock.release()

    def locate_login_dom(self, driver):
        try:
            from selenium.webdriver.common.by import By
        except Exception:
            return {"form": None, "user": None, "pass": None, "captcha": None, "submit": None}

        try:
            driver.switch_to.default_content()
        except Exception:
            pass
        rule = None
        if self.locator_rule_resolver is not None:
            try:
                rule = self.locator_rule_resolver(getattr(self, "_active_record", None))
            except Exception:
                rule = None
        dom = self.locate_login_dom_recursive(driver, depth=0, max_depth=3)
        if rule:
            ruled_dom = self.locate_login_dom_by_rule(driver, rule)
            if ruled_dom and ruled_dom.get("pass") is not None:
                return ruled_dom
        return dom or {"form": None, "user": None, "pass": None, "captcha": None, "submit": None}

    def locate_login_dom_recursive(self, driver, depth: int, max_depth: int):
        dom = self.locate_login_dom_in_current_context(driver)
        if dom and dom.get("pass") is not None:
            return dom
        if depth >= max_depth:
            return None

        try:
            from selenium.webdriver.common.by import By
        except Exception:
            return None

        for frame in driver.find_elements(By.CSS_SELECTOR, "iframe,frame"):
            try:
                driver.switch_to.frame(frame)
                nested = self.locate_login_dom_recursive(driver, depth + 1, max_depth)
                if nested and nested.get("pass") is not None:
                    return nested
            except Exception:
                pass
            finally:
                try:
                    driver.switch_to.parent_frame()
                except Exception:
                    try:
                        driver.switch_to.default_content()
                    except Exception:
                        pass
        return None

    def locate_login_dom_in_current_context(self, driver):
        try:
            from selenium.webdriver.common.by import By
        except Exception:
            return None

        form = None
        forms = driver.find_elements(By.TAG_NAME, "form")
        for candidate in forms:
            if candidate.find_elements(By.XPATH, ".//input[contains(translate(@type,'PASSWORD','password'),'password')]"):
                form = candidate
                break
        input_xpath = ".//input|.//textarea" if form else "//input|//textarea"
        button_xpath = ".//button|.//input[@type='submit']|.//input[@type='button']" if form else "//button|//input[@type='submit']|//input[@type='button']"
        search_root = form if form else driver
        inputs = search_root.find_elements(By.XPATH, input_xpath)
        buttons = search_root.find_elements(By.XPATH, button_xpath)

        user_el = None
        pass_el = None
        captcha_el = None

        for element in inputs:
            if not self.is_input_candidate_visible(element):
                continue
            hints = self.collect_dom_hints(element)
            field_type = self.safe_element_attr(element, "type").lower()
            if not pass_el and ("password" in field_type or any(keyword in hints for keyword in PASSWORD_FIELD_KEYWORDS)):
                pass_el = element
                continue

            if field_type in {"hidden", "submit", "button", "checkbox", "radio", "file"}:
                continue
            if not captcha_el and any(keyword in hints for keyword in CAPTCHA_FIELD_KEYWORDS):
                captcha_el = element
                continue
            if not user_el and any(keyword in hints for keyword in USER_FIELD_KEYWORDS):
                user_el = element

        if not user_el:
            for element in inputs:
                if element == pass_el or element == captcha_el or not self.is_input_candidate_visible(element):
                    continue
                field_type = self.safe_element_attr(element, "type").lower()
                if field_type in {"text", "email", "tel", "number", "search", ""}:
                    user_el = element
                    break

        submit_el = None
        for button in buttons:
            if self.is_input_candidate_visible(button) and self.is_submit_candidate(button):
                submit_el = button
                break
        if submit_el is None:
            custom_xpath = (
                ".//*[self::a or self::div or self::span or name()='uni-button']"
                if form
                else "//*[self::a or self::div or self::span or name()='uni-button']"
            )
            try:
                for button in search_root.find_elements(By.XPATH, custom_xpath):
                    if self.is_input_candidate_visible(button) and self.is_submit_candidate(button):
                        submit_el = button
                        break
            except Exception:
                submit_el = None
        if submit_el is None:
            for button in buttons:
                if self.is_input_candidate_visible(button):
                    submit_el = button
                    break
        return {"form": form, "user": user_el, "pass": pass_el, "captcha": captcha_el, "submit": submit_el}

    def locate_login_dom_by_rule(self, driver, rule: dict):
        selectors = rule.get("selectors") or {}
        try:
            driver.switch_to.default_content()
        except Exception:
            pass

        frame_spec = selectors.get("frame", "")
        if frame_spec:
            frame_element = self.find_element_by_locator_spec(driver, frame_spec)
            if frame_element is not None:
                try:
                    driver.switch_to.frame(frame_element)
                except Exception:
                    try:
                        driver.switch_to.default_content()
                    except Exception:
                        pass

        user_el = self.find_element_by_locator_spec(driver, selectors.get("user", ""))
        pass_el = self.find_element_by_locator_spec(driver, selectors.get("pass", ""))
        submit_el = self.find_element_by_locator_spec(driver, selectors.get("submit", ""))
        captcha_el = self.find_element_by_locator_spec(driver, selectors.get("captcha", ""))
        if pass_el is None:
            return None
        active_record = getattr(self, "_active_record", None)
        if active_record is not None:
            active_record.locator_rule_used = rule.get("rule_text", "")
        return {"form": None, "user": user_el, "pass": pass_el, "captcha": captcha_el, "submit": submit_el}

    def find_element_by_locator_spec(self, driver, spec: str):
        if not spec:
            return None
        try:
            from selenium.webdriver.common.by import By
        except Exception:
            return None

        segments = [segment.strip() for segment in spec.split(",") if segment.strip()]
        for segment in segments:
            locate_by = By.CSS_SELECTOR
            value = segment
            lowered = segment.lower()
            if lowered.startswith("css:"):
                locate_by = By.CSS_SELECTOR
                value = segment[4:].strip()
            elif lowered.startswith("xpath:"):
                locate_by = By.XPATH
                value = segment[6:].strip()
            if not value:
                continue
            try:
                elements = driver.find_elements(locate_by, value)
            except Exception:
                continue
            for element in elements:
                if self.is_input_candidate_visible(element):
                    return element
        return None

    def is_input_candidate_visible(self, element) -> bool:
        try:
            if not element.is_displayed():
                return False
            size = element.size or {}
            return size.get("width", 0) > 8 and size.get("height", 0) > 8
        except Exception:
            return False

    def safe_element_attr(self, element, attr_name):
        try:
            return (element.get_attribute(attr_name) or "").strip()
        except Exception:
            return ""

    def safe_element_text(self, element):
        try:
            text = (element.text or "").strip()
            if text:
                return text
        except Exception:
            pass
        for attr_name in ("innerText", "textContent", "value"):
            value = self.safe_element_attr(element, attr_name)
            if value:
                return value
        return ""

    def collect_dom_hints(self, element):
        parts = [
            self.safe_element_attr(element, "name"),
            self.safe_element_attr(element, "id"),
            self.safe_element_attr(element, "placeholder"),
            self.safe_element_attr(element, "aria-label"),
            self.safe_element_attr(element, "autocomplete"),
            self.safe_element_attr(element, "title"),
            self.safe_element_attr(element, "data-placeholder"),
            self.safe_element_attr(element, "data-label"),
            self.safe_element_attr(element, "formcontrolname"),
            self.safe_element_attr(element, "ng-model"),
            self.safe_element_attr(element, "role"),
            self.safe_element_attr(element, "class"),
            self.safe_element_attr(element, "type"),
            self.safe_element_text(element),
        ]
        return " ".join(part for part in parts if part).lower()

    def is_submit_candidate(self, element) -> bool:
        hints = self.collect_dom_hints(element)
        return any(keyword.lower() in hints for keyword in LOGIN_BUTTON_KEYWORDS) or any(
            token in hints for token in ("button", "btn", "submit", "login")
        )

    def solve_captcha_from_dom(self, driver, dom, record):
        captcha_input = dom.get("captcha")
        captcha_image = self.find_captcha_image(driver, dom.get("form"), captcha_input)
        if captcha_image is None:
            return ""
        try:
            image_b64 = captcha_image.screenshot_as_base64
        except Exception:
            image_b64 = ""
        if not image_b64:
            return ""
        return self.request_local_ocr(image_b64, record)

    def find_captcha_image(self, driver, form, captcha_input):
        try:
            from selenium.webdriver.common.by import By
        except Exception:
            return None

        roots = [form] if form else [driver]
        xpaths = [
            ".//img[contains(translate(@src,'CAPTCHA','captcha'),'captcha') or contains(translate(@id,'CAPTCHA','captcha'),'captcha') or contains(translate(@class,'CAPTCHA','captcha'),'captcha') or contains(translate(@name,'CAPTCHA','captcha'),'captcha')]",
            ".//canvas",
            ".//img",
        ]

        if captcha_input is not None:
            try:
                nearby = captcha_input.find_elements(By.XPATH, "./following::*[self::img or self::canvas][1]")
                if nearby:
                    return nearby[0]
            except Exception:
                pass

        for root in roots:
            for xpath in xpaths:
                current_xpath = xpath if form else xpath.replace(".//", "//", 1)
                try:
                    elements = root.find_elements(By.XPATH, current_xpath)
                except Exception:
                    continue
                for element in elements:
                    try:
                        if element.is_displayed() and element.size.get("width", 0) >= 20 and element.size.get("height", 0) >= 12:
                            return element
                    except Exception:
                        continue
        return None

    def request_local_ocr(self, image_b64: str, record=None) -> str:
        endpoint = self.ocr_endpoint
        route_rule = f"default => {endpoint}"
        if self.ocr_endpoint_resolver is not None:
            try:
                resolved = self.ocr_endpoint_resolver(record)
                if isinstance(resolved, (tuple, list)) and len(resolved) >= 2:
                    endpoint = str(resolved[0] or endpoint)
                    route_rule = str(resolved[1] or route_rule)
                elif resolved:
                    endpoint = str(resolved)
                    route_rule = f"default => {endpoint}"
            except Exception:
                endpoint = self.ocr_endpoint
                route_rule = f"default => {endpoint}"
        if record is not None:
            record.ocr_endpoint_used = endpoint
            record.ocr_route_rule = route_rule
        payload = f"<@BASE64>{image_b64}</@BASE64>"
        response = requests.post(
            endpoint,
            data=payload.encode("utf-8"),
            headers={"Content-Type": "text/plain"},
            timeout=8,
        )
        response.raise_for_status()
        return self.parse_ocr_response(response.text)

    def parse_ocr_response(self, body: str) -> str:
        text = (body or "").strip()
        if not text:
            return ""
        if text.startswith("{") and text.endswith("}"):
            try:
                payload = json.loads(text)
            except Exception:
                payload = None
            if isinstance(payload, dict):
                for key in ("return", "result", "data", "text", "code_text"):
                    value = payload.get(key)
                    if isinstance(value, str) and value.strip():
                        return value.strip()
        return text.strip().strip('"').strip("'")

    def fill_element(self, element, value: str):
        try:
            element.clear()
        except Exception:
            pass
        try:
            element.send_keys(value)
        except Exception:
            pass

    def submit_dom_form(self, driver, submit_element, pass_element):
        try:
            if submit_element is not None:
                submit_element.click()
                return True
        except Exception:
            pass
        try:
            if submit_element is not None:
                driver.execute_script("arguments[0].click();", submit_element)
                return True
        except Exception:
            pass
        try:
            pass_element.submit()
            return True
        except Exception:
            pass
        try:
            from selenium.webdriver.common.keys import Keys
            pass_element.send_keys(Keys.RETURN)
            return True
        except Exception:
            pass
        try:
            return bool(driver.execute_script(
                """
                const forms = document.getElementsByTagName('form');
                if (forms.length) { forms[0].submit(); return true; }
                return false;
                """
            ))
        except Exception:
            return False

    def is_successful_browser_login(self, driver, login_url: str) -> bool:
        try:
            driver.switch_to.default_content()
        except Exception:
            pass
        current_url = (driver.current_url or "").lower()
        if current_url and current_url != (login_url or "").lower() and not self.contains_login_marker(current_url):
            return True

        body = (driver.page_source or "").lower()
        if any(keyword in body for keyword in LOGIN_FAILURE_KEYWORDS):
            return False
        if any(keyword in body for keyword in LOGIN_SUCCESS_KEYWORDS) and not self.contains_login_marker(body):
            return True

        dom = self.locate_login_dom(driver)
        if dom["pass"] is None and not self.contains_login_marker(body):
            return True
        return False

    def load_dicts(self, dict_mode, user_dict_path="", pass_dict_path=""):
        if dict_mode == CUSTOM_BRUTE_DICT_MODE:
            users = self.read_dict_file(user_dict_path, "账号")
            passwords = self.read_dict_file(pass_dict_path, "密码")
            return users, passwords, "自定义字典"
        return self.default_user.copy(), self.default_pass.copy(), "默认中文常用字典"

    def read_dict_file(self, path, label):
        value = str(path or "").strip()
        if not value:
            raise ValueError(f"未选择{label}字典")

        file_path = Path(value).expanduser()
        if not file_path.exists():
            raise ValueError(f"{label}字典不存在")

        last_error = None
        for encoding in DICT_FILE_ENCODINGS:
            try:
                with open(file_path, "r", encoding=encoding) as handle:
                    raw_lines = handle.readlines()
                break
            except UnicodeError as exc:
                last_error = exc
        else:
            raise ValueError(f"{label}字典编码无法识别: {last_error}")

        entries = []
        seen = set()
        for raw_line in raw_lines:
            item = raw_line.strip()
            if not item or item.startswith(("#", ";", "//")):
                continue
            if item in seen:
                continue
            seen.add(item)
            entries.append(item)

        if not entries:
            raise ValueError(f"{label}字典为空")
        return entries

    def find_login_form(self, soup):
        if soup is None:
            return None
        for candidate in soup.find_all("form"):
            if candidate.find("input", {"type": re.compile("password", re.I)}):
                return candidate
        return soup.find("form")

    def extract_form_fields(self, container):
        fields = []
        for tag in container.find_all(["input", "select", "textarea", "button"]):
            field_type = (tag.get("type") or tag.name or "").strip().lower()
            submit_name = (tag.get("name") or tag.get("id") or "").strip()
            if not submit_name:
                continue

            value = (tag.get("value") or "").strip()
            if tag.name == "textarea":
                value = tag.get_text(strip=True)
            elif tag.name == "select":
                selected = tag.find("option", selected=True) or tag.find("option")
                if selected is not None:
                    value = (selected.get("value") or selected.get_text(strip=True) or "").strip()

            hints = " ".join(
                filter(
                    None,
                    [
                        tag.get("name"),
                        tag.get("id"),
                        tag.get("placeholder"),
                        tag.get("aria-label"),
                        tag.get("autocomplete"),
                        tag.get("title"),
                        tag.get("data-placeholder"),
                        tag.get("data-label"),
                        tag.get("formcontrolname"),
                        tag.get("ng-model"),
                        tag.get("class") and " ".join(tag.get("class")),
                    ],
                )
            ).lower()
            fields.append(
                {
                    "name": submit_name,
                    "type": field_type or tag.name,
                    "value": value,
                    "checked": bool(tag.has_attr("checked")),
                    "hints": hints,
                }
            )
        return fields

    def identify_login_fields(self, form_fields):
        user_key = ""
        pass_key = ""
        visible_candidates = []

        for field in form_fields:
            name = field["name"]
            field_type = field["type"]
            hints = field["hints"]

            if not pass_key and ("password" in field_type or any(keyword in hints for keyword in PASSWORD_FIELD_KEYWORDS)):
                pass_key = name
                continue

            if field_type in {"hidden", "submit", "button", "checkbox", "radio", "file"}:
                continue
            if any(keyword in hints for keyword in SKIP_USERNAME_FIELD_KEYWORDS):
                continue

            visible_candidates.append(name)
            if not user_key and any(keyword in hints for keyword in USER_FIELD_KEYWORDS):
                user_key = name

        if not user_key:
            for field in form_fields:
                if field["name"] == pass_key:
                    continue
                if field["type"] in {"text", "email", "tel", "number", "search", "textarea"}:
                    if any(keyword in field["hints"] for keyword in SKIP_USERNAME_FIELD_KEYWORDS):
                        continue
                    user_key = field["name"]
                    break

        if not user_key and len(visible_candidates) == 1:
            user_key = visible_candidates[0]
        return user_key, pass_key

    def identify_login_fields_from_summary(self, field_summary):
        user_key = ""
        pass_key = ""
        for field in (field_summary or "").split(" | "):
            if ":" not in field:
                continue
            field_type, field_name = field.split(":", 1)
            field_name = field_name.strip()
            lower_name = field_name.lower()
            if not user_key and any(keyword in lower_name for keyword in USER_FIELD_KEYWORDS):
                user_key = field_name
            if not pass_key and (
                "password" in field_type.lower()
                or any(keyword in lower_name for keyword in PASSWORD_FIELD_KEYWORDS)
            ):
                pass_key = field_name
        return user_key, pass_key

    def build_base_payload(self, form_fields, user_key, pass_key):
        payload = {}
        for field in form_fields:
            name = field["name"]
            field_type = field["type"]
            if not name or name in {user_key, pass_key}:
                continue

            if field_type == "hidden":
                payload[name] = field["value"]
            elif field_type in {"checkbox", "radio"} and field["checked"]:
                payload[name] = field["value"] or "on"
            elif field_type in {"submit", "button"} and field["value"] and name not in payload:
                payload[name] = field["value"]
            elif field_type == "select" and field["value"]:
                payload[name] = field["value"]
            elif field_type == "textarea" and field["value"]:
                payload[name] = field["value"]
        return payload

    def submit_login(self, action_url, method, payload, referer):
        headers = {"Referer": referer}
        if method == "GET":
            return self.session.get(
                action_url,
                params=payload,
                headers=headers,
                timeout=5,
                verify=False,
                allow_redirects=False,
            )
        return self.session.post(
            action_url,
            data=payload,
            headers=headers,
            timeout=5,
            verify=False,
            allow_redirects=False,
        )

    def is_successful_login(self, response, action_url):
        location = response.headers.get("Location", "")
        if response.status_code in {301, 302, 303, 307, 308} and location:
            redirect_target = urljoin(action_url, location).lower()
            if not self.contains_login_marker(redirect_target):
                return True

        body = (response.text or "").lower()
        if any(keyword in body for keyword in LOGIN_FAILURE_KEYWORDS):
            return False
        if any(keyword in body for keyword in LOGIN_SUCCESS_KEYWORDS) and not self.contains_login_marker(body):
            return True
        return False

    def contains_login_marker(self, text):
        lowered = (text or "").lower()
        return any(keyword.lower() in lowered for keyword in LOGIN_KEYWORDS)

    def get_action_url(self, record, form_action=""):
        action = form_action or record.form_action or ""
        if action.startswith("http"):
            return action
        return urljoin(record.final_url or record.target, action or record.final_url or record.target)


class SecurityAuditGUI:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("本地安全资产登录面审计工具 v2.0")
        self.root.geometry("1520x940")
        self.root.minsize(1080, 720)

        self.log_queue = queue.Queue()
        self.result_lock = threading.Lock()
        self.records_by_item = {}
        self.item_id_by_record_key = {}
        self.all_records = []
        self.is_scanning = False
        self.output_dir = Path.cwd() / "output"
        self.output_dir.mkdir(exist_ok=True)
        self.state_dir = self.output_dir / "state"
        self.state_dir.mkdir(exist_ok=True)
        self.dicts_dir = self.state_dir / "dicts"
        self.dicts_dir.mkdir(exist_ok=True)
        self.projects_dir = self.state_dir / "projects"
        self.projects_dir.mkdir(exist_ok=True)
        self.builtin_user_dict_path = self.dicts_dir / "builtin_usernames.txt"
        self.builtin_pass_dict_path = self.dicts_dir / "builtin_passwords.txt"
        self.latest_project_path = self.state_dir / "latest_project.json"
        self.current_project_path = None
        self.current_run_dir = None
        self.detail_log_path = None
        self.detail_log_lock = threading.Lock()
        self.autosave_delay_ms = 900
        self.pending_snapshot_after_id = None
        self.pending_snapshot_reason = ""
        self.source_watch_interval_ms = 2000
        self.source_watch_signature = None
        self.source_change_prompt_active = False
        self.source_change_pending = False
        self.last_loaded_path = ""
        self.preview_image = None
        self.brute_var = tk.BooleanVar(value=False)
        self.browser_render_var = tk.BooleanVar(value=True)
        self.captcha_ocr_var = tk.BooleanVar(value=True)
        self.captcha_brute_lock = threading.Semaphore(1)
        self.proxy_assignment = {}
        self.proxy_round_robin_index = 0
        self.proxy_health = {}
        self.ensure_builtin_dict_files()
        self.setup_theme()
        self.setup_ui()
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)
        self.check_log_queue()
        self.load_autosave_if_exists()
        self.watch_source_file_changes()

    def setup_theme(self):
        self.palette = {
            "bg": "#eef4f8",
            "panel": "#ffffff",
            "header": "#dce8f3",
            "border": "#bfd0df",
            "text": "#0f172a",
            "muted": "#526277",
            "muted_bg": "#e7edf4",
            "accent": "#2563eb",
            "accent_active": "#1d4ed8",
            "accent_soft": "#dbeafe",
            "success": "#0f766e",
            "success_bg": "#d9f7f1",
            "warning": "#a16207",
            "warning_bg": "#fde68a",
            "danger": "#b91c1c",
            "danger_bg": "#fecaca",
        }
        self.font_normal = ("Microsoft YaHei UI", 9)
        self.font_section = ("Microsoft YaHei UI", 10, "bold")
        self.font_title = ("Microsoft YaHei UI", 11, "bold")
        self.root.configure(bg=self.palette["bg"])
        self.root.option_add("*Font", self.font_normal)
        self.root.option_add("*TCombobox*Listbox*Background", self.palette["panel"])
        self.root.option_add("*TCombobox*Listbox*Foreground", self.palette["text"])
        self.root.option_add("*TCombobox*Listbox*selectBackground", self.palette["accent_soft"])
        self.root.option_add("*TCombobox*Listbox*selectForeground", self.palette["text"])

        self.style = ttk.Style()
        try:
            self.style.theme_use("clam")
        except tk.TclError:
            pass

        self.style.configure(".", font=self.font_normal)
        self.style.configure("TCombobox", padding=4, foreground=self.palette["text"], fieldbackground=self.palette["panel"], background=self.palette["panel"], arrowsize=13)
        self.style.map(
            "TCombobox",
            fieldbackground=[("readonly", self.palette["panel"]), ("disabled", self.palette["muted_bg"])],
            foreground=[("readonly", self.palette["text"]), ("disabled", self.palette["muted"])],
            selectbackground=[("readonly", self.palette["accent_soft"])],
            selectforeground=[("readonly", self.palette["text"])],
        )
        self.style.configure(
            "Treeview",
            background=self.palette["panel"],
            fieldbackground=self.palette["panel"],
            foreground=self.palette["text"],
            rowheight=30,
            bordercolor="#8ea3b7",
            lightcolor="#8ea3b7",
            darkcolor="#8ea3b7",
            borderwidth=1,
            relief="solid",
        )
        self.style.map("Treeview", background=[("selected", self.palette["accent_soft"])], foreground=[("selected", self.palette["text"])])
        self.style.configure(
            "Treeview.Heading",
            background=self.palette["header"],
            foreground=self.palette["text"],
            relief="solid",
            borderwidth=1,
            padding=(10, 8),
        )
        self.style.map("Treeview.Heading", background=[("active", self.palette["accent_soft"])])
        self.style.configure("TNotebook", background=self.palette["bg"], borderwidth=0)
        self.style.configure("TNotebook.Tab", padding=(12, 6), background=self.palette["header"], foreground=self.palette["muted"])
        self.style.map(
            "TNotebook.Tab",
            background=[("selected", self.palette["panel"]), ("active", self.palette["accent_soft"])],
            foreground=[("selected", self.palette["text"]), ("active", self.palette["text"])],
        )
        self.style.configure("TPanedwindow", background=self.palette["bg"])
        self.style.configure("TScrollbar", background=self.palette["header"], troughcolor=self.palette["bg"])

    def make_section(self, parent, title, padx=12, pady=10):
        return tk.LabelFrame(
            parent,
            text=title,
            padx=padx,
            pady=pady,
            bg=self.palette["panel"],
            fg=self.palette["text"],
            bd=1,
            relief=tk.GROOVE,
            font=self.font_section,
        )

    def make_label(self, parent, muted=False, **kwargs):
        kwargs.setdefault("bg", parent.cget("bg"))
        kwargs.setdefault("fg", self.palette["muted"] if muted else self.palette["text"])
        kwargs.setdefault("font", self.font_normal)
        return tk.Label(parent, **kwargs)

    def make_checkbutton(self, parent, **kwargs):
        kwargs.setdefault("bg", parent.cget("bg"))
        kwargs.setdefault("activebackground", parent.cget("bg"))
        kwargs.setdefault("fg", self.palette["text"])
        kwargs.setdefault("selectcolor", self.palette["panel"])
        kwargs.setdefault("font", self.font_normal)
        kwargs.setdefault("anchor", "w")
        return tk.Checkbutton(parent, **kwargs)

    def make_entry(self, parent, width=20, **kwargs):
        kwargs.setdefault("width", width)
        kwargs.setdefault("relief", tk.SOLID)
        kwargs.setdefault("bd", 1)
        kwargs.setdefault("highlightthickness", 1)
        kwargs.setdefault("highlightbackground", self.palette["border"])
        kwargs.setdefault("highlightcolor", self.palette["accent"])
        kwargs.setdefault("bg", self.palette["panel"])
        kwargs.setdefault("fg", self.palette["text"])
        kwargs.setdefault("insertbackground", self.palette["text"])
        kwargs.setdefault("disabledbackground", self.palette["muted_bg"])
        kwargs.setdefault("disabledforeground", self.palette["muted"])
        kwargs.setdefault("font", self.font_normal)
        return tk.Entry(parent, **kwargs)

    def make_button(self, parent, text, command, tone="neutral", **kwargs):
        tones = {
            "neutral": {"bg": self.palette["header"], "active": self.palette["accent_soft"], "fg": self.palette["text"]},
            "primary": {"bg": self.palette["accent"], "active": self.palette["accent_active"], "fg": "#ffffff"},
            "warning": {"bg": self.palette["warning_bg"], "active": "#fcd34d", "fg": self.palette["warning"]},
            "danger": {"bg": self.palette["danger_bg"], "active": "#fca5a5", "fg": self.palette["danger"]},
            "success": {"bg": self.palette["success_bg"], "active": "#99f6e4", "fg": self.palette["success"]},
        }
        colors = tones.get(tone, tones["neutral"])
        kwargs.setdefault("relief", tk.FLAT)
        kwargs.setdefault("bd", 0)
        kwargs.setdefault("highlightthickness", 0)
        kwargs.setdefault("padx", 10)
        kwargs.setdefault("pady", 5)
        kwargs.setdefault("cursor", "hand2")
        kwargs.setdefault("font", self.font_normal)
        return tk.Button(
            parent,
            text=text,
            command=command,
            bg=colors["bg"],
            fg=colors["fg"],
            activebackground=colors["active"],
            activeforeground=colors["fg"],
            **kwargs,
        )

    def configure_tree_columns(self, tree, columns):
        for spec in columns:
            name, width, anchor = spec[:3]
            minwidth = spec[3] if len(spec) >= 4 else max(36, min(width, int(width * 0.3)))
            stretch = spec[4] if len(spec) >= 5 else True
            tree.heading(name, text=name)
            tree.column(name, width=width, anchor=anchor, minwidth=max(24, int(minwidth)), stretch=stretch)
        tree.configure(style="Treeview")

    def schedule_progress_snapshot(self, reason: str = "autosave", delay_ms: int | None = None):
        self.pending_snapshot_reason = reason or self.pending_snapshot_reason or "autosave"
        wait_ms = self.autosave_delay_ms if delay_ms is None else max(0, int(delay_ms))

        def schedule():
            if self.pending_snapshot_after_id is not None:
                return
            self.pending_snapshot_after_id = self.root.after(wait_ms, self._flush_scheduled_progress_snapshot)

        self.root.after(0, schedule)

    def _flush_scheduled_progress_snapshot(self):
        self.pending_snapshot_after_id = None
        reason = self.pending_snapshot_reason or "autosave"
        self.pending_snapshot_reason = ""
        self.save_progress_snapshot(reason)

    def ensure_builtin_dict_files(self):
        self.write_builtin_dict_file(
            self.builtin_user_dict_path,
            "内置账号字典",
            DEFAULT_CN_USERNAMES,
        )
        self.write_builtin_dict_file(
            self.builtin_pass_dict_path,
            "内置密码字典",
            DEFAULT_CN_PASSWORDS,
        )

    def write_builtin_dict_file(self, path: Path, title: str, entries: list[str]):
        lines = [
            f"# {title}",
            f"# 自动生成，程序启动时会按内置默认值更新",
            "",
            *entries,
            "",
        ]
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", encoding="utf-8") as handle:
            handle.write("\n".join(lines))

    def get_source_file_signature(self, filepath: str | Path | None):
        if not filepath:
            return None
        try:
            path = Path(filepath)
            if not path.exists():
                return ("missing", str(path))
            stat = path.stat()
            return (str(path.resolve()), stat.st_mtime_ns, stat.st_size)
        except Exception:
            return None

    def mark_source_watch_baseline(self, filepath: str | Path | None = None):
        target = filepath or self.last_loaded_path or self.file_entry.get().strip()
        self.source_watch_signature = self.get_source_file_signature(target)
        self.source_change_pending = False

    def prompt_reload_for_changed_source(self):
        if self.source_change_prompt_active:
            return
        filepath = self.last_loaded_path or self.file_entry.get().strip()
        if not filepath:
            self.source_change_pending = False
            return

        self.source_change_prompt_active = True
        try:
            answer = messagebox.askyesno(
                "源文件已变化",
                f"检测到源文件内容已变化：\n{filepath}\n\n是否立即刷新目标列表？",
            )
            if answer:
                self.reload_current_source_file()
            else:
                self.mark_source_watch_baseline(filepath)
                self.log_message(f"[~] 检测到源文件变化，已跳过自动刷新: {filepath}")
        finally:
            self.source_change_prompt_active = False

    def watch_source_file_changes(self):
        try:
            filepath = self.last_loaded_path or (self.file_entry.get().strip() if hasattr(self, "file_entry") else "")
            if filepath:
                current_signature = self.get_source_file_signature(filepath)
                if self.source_watch_signature is None:
                    self.source_watch_signature = current_signature
                elif current_signature != self.source_watch_signature:
                    if self.is_scanning:
                        if not self.source_change_pending:
                            self.source_change_pending = True
                            self.log_message(f"[~] 检测到源文件变化，当前正在扫描，停止后会提示刷新: {filepath}")
                    else:
                        self.source_change_pending = True

                if self.source_change_pending and not self.is_scanning:
                    self.prompt_reload_for_changed_source()
        finally:
            self.root.after(self.source_watch_interval_ms, self.watch_source_file_changes)

    def build_browser_env(self) -> dict:
        env = os.environ.copy()
        for key in ("http_proxy", "https_proxy", "HTTP_PROXY", "HTTPS_PROXY", "ALL_PROXY", "all_proxy"):
            env.pop(key, None)
        env["CHROME_LOG_FILE"] = os.devnull
        return env

    def browser_probe_has_login_signal(self, probe: dict | None) -> bool:
        if not isinstance(probe, dict):
            return False
        field_summary = (probe.get("field_summary") or "").lower()
        return bool(probe.get("login_form") or "password:" in field_summary)

    def is_actionable_login_record(self, record: AuditRecord | None) -> bool:
        if record is None:
            return False
        field_summary = (record.field_summary or "").lower()
        return bool(record.password_field_count > 0 or "password:" in field_summary)

    def detail_log_path_for_project(self, project_path: Path | None = None) -> Path:
        path = project_path or self.current_project_path
        if path:
            return Path(path).with_suffix(".detail.log")
        return self.state_dir / "current_scan.detail.log"

    def set_detail_log_path(self, path: Path | None):
        self.detail_log_path = Path(path) if path else None
        if hasattr(self, "detail_log_path_var"):
            display = str(self.detail_log_path) if self.detail_log_path else "尚未创建"
            self.root.after(0, self.detail_log_path_var.set, f"详细日志: {display}")

    def write_detail_log(self, message: str, level: str = "INFO", record: AuditRecord | None = None):
        if not self.detail_log_path:
            return
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        parts = [f"[{timestamp}]", f"[{level}]"]
        if record is not None:
            parts.append(f"[{record.target}]")
        line = "".join(parts) + f" {message}".rstrip() + "\n"
        with self.detail_log_lock:
            self.detail_log_path.parent.mkdir(parents=True, exist_ok=True)
            with open(self.detail_log_path, "a", encoding="utf-8") as handle:
                handle.write(line)

    def start_detail_log_session(self, run_dir: Path):
        self.current_run_dir = run_dir
        detail_path = self.detail_log_path_for_project()
        self.set_detail_log_path(detail_path)
        self.write_detail_log("=" * 72, level="SESSION")
        self.write_detail_log(f"扫描开始 | run_dir={run_dir}", level="SESSION")
        if self.last_loaded_path:
            self.write_detail_log(f"源文件={self.last_loaded_path}", level="SESSION")
        self.write_detail_log(
            "配置 | "
            f"模式={self.mode_var.get()} | 并发={self.get_worker_count()} | 浏览器补扫={'启用' if self.browser_render_var.get() else '关闭'} | "
            f"弱口令={'启用' if self.brute_var.get() else '关闭'} | 验证码OCR={'启用' if self.captcha_ocr_var.get() else '关闭'}",
            level="SESSION",
        )

    def open_detail_log_file(self):
        if not self.detail_log_path or not self.detail_log_path.exists():
            messagebox.showinfo("提示", "当前还没有详细日志文件。")
            return
        os.startfile(self.detail_log_path)

    def clear_ui_log(self):
        if not hasattr(self, "log_text"):
            return
        self.log_text.config(state=tk.NORMAL)
        self.log_text.delete("1.0", tk.END)
        self.log_text.config(state=tk.DISABLED)

    def build_common_login_urls(self, url: str) -> list[str]:
        parsed = urlparse(url or "")
        if not parsed.scheme or not parsed.netloc:
            return []
        origin = f"{parsed.scheme}://{parsed.netloc}"
        candidates = []
        seen = set()
        for suffix in COMMON_LOGIN_ROUTE_SUFFIXES:
            if suffix.startswith("/#"):
                candidate = origin + suffix
            else:
                candidate = urljoin(origin + "/", suffix.lstrip("/"))
            normalized = candidate.rstrip("/")
            if normalized in seen or normalized == (url or "").rstrip("/"):
                continue
            seen.add(normalized)
            candidates.append(candidate)
        return candidates

    def should_try_common_login_routes(self, probe: dict | None) -> bool:
        if not isinstance(probe, dict) or self.browser_probe_has_login_signal(probe):
            return False
        html_text = (probe.get("html") or "").lower()
        title = (probe.get("title") or "").lower()
        input_count = int(probe.get("input_count") or 0)
        return (
            any(token in html_text for token in BROWSER_RENDER_HINTS)
            or any(keyword.lower() in title or keyword.lower() in html_text for keyword in LOGIN_KEYWORDS)
            or input_count > 0
        )

    def rank_browser_probe(self, probe: dict | None) -> int:
        if not isinstance(probe, dict):
            return -1
        if self.browser_probe_has_login_signal(probe):
            return 4
        field_summary = (probe.get("field_summary") or "").lower()
        title = (probe.get("title") or "").lower()
        html_text = (probe.get("html") or "").lower()
        input_count = int(probe.get("input_count") or 0)
        if "browser_user" in field_summary or input_count >= 2:
            return 3
        if any(keyword.lower() in title or keyword.lower() in html_text for keyword in LOGIN_KEYWORDS):
            return 2
        if probe.get("html"):
            return 1
        return 0

    def log_record_trace(self, record: AuditRecord | None, message: str, level: str = "TRACE"):
        self.write_detail_log(message, level=level, record=record)

    def find_browser_executable(self) -> str:
        candidates = [
            shutil.which("chrome.exe"),
            shutil.which("chrome"),
            shutil.which("msedge.exe"),
            shutil.which("msedge"),
            Path(os.environ.get("ProgramFiles", r"C:\Program Files")) / "Google/Chrome/Application/chrome.exe",
            Path(os.environ.get("ProgramFiles(x86)", r"C:\Program Files (x86)")) / "Google/Chrome/Application/chrome.exe",
            Path(os.environ.get("ProgramFiles", r"C:\Program Files")) / "Microsoft/Edge/Application/msedge.exe",
            Path(os.environ.get("ProgramFiles(x86)", r"C:\Program Files (x86)")) / "Microsoft/Edge/Application/msedge.exe",
        ]
        seen = set()
        for candidate in candidates:
            if not candidate:
                continue
            path = str(candidate)
            if path in seen:
                continue
            seen.add(path)
            if Path(path).exists():
                return path
        return ""

    def reserve_local_tcp_port(self) -> int:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.bind(("127.0.0.1", 0))
        port = sock.getsockname()[1]
        sock.close()
        return port

    def build_browser_launch_args(self, proxy_addr: str | None = None) -> list[str]:
        proxy_value = self.normalize_proxy_address(proxy_addr or "")
        if not proxy_value and self.use_proxy_var.get() and self.proxy_mode_var.get() == PROXY_MODE_SINGLE:
            proxy_value = self.normalize_proxy_address(self.proxy_entry.get())

        common_args = [
            "--headless=new",
            "--ignore-certificate-errors",
            "--disable-gpu",
            "--disable-popup-blocking",
            "--disable-notifications",
            "--disable-infobars",
            "--disable-blink-features=AutomationControlled",
            "--disable-features=MediaRouter,OptimizationHints,NotificationTriggers,PermissionChip,BlockInsecurePrivateNetworkRequests,PrivateNetworkAccessSendPreflights,PrivateNetworkAccessPermissionPrompt",
            "--window-size=1440,1080",
            "--log-level=3",
            "--disable-logging",
            "--no-first-run",
            "--no-default-browser-check",
            "--disable-extensions",
            "--disable-background-networking",
        ]
        if proxy_value:
            common_args.append(f"--proxy-server={proxy_value}")
        return common_args

    def launch_devtools_browser(self, proxy_addr: str | None = None) -> dict | None:
        try:
            import websocket  # noqa: F401
        except Exception:
            return None

        browser_path = self.find_browser_executable()
        if not browser_path:
            return None

        user_data_dir = tempfile.mkdtemp(prefix="passhack-cdp-")
        port = self.reserve_local_tcp_port()
        env = self.build_browser_env()

        command = [
            browser_path,
            f"--remote-debugging-port={port}",
            "--remote-allow-origins=*",
            *self.build_browser_launch_args(proxy_addr),
            f"--user-data-dir={user_data_dir}",
            "about:blank",
        ]
        process = None
        session = requests.Session()
        session.trust_env = False
        try:
            process = subprocess.Popen(command, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, env=env)
            targets = []
            for _ in range(50):
                try:
                    response = session.get(f"http://127.0.0.1:{port}/json/list", timeout=1)
                    if response.ok:
                        payload = response.json()
                        targets = [item for item in payload if item.get("type") == "page" and item.get("webSocketDebuggerUrl")]
                        if targets:
                            return {
                                "process": process,
                                "session": session,
                                "port": port,
                                "user_data_dir": user_data_dir,
                                "websocket_url": targets[0]["webSocketDebuggerUrl"],
                            }
                except Exception:
                    time.sleep(0.2)
            raise RuntimeError("DevTools target not ready")
        except Exception:
            try:
                session.close()
            except Exception:
                pass
            if process is not None:
                try:
                    process.terminate()
                    process.wait(timeout=5)
                except Exception:
                    try:
                        process.kill()
                    except Exception:
                        pass
            shutil.rmtree(user_data_dir, ignore_errors=True)
            return None

    def close_devtools_browser(self, context: dict | None):
        if not context:
            return
        session = context.get("session")
        if session is not None:
            try:
                session.close()
            except Exception:
                pass
        process = context.get("process")
        if process is not None:
            try:
                process.terminate()
                process.wait(timeout=5)
            except Exception:
                try:
                    process.kill()
                except Exception:
                    pass
        shutil.rmtree(context.get("user_data_dir", ""), ignore_errors=True)

    def probe_login_with_devtools(self, url: str, proxy_addr: str | None = None) -> dict | None:
        try:
            import websocket
        except Exception:
            return None

        context = self.launch_devtools_browser(proxy_addr)
        if not context:
            return None

        ws = None
        proxy_backup = {key: os.environ.get(key) for key in ("http_proxy", "https_proxy", "HTTP_PROXY", "HTTPS_PROXY", "ALL_PROXY", "all_proxy")}
        try:
            for key in proxy_backup:
                os.environ.pop(key, None)
            ws = websocket.create_connection(
                context["websocket_url"],
                timeout=12,
                http_proxy_host=None,
                http_proxy_port=None,
                origin=f"http://127.0.0.1:{context['port']}",
            )
            seq = {"id": 0}

            def send(method: str, params: dict | None = None) -> int:
                seq["id"] += 1
                payload = {"id": seq["id"], "method": method}
                if params is not None:
                    payload["params"] = params
                ws.send(json.dumps(payload))
                return seq["id"]

            def wait_response(target_id: int, timeout: float = 20.0) -> dict:
                deadline = time.time() + timeout
                while time.time() < deadline:
                    message = json.loads(ws.recv())
                    if message.get("id") == target_id:
                        return message
                raise TimeoutError(target_id)

            wait_response(send("Page.enable"))
            wait_response(send("Runtime.enable"))
            script = """
JSON.stringify((() => {
  const MAX_DEPTH = 3;
  const truncate = (value, limit = 160) => (value || '').replace(/\s+/g, ' ').trim().slice(0, limit);
  const isVisible = (node) => {
    if (!node) return false;
    const style = window.getComputedStyle(node);
    if (!style || style.display === 'none' || style.visibility === 'hidden' || style.opacity === '0') return false;
    const rect = node.getBoundingClientRect();
    return rect.width > 8 && rect.height > 8;
  };
  const inputs = [];
  const visitRoot = (root, depth = 0) => {
    if (!root || depth > MAX_DEPTH || !root.querySelectorAll) return;
    const labels = Array.from(root.querySelectorAll('label'));
    for (const el of Array.from(root.querySelectorAll('input, textarea'))) {
      if (!isVisible(el)) continue;
      const labelText = truncate(
        (el.id && labels.find((label) => (label.getAttribute('for') || '') === el.id)?.innerText)
        || el.closest('label')?.innerText
        || ''
      );
      const nearbyText = truncate(
        el.closest('form, [class*="login"], [id*="login"], [class*="signin"], [class*="auth"]')?.innerText
        || el.parentElement?.innerText
        || ''
      );
      inputs.push({
        tag: el.tagName.toLowerCase(),
        type: (el.getAttribute('type') || '').toLowerCase(),
        name: el.getAttribute('name') || '',
        id: el.getAttribute('id') || '',
        placeholder: el.getAttribute('placeholder') || '',
        aria: el.getAttribute('aria-label') || '',
        cls: el.getAttribute('class') || '',
        autocomplete: el.getAttribute('autocomplete') || '',
        title: el.getAttribute('title') || '',
        label: labelText,
        nearby: nearbyText
      });
    }
    for (const node of Array.from(root.querySelectorAll('*'))) {
      if (node.shadowRoot) visitRoot(node.shadowRoot, depth + 1);
    }
    for (const frame of Array.from(root.querySelectorAll('iframe, frame'))) {
      try {
        if (frame.contentDocument) visitRoot(frame.contentDocument, depth + 1);
      } catch (err) {
      }
    }
  };
  visitRoot(document, 0);
  return {
    current_url: location.href,
    html: document.documentElement.outerHTML,
    title: document.title,
    inputs: inputs
  };
})())
"""

            def collect_probe(navigate_url: str) -> dict | None:
                wait_response(send("Page.navigate", {"url": navigate_url}), 20.0)
                time.sleep(max(1.0, self.get_render_wait()))
                response = wait_response(send("Runtime.evaluate", {"expression": script, "returnByValue": True}), 20.0)
                raw_value = response.get("result", {}).get("result", {}).get("value", "")
                if not raw_value:
                    return None
                payload = json.loads(raw_value)
                inputs = payload.get("inputs") or []
                hints = []
                user_found = False
                pass_found = False
                captcha_found = False
                for item in inputs:
                    merged = " ".join(
                        [
                            item.get("name", ""),
                            item.get("id", ""),
                            item.get("placeholder", ""),
                            item.get("aria", ""),
                            item.get("autocomplete", ""),
                            item.get("cls", ""),
                            item.get("title", ""),
                            item.get("label", ""),
                            item.get("nearby", ""),
                        ]
                    ).lower()
                    field_type = (item.get("type") or "").lower()
                    if not pass_found and ("password" in field_type or any(keyword in merged for keyword in PASSWORD_FIELD_KEYWORDS)):
                        pass_found = True
                        hints.append("password:browser_password")
                        continue
                    if not captcha_found and any(keyword in merged for keyword in CAPTCHA_FIELD_KEYWORDS):
                        captcha_found = True
                        hints.append("text:browser_captcha")
                        continue
                    if field_type in {"hidden", "submit", "button", "checkbox", "radio", "file"}:
                        continue
                    if any(keyword in merged for keyword in SKIP_USERNAME_FIELD_KEYWORDS):
                        continue
                    if not user_found and (
                        any(keyword in merged for keyword in USER_FIELD_KEYWORDS)
                        or field_type in {"", "text", "email", "tel", "number", "search"}
                    ):
                        user_found = True
                        hints.append("text:browser_user")

                return {
                    "current_url": payload.get("current_url") or navigate_url,
                    "html": payload.get("html") or "",
                    "title": payload.get("title") or "",
                    "login_form": bool(pass_found),
                    "captcha_present": bool(captcha_found),
                    "field_summary": " | ".join(hints),
                    "form_method": "POST",
                    "probe_backend": "cdp",
                    "probe_url": navigate_url,
                    "input_count": len(inputs),
                }

            base_probe = collect_probe(url)
            if not base_probe:
                return None

            def append_route_attempt(route_attempts: list[str], value: str):
                normalized = (value or "").strip()
                if normalized and normalized not in route_attempts:
                    route_attempts.append(normalized)

            route_attempts = []
            append_route_attempt(route_attempts, base_probe.get("current_url") or base_probe.get("probe_url") or url)
            base_probe["route_attempts"] = list(route_attempts)
            if self.browser_probe_has_login_signal(base_probe):
                return base_probe

            best_probe = base_probe
            if self.should_try_common_login_routes(base_probe):
                for candidate_url in self.build_common_login_urls(base_probe.get("current_url") or url):
                    candidate_probe = collect_probe(candidate_url)
                    if not candidate_probe:
                        continue
                    append_route_attempt(route_attempts, candidate_probe.get("current_url") or candidate_url)
                    candidate_probe["route_candidate"] = candidate_url
                    candidate_probe["route_attempts"] = list(route_attempts)
                    if self.rank_browser_probe(candidate_probe) > self.rank_browser_probe(best_probe):
                        best_probe = candidate_probe
                    if self.browser_probe_has_login_signal(candidate_probe):
                        return candidate_probe
            best_probe["route_attempts"] = list(route_attempts)
            return best_probe
        except Exception:
            return None
        finally:
            if ws is not None:
                try:
                    ws.close()
                except Exception:
                    pass
            for key, value in proxy_backup.items():
                if value is None:
                    os.environ.pop(key, None)
                else:
                    os.environ[key] = value
            self.close_devtools_browser(context)

    def capture_screenshot_with_browser_cli(self, url: str, filename: Path, proxy_addr: str | None = None) -> Path | None:
        browser_path = self.find_browser_executable()
        if not browser_path:
            return None

        command = [
            browser_path,
            *self.build_browser_launch_args(proxy_addr),
            f"--screenshot={filename}",
            url,
        ]
        env = self.build_browser_env()
        try:
            result = subprocess.run(command, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, timeout=30, env=env)
            if result.returncode == 0 and filename.exists() and filename.stat().st_size > 0:
                return filename
        except Exception:
            return None
        return None

    def setup_ui(self):
        top_panel = tk.Frame(self.root, bg=self.palette["bg"], padx=12, pady=12)
        top_panel.pack(fill=tk.X)

        task_frame = self.make_section(top_panel, "任务与工程")
        task_frame.pack(fill=tk.X, pady=(0, 8))
        task_frame.columnconfigure(1, weight=1)

        self.make_label(task_frame, text="目标文件 (TXT/XLSX):").grid(row=0, column=0, sticky="w", pady=(0, 8))
        self.file_entry = self.make_entry(task_frame, width=60)
        self.file_entry.grid(row=0, column=1, sticky="ew", padx=(6, 6), pady=(0, 8))
        self.make_button(task_frame, text="浏览...", command=self.load_file).grid(row=0, column=2, padx=(0, 6), pady=(0, 8))
        self.make_button(task_frame, text="刷新源文件", command=self.reload_current_source_file).grid(row=0, column=3, padx=(0, 6), pady=(0, 8))
        self.make_button(task_frame, text="导出结果", command=self.save_results).grid(row=0, column=4, pady=(0, 8))

        self.summary_var = tk.StringVar(value="摘要: 总计 0 | 已完成 0 | 待处理 0 | 高 0 | 中 0 | 低 0 | 有截图 0")
        self.make_label(
            task_frame,
            textvariable=self.summary_var,
            anchor="w",
            justify=tk.LEFT,
            wraplength=980,
            muted=True,
        ).grid(row=1, column=0, columnspan=2, sticky="w")

        task_button_bar = tk.Frame(task_frame, bg=task_frame.cget("bg"))
        task_button_bar.grid(row=1, column=2, columnspan=2, sticky="e")
        self.btn_start = self.make_button(task_button_bar, text="开始审计", command=self.toggle_scan, tone="primary")
        self.btn_start.pack(side=tk.LEFT, padx=(0, 4))
        self.btn_rescan = self.make_button(task_button_bar, text="重新扫描", command=self.restart_scan, tone="warning")
        self.btn_rescan.pack(side=tk.LEFT, padx=4)
        self.btn_retry_failed = self.make_button(task_button_bar, text="重扫失败项", command=self.retry_failed_targets, tone="danger")
        self.btn_retry_failed.pack(side=tk.LEFT, padx=(4, 0))

        config_notebook = ttk.Notebook(top_panel)
        config_notebook.pack(fill=tk.X, pady=(0, 8))

        scan_tab = tk.Frame(config_notebook, bg=self.palette["bg"])
        brute_tab = tk.Frame(config_notebook, bg=self.palette["bg"])
        render_tab = tk.Frame(config_notebook, bg=self.palette["bg"])
        quick_tab = tk.Frame(config_notebook, bg=self.palette["bg"])
        log_tab = tk.Frame(config_notebook, bg=self.palette["bg"])
        config_notebook.add(scan_tab, text="扫描配置")
        config_notebook.add(brute_tab, text="弱口令检测")
        config_notebook.add(render_tab, text="浏览器 / OCR / 规则")
        config_notebook.add(quick_tab, text="快捷操作")
        config_notebook.add(log_tab, text="运行日志")

        scan_frame = self.make_section(scan_tab, "扫描配置")
        scan_frame.pack(fill=tk.X, pady=8)

        proxy_row = tk.Frame(scan_frame, bg=scan_frame.cget("bg"))
        proxy_row.pack(fill=tk.X)
        proxy_row.columnconfigure(2, weight=1)
        proxy_row.columnconfigure(6, weight=1)

        self.use_proxy_var = tk.BooleanVar(value=False)
        self.make_checkbutton(
            proxy_row,
            text="启用全局代理 (如 Burp Suite)",
            variable=self.use_proxy_var,
            command=self.toggle_proxy_state,
        ).grid(row=0, column=0, sticky="w", padx=(0, 12))

        self.make_label(proxy_row, text="地址:").grid(row=0, column=1, sticky="w", padx=(0, 4))
        self.proxy_entry = self.make_entry(proxy_row, width=22, state=tk.DISABLED)
        self.proxy_entry.insert(0, "127.0.0.1:8080")
        self.proxy_entry.grid(row=0, column=2, sticky="ew", padx=(0, 12))

        self.make_label(proxy_row, text="代理模式:").grid(row=0, column=3, sticky="w", padx=(0, 4))
        self.proxy_mode_var = tk.StringVar(value=PROXY_MODE_SINGLE)
        self.proxy_mode_box = ttk.Combobox(
            proxy_row,
            textvariable=self.proxy_mode_var,
            values=[PROXY_MODE_SINGLE, PROXY_MODE_POOL],
            width=14,
            state="readonly",
        )
        self.proxy_mode_box.grid(row=0, column=4, sticky="w", padx=(0, 12))
        self.proxy_mode_box.bind("<<ComboboxSelected>>", lambda _event: self.toggle_proxy_state())

        self.make_label(proxy_row, text="代理池:").grid(row=0, column=5, sticky="w", padx=(0, 4))
        self.proxy_pool_entry = self.make_entry(proxy_row, width=28, state=tk.DISABLED)
        self.proxy_pool_entry.grid(row=0, column=6, sticky="ew", padx=(0, 6))
        self.proxy_pool_button = self.make_button(proxy_row, text="选择...", command=self.browse_proxy_pool, state=tk.DISABLED)
        self.proxy_pool_button.grid(row=0, column=7, sticky="w")

        scan_row = tk.Frame(scan_frame, bg=scan_frame.cget("bg"))
        scan_row.pack(fill=tk.X, pady=(8, 0))

        self.make_label(scan_row, text="失败重试:").grid(row=0, column=0, sticky="w")
        self.proxy_retry_var = tk.StringVar(value="2")
        self.proxy_retry_box = ttk.Combobox(scan_row, textvariable=self.proxy_retry_var, values=["1", "2", "3", "4", "5"], width=4, state="readonly")
        self.proxy_retry_box.grid(row=0, column=1, sticky="w", padx=(4, 12))

        self.make_label(scan_row, text="熔断阈值:").grid(row=0, column=2, sticky="w")
        self.proxy_fail_threshold_var = tk.StringVar(value="2")
        self.proxy_fail_threshold_box = ttk.Combobox(scan_row, textvariable=self.proxy_fail_threshold_var, values=["1", "2", "3", "4", "5"], width=4, state="readonly")
        self.proxy_fail_threshold_box.grid(row=0, column=3, sticky="w", padx=(4, 12))

        self.make_label(scan_row, text="冷却:").grid(row=0, column=4, sticky="w")
        self.proxy_cooldown_var = tk.StringVar(value="120")
        self.proxy_cooldown_box = ttk.Combobox(scan_row, textvariable=self.proxy_cooldown_var, values=["30", "60", "120", "300", "600"], width=5, state="readonly")
        self.proxy_cooldown_box.grid(row=0, column=5, sticky="w", padx=(4, 12))

        self.make_label(scan_row, text="识别模式:").grid(row=0, column=6, sticky="w")
        self.mode_var = tk.StringVar(value="规则模式")
        ttk.Combobox(scan_row, textvariable=self.mode_var, values=["规则模式", "NLP模式"], width=12, state="readonly").grid(row=0, column=7, sticky="w", padx=(4, 12))

        self.capture_var = tk.BooleanVar(value=True)
        self.make_checkbutton(scan_row, text="尝试截图", variable=self.capture_var).grid(row=0, column=8, sticky="w", padx=(0, 12))

        self.make_label(scan_row, text="截图策略:").grid(row=0, column=9, sticky="w")
        self.capture_policy_var = tk.StringVar(value=CAPTURE_POLICY_HIT)
        ttk.Combobox(
            scan_row,
            textvariable=self.capture_policy_var,
            values=CAPTURE_POLICY_OPTIONS,
            width=14,
            state="readonly",
        ).grid(row=0, column=10, sticky="w", padx=(4, 0))

        advanced_row = tk.Frame(scan_frame, bg=scan_frame.cget("bg"))
        advanced_row.pack(fill=tk.X, pady=(8, 0))

        self.follow_redirect_var = tk.BooleanVar(value=True)
        self.make_checkbutton(advanced_row, text="跟随重定向", variable=self.follow_redirect_var).grid(row=0, column=0, sticky="w", padx=(0, 12))

        self.make_label(advanced_row, text="并发数:").grid(row=0, column=1, sticky="w")
        self.worker_var = tk.StringVar(value="4")
        ttk.Combobox(advanced_row, textvariable=self.worker_var, values=["1", "2", "4", "6", "8", "12"], width=4, state="readonly").grid(row=0, column=2, sticky="w", padx=(4, 12))

        self.make_label(advanced_row, text="截图节流:").grid(row=0, column=3, sticky="w")
        self.capture_delay_var = tk.StringVar(value="0.4")
        ttk.Combobox(advanced_row, textvariable=self.capture_delay_var, values=["0", "0.2", "0.4", "0.8", "1.2"], width=5, state="readonly").grid(row=0, column=4, sticky="w", padx=(4, 12))

        self.make_label(advanced_row, text="列表筛选:").grid(row=0, column=5, sticky="w")
        self.filter_var = tk.StringVar(value="全部")
        filter_box = ttk.Combobox(
            advanced_row,
            textvariable=self.filter_var,
            values=["全部", "仅高风险", "仅疑似登录页", "仅已完成", "仅失败", "仅有截图", "仅有表单字段"],
            width=13,
            state="readonly",
        )
        filter_box.grid(row=0, column=6, sticky="w", padx=(4, 0))
        filter_box.bind("<<ComboboxSelected>>", lambda _event: self.apply_filter())

        brute_frame = self.make_section(brute_tab, "弱口令检测")
        brute_frame.pack(fill=tk.X, pady=8)

        brute_row1 = tk.Frame(brute_frame, bg=brute_frame.cget("bg"))
        brute_row1.pack(fill=tk.X)
        brute_row1.columnconfigure(4, weight=1)

        self.make_checkbutton(
            brute_row1,
            text="启用弱口令检测",
            variable=self.brute_var,
            command=self.toggle_brute_dict_state,
        ).grid(row=0, column=0, sticky="w", padx=(0, 12))

        self.make_label(brute_row1, text="字典模式:").grid(row=0, column=1, sticky="w")
        self.dict_mode_var = tk.StringVar(value=DEFAULT_BRUTE_DICT_MODE)
        self.dict_mode_box = ttk.Combobox(
            brute_row1,
            textvariable=self.dict_mode_var,
            values=[DEFAULT_BRUTE_DICT_MODE, CUSTOM_BRUTE_DICT_MODE],
            width=20,
            state="readonly",
        )
        self.dict_mode_box.grid(row=0, column=2, sticky="w", padx=(4, 12))
        self.dict_mode_box.bind("<<ComboboxSelected>>", lambda _event: self.toggle_brute_dict_state())

        self.make_label(brute_row1, text="账号字典:").grid(row=0, column=3, sticky="w")
        self.user_dict_entry = self.make_entry(brute_row1, width=36)
        self.user_dict_entry.grid(row=0, column=4, sticky="ew", padx=(4, 6))
        self.user_dict_button = self.make_button(brute_row1, text="选择...", command=self.browse_user_dict)
        self.user_dict_button.grid(row=0, column=5, sticky="w")

        brute_row2 = tk.Frame(brute_frame, bg=brute_frame.cget("bg"))
        brute_row2.pack(fill=tk.X, pady=(8, 0))
        brute_row2.columnconfigure(1, weight=1)

        self.make_label(brute_row2, text="密码字典:").grid(row=0, column=0, sticky="w")
        self.pass_dict_entry = self.make_entry(brute_row2, width=36)
        self.pass_dict_entry.grid(row=0, column=1, sticky="ew", padx=(4, 6))
        self.pass_dict_button = self.make_button(brute_row2, text="选择...", command=self.browse_pass_dict)
        self.pass_dict_button.grid(row=0, column=2, sticky="w")
        self.load_builtin_dict_button = self.make_button(brute_row2, text="载入内置文件", command=self.load_builtin_dict_files_into_form)
        self.load_builtin_dict_button.grid(row=0, column=3, sticky="w", padx=(8, 0))
        self.open_builtin_dict_button = self.make_button(brute_row2, text="打开内置字典", command=self.open_builtin_dict_dir)
        self.open_builtin_dict_button.grid(row=0, column=4, sticky="w", padx=(8, 0))
        self.make_label(brute_row2, text="默认模式会直接使用内置常用账号/密码字典。", muted=True).grid(row=0, column=5, sticky="w", padx=(12, 0))

        self.toggle_brute_dict_state()

        render_frame = self.make_section(render_tab, "浏览器补扫 / OCR / 定位规则")
        render_frame.pack(fill=tk.X, pady=8)

        render_row1 = tk.Frame(render_frame, bg=render_frame.cget("bg"))
        render_row1.pack(fill=tk.X)
        render_row1.columnconfigure(3, weight=1)

        self.make_checkbutton(render_row1, text="启用浏览器补扫识别", variable=self.browser_render_var).grid(row=0, column=0, sticky="w", padx=(0, 12))
        self.make_label(render_row1, text="渲染等待:").grid(row=0, column=1, sticky="w")
        self.render_wait_var = tk.StringVar(value="2.5")
        ttk.Combobox(render_row1, textvariable=self.render_wait_var, values=["1.2", "2.0", "2.5", "3.5", "5.0", "8.0"], width=5, state="readonly").grid(row=0, column=2, sticky="w", padx=(4, 8))
        self.make_label(render_row1, text="秒，仅对静态识别不足目标触发", muted=True).grid(row=0, column=3, sticky="w")
        self.make_checkbutton(render_row1, text="启用本地验证码OCR", variable=self.captcha_ocr_var).grid(row=0, column=4, sticky="e", padx=(12, 0))

        render_row2 = tk.Frame(render_frame, bg=render_frame.cget("bg"))
        render_row2.pack(fill=tk.X, pady=(8, 0))
        render_row2.columnconfigure(1, weight=1)
        render_row2.columnconfigure(3, weight=1)

        self.make_label(render_row2, text="OCR接口:").grid(row=0, column=0, sticky="w")
        self.ocr_endpoint_entry = self.make_entry(render_row2, width=32)
        self.ocr_endpoint_entry.insert(0, DEFAULT_OCR_ENDPOINT)
        self.ocr_endpoint_entry.grid(row=0, column=1, sticky="ew", padx=(4, 12))

        self.make_label(render_row2, text="映射文件:").grid(row=0, column=2, sticky="w")
        self.ocr_route_entry = self.make_entry(render_row2, width=30)
        self.ocr_route_entry.grid(row=0, column=3, sticky="ew", padx=(4, 6))
        self.ocr_route_button = self.make_button(render_row2, text="选择...", command=self.browse_ocr_route_file)
        self.ocr_route_button.grid(row=0, column=4, sticky="w")
        self.ocr_route_template_button = self.make_button(render_row2, text="模板", command=self.open_ocr_route_template)
        self.ocr_route_template_button.grid(row=0, column=5, sticky="w", padx=(4, 0))
        self.ocr_route_import_button = self.make_button(render_row2, text="导入", command=self.import_ocr_route_library, tone="success")
        self.ocr_route_import_button.grid(row=0, column=6, sticky="w", padx=(4, 0))
        self.ocr_route_export_button = self.make_button(render_row2, text="导出", command=self.export_ocr_route_library)
        self.ocr_route_export_button.grid(row=0, column=7, sticky="w", padx=(4, 0))

        render_row3 = tk.Frame(render_frame, bg=render_frame.cget("bg"))
        render_row3.pack(fill=tk.X, pady=(8, 0))
        render_row3.columnconfigure(1, weight=1)

        self.make_label(render_row3, text="定位规则:").grid(row=0, column=0, sticky="w")
        self.locator_rule_entry = self.make_entry(render_row3, width=42)
        self.locator_rule_entry.grid(row=0, column=1, sticky="ew", padx=(4, 6))
        self.locator_rule_button = self.make_button(render_row3, text="选择...", command=self.browse_locator_rule_file)
        self.locator_rule_button.grid(row=0, column=2, sticky="w")
        self.locator_rule_template_button = self.make_button(render_row3, text="模板", command=self.open_locator_rule_template)
        self.locator_rule_template_button.grid(row=0, column=3, sticky="w", padx=(4, 0))
        self.locator_rule_quick_button = self.make_button(render_row3, text="录入规则", command=self.open_locator_rule_quick_editor, tone="primary")
        self.locator_rule_quick_button.grid(row=0, column=4, sticky="w", padx=(4, 0))
        self.make_label(
            render_frame,
            text="录入规则 = 把这个站点的登录框定位方式保存到规则文件，不是填账号密码。下次扫描到同类站点会直接按规则找登录框。",
            muted=True,
            anchor="w",
            justify=tk.LEFT,
            wraplength=1280,
        ).pack(fill=tk.X, pady=(8, 0))

        action_frame = self.make_section(quick_tab, "快捷操作")
        action_frame.pack(fill=tk.X, pady=8)
        action_buttons = [
            ("工程列表", self.open_project_manager),
            ("画像统计", self.open_profile_dashboard),
            ("代理状态", self.open_proxy_status_window),
            ("放大查看", self.open_preview_zoom),
            ("查看截图", self.open_selected_screenshot),
            ("结果详情", self.open_selected_detail),
            ("打开目标", self.open_selected_target),
            ("打开工程目录", self.open_output_dir),
            ("导出证据页", self.export_evidence_page),
        ]
        for index, (text, command) in enumerate(action_buttons):
            action_frame.columnconfigure(index % 5, weight=1)
            self.make_button(action_frame, text=text, command=command).grid(
                row=index // 5,
                column=index % 5,
                sticky="ew",
                padx=4,
                pady=4,
            )

        log_frame = self.make_section(log_tab, "运行日志", padx=10, pady=10)
        log_frame.pack(fill=tk.BOTH, expand=True, pady=8)
        log_head = tk.Frame(log_frame, bg=log_frame.cget("bg"))
        log_head.pack(fill=tk.X, pady=(0, 6))
        self.detail_log_path_var = tk.StringVar(value="详细日志: 尚未创建")
        self.make_label(
            log_head,
            textvariable=self.detail_log_path_var,
            anchor="w",
            justify=tk.LEFT,
            muted=True,
            wraplength=1160,
        ).pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.make_button(log_head, text="打开详细日志", command=self.open_detail_log_file).pack(side=tk.RIGHT)
        self.make_button(log_head, text="清空界面日志", command=self.clear_ui_log).pack(side=tk.RIGHT, padx=(0, 6))

        self.log_text = scrolledtext.ScrolledText(
            log_frame,
            height=9,
            state=tk.DISABLED,
            wrap=tk.WORD,
            bg="#f8fafc",
            fg=self.palette["text"],
            insertbackground=self.palette["text"],
            relief=tk.SOLID,
            bd=1,
            font=("Consolas", 9),
        )
        self.log_text.pack(fill=tk.BOTH, expand=True)

        mid_frame = tk.Frame(self.root, bg=self.palette["bg"], padx=12)
        mid_frame.pack(fill=tk.BOTH, expand=True)

        split_pane = ttk.Panedwindow(mid_frame, orient=tk.HORIZONTAL)
        split_pane.pack(fill=tk.BOTH, expand=True)

        left_panel = tk.Frame(split_pane, bg=self.palette["panel"], bd=1, relief=tk.GROOVE)
        preview_frame = tk.LabelFrame(
            split_pane,
            text="详情预览",
            padx=10,
            pady=10,
            bg=self.palette["panel"],
            fg=self.palette["text"],
            bd=1,
            relief=tk.GROOVE,
            font=self.font_section,
        )
        split_pane.add(left_panel, weight=5)
        split_pane.add(preview_frame, weight=2)

        self.make_label(left_panel, text="扫描结果", font=self.font_section).pack(anchor=tk.W, padx=10, pady=(10, 6))
        tree_frame = tk.Frame(left_panel, bg=self.palette["panel"])
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)

        columns = ("ID", "目标资产", "状态", "页面标题", "风险级别", "审计结果")
        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=12)
        v_scroll = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        h_scroll = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)

        self.configure_tree_columns(self.tree, [
            ("ID", 56, tk.CENTER),
            ("目标资产", 280, tk.W),
            ("状态", 92, tk.CENTER),
            ("页面标题", 260, tk.W),
            ("风险级别", 80, tk.CENTER),
            ("审计结果", 720, tk.W),
        ])

        self.tree.tag_configure("risk_high", background="#ffe1e1", foreground=self.palette["text"])
        self.tree.tag_configure("risk_medium", background="#fff0cf", foreground=self.palette["text"])
        self.tree.tag_configure("risk_low", background="#edf9ed", foreground=self.palette["text"])
        self.tree.tag_configure("status_error", background="#f3d6d6", foreground=self.palette["text"])
        self.tree.bind("<Double-1>", self.on_tree_double_click)
        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)

        self.tree.grid(row=0, column=0, sticky="nsew")
        v_scroll.grid(row=0, column=1, sticky="ns")
        h_scroll.grid(row=1, column=0, sticky="ew")

        preview_frame.configure(width=480)
        preview_frame.pack_propagate(False)

        self.preview_title_var = tk.StringVar(value="未选择记录")
        self.make_label(
            preview_frame,
            textvariable=self.preview_title_var,
            anchor="w",
            justify=tk.LEFT,
            wraplength=500,
            font=self.font_title,
        ).pack(fill=tk.X)

        self.preview_meta_var = tk.StringVar(value="目标、风险、最终URL 会显示在这里")
        self.make_label(
            preview_frame,
            textvariable=self.preview_meta_var,
            anchor="w",
            justify=tk.LEFT,
            wraplength=500,
            muted=True,
        ).pack(fill=tk.X, pady=(6, 8))

        nav_frame = tk.Frame(preview_frame, bg=preview_frame.cget("bg"))
        nav_frame.pack(fill=tk.X, pady=(0, 8))
        self.make_button(nav_frame, text="上一条", command=lambda: self.select_relative_record(-1)).pack(side=tk.LEFT)
        self.make_button(nav_frame, text="下一条", command=lambda: self.select_relative_record(1)).pack(side=tk.LEFT, padx=(6, 0))

        self.preview_image_label = tk.Label(
            preview_frame,
            text="暂无截图",
            bg="#f8fafc",
            fg=self.palette["muted"],
            width=56,
            height=16,
            relief=tk.SOLID,
            bd=1,
            anchor="center",
        )
        self.preview_image_label.pack(fill=tk.BOTH, pady=(0, 8))
        self.preview_image_label.bind("<Double-1>", lambda _event: self.open_preview_zoom())

        self.make_label(preview_frame, text="表单与取证信息:").pack(anchor=tk.W, pady=(10, 4))
        preview_text_frame = tk.Frame(preview_frame, bg=preview_frame.cget("bg"))
        preview_text_frame.pack(fill=tk.BOTH, expand=True)
        preview_text_frame.columnconfigure(0, weight=1)
        preview_text_frame.rowconfigure(0, weight=1)

        preview_text_vscroll = ttk.Scrollbar(preview_text_frame, orient="vertical")
        preview_text_hscroll = ttk.Scrollbar(preview_text_frame, orient="horizontal")
        self.preview_text = tk.Text(
            preview_text_frame,
            height=14,
            state=tk.DISABLED,
            wrap=tk.NONE,
            yscrollcommand=preview_text_vscroll.set,
            xscrollcommand=preview_text_hscroll.set,
            bg="#f8fafc",
            fg=self.palette["text"],
            insertbackground=self.palette["text"],
            relief=tk.SOLID,
            bd=1,
            highlightthickness=1,
            highlightbackground=self.palette["border"],
            highlightcolor=self.palette["accent"],
            font=self.font_normal,
        )
        preview_text_vscroll.config(command=self.preview_text.yview)
        preview_text_hscroll.config(command=self.preview_text.xview)
        self.preview_text.grid(row=0, column=0, sticky="nsew")
        preview_text_vscroll.grid(row=0, column=1, sticky="ns")
        preview_text_hscroll.grid(row=1, column=0, sticky="ew")
        self.toggle_proxy_state()

    def toggle_proxy_state(self):
        enabled = self.use_proxy_var.get()
        pool_mode = self.proxy_mode_var.get() == PROXY_MODE_POOL

        self.proxy_mode_box.config(state="readonly" if enabled else tk.DISABLED)
        self.proxy_retry_box.config(state="readonly" if enabled else tk.DISABLED)
        self.proxy_fail_threshold_box.config(state="readonly" if enabled else tk.DISABLED)
        self.proxy_cooldown_box.config(state="readonly" if enabled else tk.DISABLED)
        self.proxy_entry.config(state=tk.NORMAL if enabled and not pool_mode else tk.DISABLED)
        self.proxy_pool_entry.config(state=tk.NORMAL if enabled and pool_mode else tk.DISABLED)
        self.proxy_pool_button.config(state=tk.NORMAL if enabled and pool_mode else tk.DISABLED)

    def disable_proxy_controls_for_scan(self):
        self.proxy_entry.config(state=tk.DISABLED)
        self.proxy_mode_box.config(state=tk.DISABLED)
        self.proxy_pool_entry.config(state=tk.DISABLED)
        self.proxy_pool_button.config(state=tk.DISABLED)
        self.proxy_retry_box.config(state=tk.DISABLED)
        self.proxy_fail_threshold_box.config(state=tk.DISABLED)
        self.proxy_cooldown_box.config(state=tk.DISABLED)

    def browse_ocr_route_file(self):
        path = filedialog.askopenfilename(
            title="选择OCR映射文件",
            filetypes=[("Text files", "*.txt *.cfg *.ini *.lst"), ("All files", "*.*")],
        )
        if not path:
            return
        self.ocr_route_entry.delete(0, tk.END)
        self.ocr_route_entry.insert(0, path)

    def browse_locator_rule_file(self):
        path = filedialog.askopenfilename(
            title="选择定位规则文件",
            filetypes=[("Text files", "*.txt *.cfg *.ini *.lst"), ("All files", "*.*")],
        )
        if not path:
            return
        self.locator_rule_entry.delete(0, tk.END)
        self.locator_rule_entry.insert(0, path)

    def get_ocr_rule_library_dir(self) -> Path:
        path = self.state_dir / "ocr_rules"
        path.mkdir(parents=True, exist_ok=True)
        return path

    def get_locator_rule_template_path(self) -> Path:
        return self.state_dir / "locator_rule_template.txt"

    def ensure_locator_rule_file(self) -> Path:
        path_value = self.locator_rule_entry.get().strip()
        if path_value:
            path = Path(path_value)
            path.parent.mkdir(parents=True, exist_ok=True)
            if not path.exists():
                path.write_text("", encoding="utf-8")
            return path
        template_path = self.ensure_locator_rule_template_file()
        self.locator_rule_entry.delete(0, tk.END)
        self.locator_rule_entry.insert(0, str(template_path))
        return template_path

    def ensure_locator_rule_template_file(self) -> Path:
        template_path = self.get_locator_rule_template_path()
        if not template_path.exists():
            template_path.parent.mkdir(parents=True, exist_ok=True)
            template_path.write_text(
                "# 定位规则示例\n"
                "# 规则格式: scope:pattern => key=value; key=value\n"
                "# 支持 scope: host / target / url / title / default\n"
                "# 支持 key: user / pass / submit / captcha / frame\n"
                "# value 支持 css: 或 xpath: 前缀; 不写前缀时默认按 css 处理\n"
                "\n"
                "# 示例: 当前这类宅基地系统\n"
                "host:27.150.180.183 => user=css:input[placeholder*='账号'], input[placeholder*='手机号']; pass=css:input[type='password']; submit=xpath://button[contains(.,'登录')]\n"
                "# 示例: 登录框在 iframe 里\n"
                "title:登录 => frame=css:iframe; user=css:input[name='username']; pass=css:input[type='password']; submit=css:button[type='submit']\n"
                "# 兜底规则可选\n"
                "default => user=css:input[type='text']; pass=css:input[type='password']\n",
                encoding="utf-8",
            )
        return template_path

    def open_locator_rule_template(self):
        template_path = self.ensure_locator_rule_template_file()
        self.locator_rule_entry.delete(0, tk.END)
        self.locator_rule_entry.insert(0, str(template_path))
        self.log_message(f"[*] 已准备定位规则模板: {template_path}")
        os.startfile(template_path)

    def append_locator_rule(self, rule_text: str):
        target_file = self.ensure_locator_rule_file()
        existing = ""
        if target_file.exists():
            existing = target_file.read_text(encoding="utf-8", errors="ignore")
        content = existing.rstrip() + ("\n" if existing.strip() else "") + rule_text.strip() + "\n"
        target_file.write_text(content, encoding="utf-8")
        self.locator_rule_entry.delete(0, tk.END)
        self.locator_rule_entry.insert(0, str(target_file))
        self.log_message(f"[*] 已写入定位规则: {rule_text}")

    def import_ocr_route_library(self):
        source = filedialog.askopenfilename(
            title="导入站点规则库",
            filetypes=[("Text files", "*.txt *.cfg *.ini *.lst"), ("All files", "*.*")],
        )
        if not source:
            return
        src_path = Path(source)
        target_dir = self.get_ocr_rule_library_dir()
        target_path = target_dir / src_path.name
        if target_path.exists():
            stem = src_path.stem
            suffix = src_path.suffix or ".txt"
            target_path = target_dir / f"{stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{suffix}"
        target_path.write_text(src_path.read_text(encoding="utf-8", errors="ignore"), encoding="utf-8")
        self.ocr_route_entry.delete(0, tk.END)
        self.ocr_route_entry.insert(0, str(target_path))
        self.log_message(f"[*] 已导入站点规则库: {target_path}")

    def export_ocr_route_library(self):
        source_value = self.ocr_route_entry.get().strip()
        if not source_value:
            messagebox.showinfo("提示", "当前没有可导出的站点规则库。")
            return
        source = Path(source_value)
        if not source.exists():
            messagebox.showwarning("提示", f"规则库文件不存在:\n{source}")
            return
        output = filedialog.asksaveasfilename(
            title="导出站点规则库",
            initialfile=source.name,
            defaultextension=source.suffix or ".txt",
            filetypes=[("Text files", "*.txt *.cfg *.ini *.lst"), ("All files", "*.*")],
        )
        if not output:
            return
        output_path = Path(output)
        output_path.write_text(source.read_text(encoding="utf-8", errors="ignore"), encoding="utf-8")
        self.log_message(f"[*] 已导出站点规则库: {output_path}")

    def get_ocr_route_template_path(self) -> Path:
        return self.state_dir / "ocr_route_template.txt"

    def ensure_ocr_route_template_file(self) -> Path:
        template_path = self.get_ocr_route_template_path()
        if not template_path.exists():
            template_path.parent.mkdir(parents=True, exist_ok=True)
            template_path.write_text(
                "# OCR 映射文件示例\n"
                "# 规则格式: scope:pattern => endpoint\n"
                "# 支持 scope: host / target / url / title / default\n"
                "# 匹配逻辑: 包含匹配, 按文件顺序命中第一条\n"
                "# 推荐先写 host 级别精确规则, 最后保留 default 兜底\n"
                "\n"
                "# 示例: 某些站点是纯数字图形验证码\n"
                "host:47.109.128.200 => http://127.0.0.1:8888/reg00\n"
                "# 示例: 标题里带登录页的算术验证码\n"
                "title:登录页 => http://127.0.0.1:8888/reg01\n"
                "# 示例: 某类 URL 走自定义模型\n"
                "url:/user/login => http://127.0.0.1:8888/reg\n"
                "# 兜底规则\n"
                "default => http://127.0.0.1:8888/reg\n",
                encoding="utf-8",
            )
        return template_path

    def open_ocr_route_template(self):
        template_path = self.ensure_ocr_route_template_file()
        self.ocr_route_entry.delete(0, tk.END)
        self.ocr_route_entry.insert(0, str(template_path))
        self.log_message(f"[*] 已准备OCR映射模板: {template_path}")
        os.startfile(template_path)

    def browse_proxy_pool(self):
        path = filedialog.askopenfilename(
            title="选择代理池文件",
            filetypes=[("Text files", "*.txt *.lst *.csv"), ("All files", "*.*")],
        )
        if not path:
            return
        self.proxy_pool_entry.config(state=tk.NORMAL)
        self.proxy_pool_entry.delete(0, tk.END)
        self.proxy_pool_entry.insert(0, path)
        self.toggle_proxy_state()

    def normalize_proxy_address(self, value: str) -> str:
        proxy = (value or "").strip()
        if not proxy:
            return ""
        if "://" not in proxy:
            proxy = f"http://{proxy}"
        return proxy

    def get_proxy_retry_count(self) -> int:
        try:
            return max(1, min(5, int(self.proxy_retry_var.get())))
        except Exception:
            return 2

    def get_proxy_fail_threshold(self) -> int:
        try:
            return max(1, min(5, int(self.proxy_fail_threshold_var.get())))
        except Exception:
            return 2

    def get_proxy_cooldown_seconds(self) -> int:
        try:
            return max(10, min(3600, int(self.proxy_cooldown_var.get())))
        except Exception:
            return 120

    def read_text_lines(self, path: Path) -> list[str]:
        last_error = None
        for encoding in DICT_FILE_ENCODINGS:
            try:
                with open(path, "r", encoding=encoding) as handle:
                    return handle.readlines()
            except UnicodeError as exc:
                last_error = exc
        if last_error:
            raise last_error
        return []

    def load_ocr_route_rules(self) -> list[dict]:
        path_value = self.ocr_route_entry.get().strip()
        if not path_value:
            return []
        path = Path(path_value)
        if not path.exists():
            return []
        try:
            lines = self.read_text_lines(path)
        except UnicodeError:
            return []

        rules = []
        for line in lines:
            raw = line.strip()
            if not raw or raw.startswith(("#", ";", "//")):
                continue
            if "=>" not in raw:
                continue
            left, right = raw.split("=>", 1)
            matcher = left.strip()
            endpoint = right.strip()
            if not matcher or not endpoint:
                continue
            scope = "host"
            pattern = matcher
            if ":" in matcher:
                possible_scope, possible_pattern = matcher.split(":", 1)
                normalized_scope = possible_scope.strip().lower()
                if normalized_scope in {"host", "url", "title", "target", "default"}:
                    scope = normalized_scope
                    pattern = possible_pattern.strip()
            if matcher.strip().lower() in {"default", "*"}:
                scope = "default"
                pattern = "*"
            rules.append({"scope": scope, "pattern": pattern.lower(), "endpoint": endpoint})
        return rules

    def load_locator_rules(self) -> list[dict]:
        path_value = self.locator_rule_entry.get().strip()
        if not path_value:
            return []
        path = Path(path_value)
        if not path.exists():
            return []
        try:
            lines = self.read_text_lines(path)
        except UnicodeError:
            return []

        rules = []
        for line in lines:
            raw = line.strip()
            if not raw or raw.startswith(("#", ";", "//")):
                continue
            if "=>" not in raw:
                continue
            left, right = raw.split("=>", 1)
            matcher = left.strip()
            spec = right.strip()
            if not matcher or not spec:
                continue

            scope = "host"
            pattern = matcher
            if ":" in matcher:
                possible_scope, possible_pattern = matcher.split(":", 1)
                normalized_scope = possible_scope.strip().lower()
                if normalized_scope in {"host", "url", "title", "target", "default"}:
                    scope = normalized_scope
                    pattern = possible_pattern.strip()
            if matcher.strip().lower() in {"default", "*"}:
                scope = "default"
                pattern = "*"

            selectors = {}
            for part in spec.split(";"):
                item = part.strip()
                if not item or "=" not in item:
                    continue
                key, value = item.split("=", 1)
                key = key.strip().lower()
                value = value.strip()
                if key in {"user", "pass", "submit", "captcha", "frame"} and value:
                    selectors[key] = value
            if selectors:
                rules.append(
                    {
                        "scope": scope,
                        "pattern": pattern.lower(),
                        "rule_text": raw,
                        "selectors": selectors,
                    }
                )
        return rules

    def resolve_locator_rule_for_record(self, record: AuditRecord | None) -> dict | None:
        if record is None:
            return None
        target = (record.target or "").lower()
        final_url = (record.final_url or "").lower()
        title = (record.title or "").lower()
        default_rule = None
        for rule in self.load_locator_rules():
            scope = rule["scope"]
            pattern = rule["pattern"]
            if scope == "default":
                default_rule = rule
                continue
            if scope in {"host", "target"} and pattern and pattern in target:
                return rule
            if scope == "url" and pattern and pattern in final_url:
                return rule
            if scope == "title" and pattern and pattern in title:
                return rule
        return default_rule

    def resolve_ocr_endpoint_for_record(self, record: AuditRecord | None) -> str:
        endpoint, _rule = self.resolve_ocr_route_for_record(record)
        return endpoint

    def resolve_ocr_route_for_record(self, record: AuditRecord | None) -> tuple[str, str]:
        default_endpoint = self.ocr_endpoint_entry.get().strip() or DEFAULT_OCR_ENDPOINT
        default_rule = "default => " + default_endpoint
        if record is None:
            return default_endpoint, default_rule

        target = (record.target or "").lower()
        final_url = (record.final_url or "").lower()
        title = (record.title or "").lower()
        for rule in self.load_ocr_route_rules():
            scope = rule["scope"]
            pattern = rule["pattern"]
            if scope == "default":
                return rule["endpoint"], f"default => {rule['endpoint']}"
            if scope in {"host", "target"} and pattern and pattern in target:
                return rule["endpoint"], f"{scope}:{pattern} => {rule['endpoint']}"
            if scope == "url" and pattern and pattern in final_url:
                return rule["endpoint"], f"url:{pattern} => {rule['endpoint']}"
            if scope == "title" and pattern and pattern in title:
                return rule["endpoint"], f"title:{pattern} => {rule['endpoint']}"
        return default_endpoint, default_rule

    def load_proxy_pool(self) -> list[str]:
        path = self.proxy_pool_entry.get().strip()
        if not path:
            return []
        file_path = Path(path)
        if not file_path.exists():
            return []

        proxies = []
        seen = set()
        try:
            lines = self.read_text_lines(file_path)
        except UnicodeError:
            return []

        for line in lines:
            item = self.normalize_proxy_address(line.strip())
            if not item or item.startswith(("#", ";", "//")):
                continue
            if item in seen:
                continue
            seen.add(item)
            proxies.append(item)
        return proxies

    def get_proxy_health(self, proxy_addr: str) -> dict:
        state = self.proxy_health.setdefault(
            proxy_addr,
            {
                "fail_count": 0,
                "cooldown_until": 0.0,
                "last_error": "",
                "last_ok": 0.0,
            },
        )
        return state

    def is_proxy_in_cooldown(self, proxy_addr: str) -> bool:
        if not proxy_addr:
            return False
        state = self.get_proxy_health(proxy_addr)
        return state.get("cooldown_until", 0.0) > time.time()

    def mark_proxy_success(self, proxy_addr: str):
        if not proxy_addr or self.proxy_mode_var.get() != PROXY_MODE_POOL:
            return
        state = self.get_proxy_health(proxy_addr)
        state["fail_count"] = 0
        state["cooldown_until"] = 0.0
        state["last_error"] = ""
        state["last_ok"] = time.time()

    def mark_proxy_failure(self, proxy_addr: str, error_summary: str):
        if not proxy_addr or self.proxy_mode_var.get() != PROXY_MODE_POOL:
            return
        state = self.get_proxy_health(proxy_addr)
        state["fail_count"] = int(state.get("fail_count", 0)) + 1
        state["last_error"] = error_summary
        if state["fail_count"] >= self.get_proxy_fail_threshold():
            cooldown_until = time.time() + self.get_proxy_cooldown_seconds()
            state["cooldown_until"] = cooldown_until
            self.log_queue.put(
                f"[~] 代理进入冷却: {proxy_addr} | 失败次数={state['fail_count']} | "
                f"冷却 {self.get_proxy_cooldown_seconds()}s | 原因={error_summary}"
            )

    def get_proxy_recovery_eta(self, proxy_addr: str) -> int:
        if not proxy_addr:
            return 0
        state = self.get_proxy_health(proxy_addr)
        return max(0, int(state.get("cooldown_until", 0.0) - time.time()))

    def format_time_value(self, timestamp: float) -> str:
        if not timestamp:
            return "-"
        try:
            return datetime.fromtimestamp(timestamp).strftime("%H:%M:%S")
        except Exception:
            return "-"

    def format_duration(self, seconds: int) -> str:
        value = max(0, int(seconds))
        if value < 60:
            return f"{value}s"
        minutes, remain = divmod(value, 60)
        if minutes < 60:
            return f"{minutes}m{remain:02d}s"
        hours, remain_minutes = divmod(minutes, 60)
        return f"{hours}h{remain_minutes:02d}m"

    def get_proxy_assignment_count(self, proxy_addr: str) -> int:
        return sum(1 for assigned in self.proxy_assignment.values() if assigned == proxy_addr)

    def get_proxy_status_rows(self) -> list[dict]:
        rows = []
        if self.proxy_mode_var.get() == PROXY_MODE_POOL:
            proxies = self.load_proxy_pool()
        else:
            proxy = self.normalize_proxy_address(self.proxy_entry.get())
            proxies = [proxy] if proxy else []

        for proxy in proxies:
            state = self.get_proxy_health(proxy)
            cooldown_left = self.get_proxy_recovery_eta(proxy)
            status = "冷却中" if cooldown_left > 0 else "可用"
            rows.append(
                {
                    "proxy": proxy,
                    "status": status,
                    "fail_count": state.get("fail_count", 0),
                    "cooldown": self.format_duration(cooldown_left) if cooldown_left > 0 else "-",
                    "assigned": self.get_proxy_assignment_count(proxy),
                    "last_error": state.get("last_error", "") or "-",
                    "last_ok": self.format_time_value(state.get("last_ok", 0.0)),
                }
            )
        return rows

    def get_proxy_candidates_for_record(self, record: AuditRecord) -> list[str | None]:
        if not self.use_proxy_var.get():
            return [None]

        if self.proxy_mode_var.get() == PROXY_MODE_SINGLE:
            proxy = self.normalize_proxy_address(self.proxy_entry.get())
            return [proxy] if proxy else [None]

        pool = self.load_proxy_pool()
        if not pool:
            return [None]

        key = record.target
        assigned = self.proxy_assignment.get(key)
        if assigned not in pool:
            assigned = pool[self.proxy_round_robin_index % len(pool)]
            self.proxy_assignment[key] = assigned
            self.proxy_round_robin_index = (self.proxy_round_robin_index + 1) % max(1, len(pool))

        start_index = pool.index(assigned)
        ordered = pool[start_index:] + pool[:start_index]
        healthy = [proxy for proxy in ordered if not self.is_proxy_in_cooldown(proxy)]
        if healthy:
            return healthy[: self.get_proxy_retry_count()]

        fallback = sorted(
            ordered,
            key=lambda proxy: self.get_proxy_health(proxy).get("cooldown_until", 0.0),
        )
        if fallback:
            wait_seconds = self.get_proxy_recovery_eta(fallback[0])
            self.log_queue.put(
                f"[~] 代理池当前全部处于冷却，临时尝试最先恢复的代理: {fallback[0]} | 剩余约 {wait_seconds}s"
            )
        return fallback[: self.get_proxy_retry_count()] or [None]

    def validate_proxy_settings(self) -> str:
        if not self.use_proxy_var.get():
            return ""
        if self.proxy_mode_var.get() == PROXY_MODE_SINGLE:
            if not self.normalize_proxy_address(self.proxy_entry.get()):
                return "已启用代理，但未填写代理地址。"
            return ""
        pool = self.load_proxy_pool()
        if not self.proxy_pool_entry.get().strip():
            return "已启用代理池轮换，但未选择代理池文件。"
        if not pool:
            return "代理池文件为空、编码无法识别，或文件内容无有效代理。"
        return ""

    def should_retry_with_next_proxy(self, exc: Exception, attempt_index: int, total_attempts: int) -> bool:
        if attempt_index >= total_attempts:
            return False
        if not isinstance(exc, requests.exceptions.RequestException):
            return False
        summary, _detail = describe_request_exception(exc)
        if isinstance(exc, requests.exceptions.ProxyError):
            return True
        return summary in RETRYABLE_FAILURE_PREFIXES

    def toggle_brute_dict_state(self):
        brute_enabled = self.brute_var.get()
        custom_mode = self.dict_mode_var.get() == CUSTOM_BRUTE_DICT_MODE

        self.dict_mode_box.config(state="readonly" if brute_enabled else tk.DISABLED)
        entry_state = tk.NORMAL if brute_enabled and custom_mode else tk.DISABLED
        button_state = tk.NORMAL if brute_enabled and custom_mode else tk.DISABLED

        self.user_dict_entry.config(state=entry_state)
        self.pass_dict_entry.config(state=entry_state)
        self.user_dict_button.config(state=button_state)
        self.pass_dict_button.config(state=button_state)
        self.load_builtin_dict_button.config(state=button_state)
        self.open_builtin_dict_button.config(state=tk.NORMAL)

    def browse_user_dict(self):
        path = filedialog.askopenfilename(
            title="选择账号字典",
            filetypes=[("Text files", "*.txt *.dic *.lst *.csv"), ("All files", "*.*")],
        )
        if not path:
            return
        self.user_dict_entry.config(state=tk.NORMAL)
        self.user_dict_entry.delete(0, tk.END)
        self.user_dict_entry.insert(0, path)
        self.toggle_brute_dict_state()

    def browse_pass_dict(self):
        path = filedialog.askopenfilename(
            title="选择密码字典",
            filetypes=[("Text files", "*.txt *.dic *.lst *.csv"), ("All files", "*.*")],
        )
        if not path:
            return
        self.pass_dict_entry.config(state=tk.NORMAL)
        self.pass_dict_entry.delete(0, tk.END)
        self.pass_dict_entry.insert(0, path)
        self.toggle_brute_dict_state()

    def load_builtin_dict_files_into_form(self):
        self.ensure_builtin_dict_files()
        self.brute_var.set(True)
        self.dict_mode_var.set(CUSTOM_BRUTE_DICT_MODE)
        self.user_dict_entry.config(state=tk.NORMAL)
        self.user_dict_entry.delete(0, tk.END)
        self.user_dict_entry.insert(0, str(self.builtin_user_dict_path))
        self.pass_dict_entry.config(state=tk.NORMAL)
        self.pass_dict_entry.delete(0, tk.END)
        self.pass_dict_entry.insert(0, str(self.builtin_pass_dict_path))
        self.toggle_brute_dict_state()
        self.log_message(
            f"[*] 已载入内置字典文件: 账号={self.builtin_user_dict_path} | 密码={self.builtin_pass_dict_path}"
        )

    def open_builtin_dict_dir(self):
        self.ensure_builtin_dict_files()
        os.startfile(self.dicts_dir)

    def normalize_target(self, target: str) -> str:
        value = target.strip()
        if not value:
            return ""
        if not re.match(r"^https?://", value, flags=re.I):
            value = f"http://{value}"
        return value

    def collect_targets_from_source(self, path: Path) -> list[str]:
        seen = set()
        targets = []
        raw_targets = self.read_targets_from_file(path)
        for raw_line in raw_targets:
            target = self.normalize_target(raw_line)
            if not target or target in seen:
                continue
            seen.add(target)
            targets.append(target)
        return targets

    def sync_records_with_source(self, filepath: str) -> dict[str, int]:
        source_path = Path(filepath)
        targets = self.collect_targets_from_source(source_path)
        existing_by_target = {record.target: record for record in self.all_records}
        existing_targets = set(existing_by_target)
        next_record_id = max((record.record_id for record in self.all_records), default=0) + 1

        synced_records = []
        added = 0
        kept = 0
        for target in targets:
            record = existing_by_target.get(target)
            if record is None:
                record = AuditRecord(record_id=next_record_id, target=target)
                next_record_id += 1
                added += 1
            else:
                kept += 1
            synced_records.append(record)

        removed = len(existing_targets - set(targets))
        self.all_records = synced_records
        return {
            "total": len(synced_records),
            "added": added,
            "removed": removed,
            "kept": kept,
        }

    def load_file(self):
        filepath = filedialog.askopenfilename(
            filetypes=[
                ("Supported files", "*.txt *.xlsx"),
                ("Text files", "*.txt"),
                ("Excel files", "*.xlsx"),
            ]
        )
        if not filepath:
            return

        project_path = self.project_path_for_source(filepath)
        if project_path.exists():
            self.load_project_file(project_path, sync_source=True)
            self.log_message(f"[*] 已载入历史工程进度: {filepath}")
            return

        self.file_entry.delete(0, tk.END)
        self.file_entry.insert(0, filepath)
        self.last_loaded_path = filepath
        self.current_project_path = project_path
        self.set_detail_log_path(self.detail_log_path_for_project(project_path))
        self.records_by_item.clear()
        self.item_id_by_record_key.clear()
        self.all_records.clear()
        self.proxy_assignment.clear()
        self.proxy_round_robin_index = 0
        self.proxy_health.clear()
        for item in self.tree.get_children():
            self.tree.delete(item)

        try:
            sync_stats = self.sync_records_with_source(filepath)
            self.rebuild_tree()
            self.ensure_selection()
            self.save_progress_snapshot(reason="load")
            self.mark_source_watch_baseline(filepath)
            self.log_message(f"[*] 加载文件: {filepath}")
            self.log_message(
                f"[*] 成功导入 {sync_stats['total']} 个目标资产，已自动去重与协议补全。"
            )
        except Exception as exc:
            self.log_message(f"[-] 读取文件失败: {exc}")
            messagebox.showerror("错误", f"读取文件失败: {exc}")

    def reload_current_source_file(self):
        if self.is_scanning:
            messagebox.showwarning("提示", "当前正在扫描，请先停止后再刷新源文件。")
            return
        filepath = self.file_entry.get().strip() or self.last_loaded_path
        if not filepath:
            messagebox.showwarning("提示", "请先加载目标文件。")
            return
        source_path = Path(filepath)
        if not source_path.exists():
            messagebox.showwarning("提示", f"源文件不存在:\n{source_path}")
            return

        try:
            self.last_loaded_path = str(source_path)
            if not self.current_project_path:
                self.current_project_path = self.project_path_for_source(str(source_path))
            stats = self.sync_records_with_source(str(source_path))
            self.rebuild_tree()
            self.ensure_selection()
            self.save_progress_snapshot(reason="source-refresh")
            self.mark_source_watch_baseline(source_path)
            self.log_message(
                f"[*] 已按源文件刷新列表: 总计 {stats['total']} | 新增 {stats['added']} | 删除 {stats['removed']} | 保留 {stats['kept']}"
            )
        except Exception as exc:
            self.log_message(f"[-] 刷新源文件失败: {exc}")
            messagebox.showerror("错误", f"刷新源文件失败: {exc}")

    def read_targets_from_file(self, path: Path) -> list[str]:
        if path.suffix.lower() == ".xlsx":
            return self.read_targets_from_xlsx(path)
        with open(path, "r", encoding="utf-8") as handle:
            return [line.strip() for line in handle]

    def read_targets_from_xlsx(self, path: Path) -> list[str]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        targets = []
        header_aliases = {
            "host": "host",
            "网址": "host",
            "地址": "host",
            "资产": "host",
            "目标": "host",
            "url": "host",
            "site": "host",
            "ip": "ip",
            "端口": "port",
            "协议": "scheme",
        }
        header_row_found = False
        header_map = {}

        for row in sheet.iter_rows(values_only=True):
            values = ["" if cell is None else str(cell).strip() for cell in row]
            if not any(values):
                continue

            normalized_row = [value.lower() for value in values]
            current_map = {}
            for idx, value in enumerate(normalized_row):
                if value in header_aliases:
                    current_map[header_aliases[value]] = idx

            if not header_row_found:
                if "host" in current_map or {"ip", "port", "scheme"}.issubset(current_map):
                    header_row_found = True
                    header_map = current_map
                continue

            target = self.extract_target_from_row(values, header_map)
            if target:
                targets.append(target)
        workbook.close()
        return targets

    def extract_target_from_row(self, values: list[str], header_map: dict[str, int]) -> str:
        host = self.get_cell(values, header_map.get("host"))
        if host:
            return host

        ip = self.get_cell(values, header_map.get("ip"))
        port = self.get_cell(values, header_map.get("port"))
        scheme = self.get_cell(values, header_map.get("scheme")) or "http"

        if not ip:
            return ""

        port = self.normalize_port(port)
        if port:
            return f"{scheme}://{ip}:{port}"
        return f"{scheme}://{ip}"

    def get_cell(self, values: list[str], index: int | None) -> str:
        if index is None or index >= len(values):
            return ""
        return str(values[index]).strip()

    def normalize_port(self, value: str) -> str:
        if not value:
            return ""
        port = value.strip()
        if re.fullmatch(r"\d+\.0", port):
            return port[:-2]
        return port

    def rebuild_tree(self):
        self.records_by_item.clear()
        self.item_id_by_record_key.clear()
        for item in self.tree.get_children():
            self.tree.delete(item)
        for record in self.filtered_records():
            item_id = self.tree.insert(
                "",
                tk.END,
                values=(
                    record.record_id,
                    record.target,
                    record.status,
                    record.title or "-",
                    record.risk_level,
                    record.result,
                ),
                tags=(self.tag_for_record(record),),
            )
            self.records_by_item[item_id] = record
            self.item_id_by_record_key[record.record_id] = item_id
        self.update_summary()

    def filtered_records(self) -> list[AuditRecord]:
        current_filter = self.filter_var.get()
        if current_filter == "仅高风险":
            return [record for record in self.all_records if record.risk_level == "高"]
        if current_filter == "仅疑似登录页":
            return [record for record in self.all_records if record.login_form]
        if current_filter == "仅已完成":
            return [record for record in self.all_records if record.status == "已完成"]
        if current_filter == "仅失败":
            return [record for record in self.all_records if record.risk_level == "失败"]
        if current_filter == "仅有截图":
            return [record for record in self.all_records if record.screenshot_path]
        if current_filter == "仅有表单字段":
            return [record for record in self.all_records if record.field_summary]
        return list(self.all_records)

    def apply_filter(self):
        self.rebuild_tree()
        self.ensure_selection()
        self.log_message(f"[*] 当前列表筛选: {self.filter_var.get()}")

    def save_results(self):
        records = self.get_records_snapshot()
        if not records:
            messagebox.showwarning("提示", "没有可导出的数据。")
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[
                ("CSV files", "*.csv"),
                ("JSON files", "*.json"),
                ("HTML files", "*.html"),
            ],
            title="保存审计结果",
        )
        if not filepath:
            return

        output = Path(filepath)
        try:
            if output.suffix.lower() == ".json":
                self.export_json(output, records)
            elif output.suffix.lower() == ".html":
                self.export_html(output, records)
            else:
                self.export_csv(output, records)
            self.log_message(f"[*] 审计结果已导出至: {output}")
            messagebox.showinfo("成功", f"导出成功: {output.name}")
        except Exception as exc:
            self.log_message(f"[-] 导出失败: {exc}")
            messagebox.showerror("错误", f"导出失败: {exc}")

    def export_csv(self, output: Path, records: list[AuditRecord]):
        with open(output, "w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.writer(handle)
            writer.writerow(
                [
                    "ID",
                    "目标资产",
                    "最终URL",
                    "状态",
                    "页面标题",
                    "风险级别",
                    "登录评分",
                    "审计结果",
                    "是否疑似登录页",
                    "密码框数量",
                    "验证码",
                    "多因素",
                    "锁定提示",
                    "默认账号提示",
                    "表单Action",
                    "表单Method",
                    "代理出口",
                    "OCR接口",
                    "OCR命中规则",
                    "字段摘要",
                    "截图路径",
                    "错误信息",
                ]
            )
            for record in records:
                writer.writerow(
                    [
                        record.record_id,
                        record.target,
                        record.final_url,
                        record.status,
                        record.title,
                        record.risk_level,
                        record.login_score,
                        record.result,
                        "是" if record.login_form else "否",
                        record.password_field_count,
                        "是" if record.captcha_present else "否",
                        "是" if record.mfa_present else "否",
                        "是" if record.lockout_hint else "否",
                        "是" if record.default_hint else "否",
                        record.form_action,
                        record.form_method,
                        record.proxy_used,
                        record.ocr_endpoint_used,
                        record.ocr_route_rule,
                        record.field_summary,
                        record.screenshot_path,
                        record.error,
                    ]
                )

    def export_json(self, output: Path, records: list[AuditRecord]):
        payload = {
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "record_count": len(records),
            "profile": self.get_profile_dashboard_data(records),
            "records": [asdict(record) for record in records],
        }
        with open(output, "w", encoding="utf-8") as handle:
            json.dump(payload, handle, ensure_ascii=False, indent=2)

    def export_html(self, output: Path, records: list[AuditRecord]):
        profile = self.get_profile_dashboard_data(records)
        rows = []
        for record in records:
            screenshot = self.render_html_screenshot(record)
            rows.append(
                "<tr>"
                f"<td>{record.record_id}</td>"
                f"<td>{html.escape(record.target)}</td>"
                f"<td>{html.escape(record.status)}</td>"
                f"<td>{html.escape(record.title or '-')}</td>"
                f"<td>{html.escape(record.risk_level)}</td>"
                f"<td>{record.login_score}</td>"
                f"<td>{html.escape(record.result)}</td>"
                f"<td>{html.escape(record.proxy_used or '-')}</td>"
                f"<td>{html.escape(record.ocr_endpoint_used or '-')}</td>"
                f"<td>{html.escape(record.ocr_route_rule or '-')}</td>"
                f"<td>{html.escape(record.form_method or '-')}</td>"
                f"<td>{html.escape(record.field_summary or '-')}</td>"
                f"<td>{screenshot}</td>"
                "</tr>"
            )

        profile_cards = []
        for item in profile["summary_items"]:
            profile_cards.append(
                "<div class='metric'>"
                f"<div class='metric-label'>{html.escape(item['label'])}</div>"
                f"<div class='metric-value'>{html.escape(str(item['value']))}</div>"
                "</div>"
            )

        profile_sections = [
            self.render_profile_table_html("站点画像分布", profile["profile_tags"], ["画像标签", "数量"]),
            self.render_profile_table_html("OCR规则命中", profile["ocr_rules"], ["规则", "数量"]),
            self.render_profile_table_html("代理出口分布", profile["proxy_usage"], ["代理出口", "数量"]),
            self.render_profile_table_html("失败原因分布", profile["failures"], ["失败类型", "数量"]),
        ]

        page = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<title>登录面审计报告</title>
<style>
body {{ font-family: "Microsoft YaHei", sans-serif; margin: 24px; background: #f5f7fb; color: #1f2937; }}
h1 {{ margin-bottom: 8px; }}
p {{ color: #4b5563; }}
.metrics {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(120px, 1fr)); gap: 12px; margin: 18px 0 24px; }}
.metric {{ background: white; border: 1px solid #dbeafe; border-radius: 10px; padding: 12px; box-shadow: 0 6px 18px rgba(15,23,42,.05); }}
.metric-label {{ color: #64748b; font-size: 13px; }}
.metric-value {{ color: #0f172a; font-size: 24px; font-weight: 700; margin-top: 6px; }}
.profile-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(300px, 1fr)); gap: 16px; margin-bottom: 24px; }}
.profile-card {{ background: white; border: 1px solid #dbeafe; border-radius: 12px; padding: 14px; box-shadow: 0 6px 20px rgba(15,23,42,.05); }}
.profile-card h2 {{ margin: 0 0 10px; font-size: 17px; }}
table {{ width: 100%; border-collapse: separate; border-spacing: 0; background: white; border: 1px solid #94a3b8; border-radius: 10px; overflow: hidden; }}
th, td {{ padding: 12px 14px; text-align: left; vertical-align: top; border-right: 1px solid #b8c4d2; border-bottom: 1px solid #b8c4d2; }}
th:last-child, td:last-child {{ border-right: none; }}
tbody tr:last-child td {{ border-bottom: none; }}
tbody tr:nth-child(even) {{ background: #f8fbff; }}
th {{ background: #e8f1fb; }}
a {{ color: #1d4ed8; text-decoration: none; }}
img {{ max-width: 280px; max-height: 180px; border: 1px solid #b8c4d2; border-radius: 6px; display: block; margin-bottom: 6px; }}
.high {{ color: #b91c1c; font-weight: bold; }}
.medium {{ color: #b45309; font-weight: bold; }}
.low {{ color: #166534; font-weight: bold; }}
</style>
</head>
<body>
<h1>登录面审计报告</h1>
<p>生成时间: {html.escape(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))}</p>
<p>记录总数: {len(records)}</p>
<section class="metrics">{''.join(profile_cards)}</section>
<section class="profile-grid">{''.join(profile_sections)}</section>
<table>
<thead>
<tr><th>ID</th><th>目标资产</th><th>状态</th><th>页面标题</th><th>风险级别</th><th>登录评分</th><th>审计结果</th><th>代理出口</th><th>OCR接口</th><th>OCR规则</th><th>表单方法</th><th>字段摘要</th><th>截图</th></tr>
</thead>
<tbody>
{''.join(rows)}
</tbody>
</table>
</body>
</html>"""
        with open(output, "w", encoding="utf-8") as handle:
            handle.write(page)

    def export_evidence_page(self):
        records = [record for record in self.get_records_snapshot() if record.screenshot_path or record.field_summary]
        if not records:
            messagebox.showinfo("提示", "当前没有可导出的证据记录。")
            return
        output = filedialog.asksaveasfilename(
            defaultextension=".html",
            filetypes=[("HTML files", "*.html")],
            title="导出证据页",
        )
        if not output:
            return
        self.export_evidence_html(Path(output), records)
        self.log_message(f"[*] 证据页已导出至: {output}")
        messagebox.showinfo("成功", "证据页导出成功。")

    def export_evidence_html(self, output: Path, records: list[AuditRecord]):
        profile = self.get_profile_dashboard_data(records)
        cards = []
        for record in records:
            screenshot = self.render_html_screenshot(record)
            cards.append(
                "<section class='card'>"
                f"<h2>{html.escape(record.title or record.target)}</h2>"
                f"<p><strong>目标:</strong> {html.escape(record.target)}</p>"
                f"<p><strong>最终URL:</strong> {html.escape(record.final_url or '-')}</p>"
                f"<p><strong>状态:</strong> {html.escape(record.status)} | <strong>风险:</strong> {html.escape(record.risk_level)}</p>"
                f"<p><strong>代理出口:</strong> {html.escape(record.proxy_used or '-')}</p>"
                f"<p><strong>OCR接口:</strong> {html.escape(record.ocr_endpoint_used or '-')}</p>"
                f"<p><strong>OCR规则:</strong> {html.escape(record.ocr_route_rule or '-')}</p>"
                f"<p><strong>表单方法:</strong> {html.escape(record.form_method or '-')}</p>"
                f"<p><strong>表单Action:</strong> {html.escape(record.form_action or '-')}</p>"
                f"<p><strong>字段摘要:</strong> {html.escape(record.field_summary or '-')}</p>"
                f"<p><strong>审计结果:</strong> {html.escape(record.result)}</p>"
                f"<div class='shot'>{screenshot}</div>"
                "</section>"
            )
        page = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<title>登录面证据页</title>
<style>
body {{ font-family: "Microsoft YaHei", sans-serif; margin: 24px; background: #f8fafc; color: #0f172a; }}
h1 {{ margin-bottom: 8px; }}
.grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(360px, 1fr)); gap: 16px; }}
.card {{ background: white; border: 1px solid #dbeafe; border-radius: 12px; padding: 16px; box-shadow: 0 8px 24px rgba(15,23,42,.06); }}
.card h2 {{ margin-top: 0; font-size: 18px; }}
.card p {{ margin: 8px 0; line-height: 1.5; word-break: break-all; }}
img {{ max-width: 100%; border-radius: 8px; border: 1px solid #cbd5e1; }}
a {{ color: #2563eb; text-decoration: none; }}
</style>
</head>
<body>
<h1>登录面证据页</h1>
<p>生成时间: {html.escape(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))}</p>
<p>画像摘要: {' | '.join(f"{item['label']} {item['value']}" for item in profile['summary_items'][:6])}</p>
<div class="grid">
{''.join(cards)}
</div>
</body>
</html>"""
        with open(output, "w", encoding="utf-8") as handle:
            handle.write(page)

    def extract_primary_result(self, text: str) -> str:
        items = self.split_pipe_items(text)
        return items[0] if items else "-"

    def get_record_profile_tags(self, record: AuditRecord) -> list[str]:
        tags = []
        result_text = record.result or ""
        if record.risk_level == "失败":
            tags.append("访问失败")
        if record.login_form:
            tags.append("登录页")
            if record.captcha_present:
                tags.append("验证码登录页")
            else:
                tags.append("无验证码登录页")
            if record.mfa_present:
                tags.append("多因素登录页")
            if record.lockout_hint:
                tags.append("有锁定提示")
            else:
                tags.append("无锁定提示")
            if record.default_hint:
                tags.append("默认口令提示")
        else:
            tags.append("非登录页")

        if "浏览器渲染补扫命中" in result_text:
            tags.append("浏览器补扫命中")
        if "浏览器渲染补扫未命中" in result_text:
            tags.append("浏览器补扫未命中")
        if record.ocr_endpoint_used:
            tags.append("已使用OCR")
        if BRUTE_FORCE_SUCCESS_PREFIX in result_text:
            tags.append("弱口令命中")
        return tags

    def counter_to_rows(self, counter: Counter, top_n: int = 10) -> list[dict]:
        rows = []
        for key, value in counter.most_common(top_n):
            rows.append({"name": key, "count": value})
        return rows

    def get_profile_dashboard_data(self, records: list[AuditRecord] | None = None) -> dict:
        target_records = list(records if records is not None else self.all_records)
        total = len(target_records)
        completed = sum(1 for record in target_records if record.status == "已完成")
        login_pages = sum(1 for record in target_records if record.login_form)
        captcha_pages = sum(1 for record in target_records if record.captcha_present)
        browser_hits = sum(1 for record in target_records if "浏览器渲染补扫命中" in (record.result or ""))
        ocr_used = sum(1 for record in target_records if record.ocr_endpoint_used)
        weak_hits = sum(1 for record in target_records if BRUTE_FORCE_SUCCESS_PREFIX in (record.result or ""))
        failed = sum(1 for record in target_records if record.risk_level == "失败")

        profile_tags = Counter()
        ocr_rules = Counter()
        proxy_usage = Counter()
        failures = Counter()

        for record in target_records:
            profile_tags.update(self.get_record_profile_tags(record))
            proxy_usage[record.proxy_used or "直连"] += 1
            if record.ocr_route_rule:
                ocr_rules[record.ocr_route_rule] += 1
            if record.risk_level == "失败":
                failures[self.extract_primary_result(record.result)] += 1

        return {
            "summary_items": [
                {"label": "总目标", "value": total},
                {"label": "已完成", "value": completed},
                {"label": "登录页", "value": login_pages},
                {"label": "验证码页", "value": captcha_pages},
                {"label": "补扫命中", "value": browser_hits},
                {"label": "已用OCR", "value": ocr_used},
                {"label": "弱口令命中", "value": weak_hits},
                {"label": "访问失败", "value": failed},
            ],
            "profile_tags": self.counter_to_rows(profile_tags, top_n=12),
            "ocr_rules": self.counter_to_rows(ocr_rules, top_n=12),
            "proxy_usage": self.counter_to_rows(proxy_usage, top_n=12),
            "failures": self.counter_to_rows(failures, top_n=12),
        }

    def render_profile_table_html(self, title: str, rows: list[dict], headers: list[str]) -> str:
        body_rows = []
        for row in rows or [{"name": "-", "count": 0}]:
            body_rows.append(
                "<tr>"
                f"<td>{html.escape(str(row.get('name', '-')))}</td>"
                f"<td>{html.escape(str(row.get('count', 0)))}</td>"
                "</tr>"
            )
        return (
            "<section class='profile-card'>"
            f"<h2>{html.escape(title)}</h2>"
            "<table>"
            "<thead>"
            f"<tr><th>{html.escape(headers[0])}</th><th>{html.escape(headers[1])}</th></tr>"
            "</thead>"
            "<tbody>"
            + "".join(body_rows)
            + "</tbody></table></section>"
        )

    def build_requests_session(self, proxy_addr: str | None = None) -> requests.Session:
        session = requests.Session()
        session.trust_env = False
        session.headers.update(
            {
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
                )
            }
        )
        proxy_value = self.normalize_proxy_address(proxy_addr or "")
        if not proxy_value and self.use_proxy_var.get() and self.proxy_mode_var.get() == PROXY_MODE_SINGLE:
            proxy_value = self.normalize_proxy_address(self.proxy_entry.get())
        if proxy_value:
            session.proxies = {
                "http": proxy_value,
                "https": proxy_value,
            }
        return session

    def get_worker_count(self) -> int:
        try:
            return max(1, min(12, int(self.worker_var.get())))
        except Exception:
            return 4

    def get_capture_delay(self) -> float:
        try:
            return max(0.0, min(3.0, float(self.capture_delay_var.get())))
        except Exception:
            return 0.4

    def get_render_wait(self) -> float:
        try:
            return max(0.8, min(10.0, float(self.render_wait_var.get())))
        except Exception:
            return 2.5

    def reset_record_for_rescan(self, record: AuditRecord):
        fresh = AuditRecord(record_id=record.record_id, target=record.target)
        for field_name, value in fresh.__dict__.items():
            setattr(record, field_name, value)

    def get_retryable_failure_records(self) -> list[AuditRecord]:
        return [
            record
            for record in self.all_records
            if is_retryable_failure_result(record.risk_level, record.result)
        ]

    def restart_scan(self):
        if not self.all_records:
            messagebox.showwarning("提示", "请先导入目标文件。")
            return
        if self.is_scanning:
            messagebox.showwarning("提示", "当前正在扫描，请先停止后再重新扫描。")
            return

        confirmed = messagebox.askyesno(
            "重新扫描",
            "将清空当前工程的扫描结果并从头重新扫描全部目标。\n\n是否继续？",
        )
        if not confirmed:
            return

        with self.result_lock:
            for record in self.all_records:
                self.reset_record_for_rescan(record)

        self.rebuild_tree()
        self.ensure_selection()
        self.save_progress_snapshot(reason="rescan-reset")
        self.log_message(f"[*] 已重置 {len(self.all_records)} 条记录，准备重新扫描。")
        self.toggle_scan()

    def retry_failed_targets(self):
        if not self.all_records:
            messagebox.showwarning("提示", "请先导入目标文件。")
            return
        if self.is_scanning:
            messagebox.showwarning("提示", "当前正在扫描，请先停止后再重扫失败项。")
            return

        retry_records = self.get_retryable_failure_records()
        if not retry_records:
            messagebox.showinfo("提示", "当前没有可重扫的失败项。")
            return

        categories = []
        seen_categories = set()
        for record in retry_records:
            summary = (record.result or "").strip()
            if summary and summary not in seen_categories:
                seen_categories.add(summary)
                categories.append(summary)
            if len(categories) >= 3:
                break

        message = [f"将重扫 {len(retry_records)} 条失败目标。", "", "仅会重置以下可重试失败项："]
        for item in categories:
            message.append(f"- {item}")
        if len(retry_records) > len(categories):
            message.append("- 其余同类失败项")
        message.extend(["", "是否继续？"])

        confirmed = messagebox.askyesno("重扫失败项", "\n".join(message))
        if not confirmed:
            return

        with self.result_lock:
            for record in retry_records:
                self.reset_record_for_rescan(record)

        self.rebuild_tree()
        self.ensure_selection()
        self.save_progress_snapshot(reason="retry-failed-reset")
        self.log_message(f"[*] 已重置 {len(retry_records)} 条失败目标，准备重扫失败项。")
        self.toggle_scan()

    def toggle_scan(self):
        if not self.all_records:
            messagebox.showwarning("提示", "请先导入目标文件。")
            return

        if not self.is_scanning:
            proxy_error = self.validate_proxy_settings()
            if proxy_error:
                messagebox.showwarning("提示", proxy_error)
                return
            self.is_scanning = True
            self.btn_start.config(text="停止审计", bg="#ff8d7b")
            self.btn_retry_failed.config(state=tk.DISABLED)
            self.btn_rescan.config(state=tk.DISABLED)
            self.disable_proxy_controls_for_scan()
            self.save_progress_snapshot(reason="scan-start")
            self.log_message("[*] 初始化登录面审计引擎...")
            self.log_message(f"[*] 当前识别模式: {self.mode_var.get()}")
            self.log_message(f"[*] 当前并发数: {self.get_worker_count()}")
            self.log_message(
                f"[*] 截图策略: {self.capture_policy_var.get()} | 截图节流: {self.get_capture_delay():.1f}s"
            )
            self.log_message(
                f"[*] 浏览器补扫识别: {'启用' if self.browser_render_var.get() else '关闭'} | 渲染等待: {self.get_render_wait():.1f}s"
            )
            if not self.browser_render_var.get():
                self.log_message("[~] 当前已关闭浏览器补扫识别，Vue/React/前端渲染登录页可能识别不到登录框。")
            self.log_message(
                f"[*] 验证码OCR: {'启用' if self.captcha_ocr_var.get() else '关闭'} | 接口: {self.ocr_endpoint_entry.get().strip() or DEFAULT_OCR_ENDPOINT}"
            )
            if self.capture_var.get() and self.capture_policy_var.get() == CAPTURE_POLICY_HIGH:
                self.log_message("[~] 当前截图策略为“仅高风险”，低/中风险或未命中的目标不会截图。")
            if self.ocr_route_entry.get().strip():
                self.log_message(f"[*] OCR映射文件: {self.ocr_route_entry.get().strip()}")
            if self.brute_var.get():
                if self.dict_mode_var.get() == CUSTOM_BRUTE_DICT_MODE:
                    self.log_message(
                        "[*] 弱口令检测: 已启用 | "
                        f"账号字典={self.user_dict_entry.get().strip() or '未选择'} | "
                        f"密码字典={self.pass_dict_entry.get().strip() or '未选择'}"
                    )
                else:
                    self.log_message("[*] 弱口令检测: 已启用 | 字典=默认中文常用账号/密码")
            else:
                self.log_message("[*] 弱口令检测: 未启用")
            self.log_message(f"[*] 本次将从未完成目标继续扫描，已完成记录会自动跳过。")
            if self.use_proxy_var.get():
                if self.proxy_mode_var.get() == PROXY_MODE_POOL:
                    self.log_message(
                        f"[*] 已启用代理池轮换: 文件={self.proxy_pool_entry.get().strip() or '-'} | "
                        f"失败重试={self.get_proxy_retry_count()} 次 | "
                        f"熔断阈值={self.get_proxy_fail_threshold()} | 冷却={self.get_proxy_cooldown_seconds()}s"
                    )
                else:
                    self.log_message(
                        f"[*] 已启用代理: {self.normalize_proxy_address(self.proxy_entry.get()) or '-'} | "
                        f"模式={self.proxy_mode_var.get()}"
                    )
            threading.Thread(target=self.scan_engine, daemon=True).start()
        else:
            self.is_scanning = False
            self.save_progress_snapshot(reason="manual-stop")
            self.log_message("[-] 用户终止扫描任务。")

    def scan_engine(self):
        run_dir = self.output_dir / datetime.now().strftime("%Y%m%d_%H%M%S")
        run_dir.mkdir(parents=True, exist_ok=True)
        self.start_detail_log_session(run_dir)
        self.log_queue.put(f"[*] 详细日志: {self.detail_log_path}")
        screenshot_dir = run_dir / "screenshots"
        screenshot_dir.mkdir(exist_ok=True)
        pending_records = []
        for record in self.all_records:
            if record.status == "已完成":
                continue
            if record.status == "扫描中...":
                record.status = "等待中"
            pending_records.append(record)

        if not pending_records:
            self.log_queue.put("[*] 当前工程没有待处理目标。")
            self.save_progress_snapshot(reason="scan-finished")
            self.root.after(0, self.reset_scan_button)
            return

        executor = concurrent.futures.ThreadPoolExecutor(
            max_workers=self.get_worker_count(),
            thread_name_prefix="audit",
        )
        future_map = {}
        try:
            for record in pending_records:
                if not self.is_scanning:
                    break
                item_id = self.find_item_id_by_record(record)
                if item_id:
                    self.root.after(0, self.update_tree_item, item_id, "扫描中...", record.title or "-", record.risk_level, "正在分析")
                record.status = "扫描中..."
                self.log_queue.put(f"[>] 正在分析: {record.target}")
                future = executor.submit(self.inspect_target_threadsafe, record, screenshot_dir)
                future_map[future] = record

            while future_map:
                done, _pending = concurrent.futures.wait(
                    future_map.keys(),
                    timeout=0.2,
                    return_when=concurrent.futures.FIRST_COMPLETED,
                )
                if not self.is_scanning:
                    for future in list(future_map):
                        future.cancel()
                    break
                for future in done:
                    record = future_map.pop(future)
                    item_id = self.find_item_id_by_record(record)
                    try:
                        analyzed = future.result()
                        if item_id:
                            self.root.after(
                                0,
                                self.update_tree_item,
                                item_id,
                                analyzed.status,
                                analyzed.title or "-",
                                analyzed.risk_level,
                                analyzed.result,
                            )
                        else:
                            self.root.after(0, self.apply_filter)
                        self.schedule_progress_snapshot(reason="record-updated")
                    except requests.exceptions.ProxyError:
                        record.status = "已完成"
                        record.risk_level = "失败"
                        record.result = "代理连接失败，请检查代理配置或代理服务状态"
                        record.error = "ProxyError: 代理连接失败"
                        if item_id:
                            self.root.after(0, self.update_tree_item, item_id, record.status, "-", record.risk_level, record.result)
                        self.log_queue.put("[-] 代理连接失败，请检查代理服务。")
                        self.schedule_progress_snapshot(reason="proxy-error")
                        self.is_scanning = False
                        for other_future in list(future_map):
                            other_future.cancel()
                        future_map.clear()
                        break
                    except Exception as exc:
                        result_summary, error_text = describe_request_exception(exc)
                        record.status = "已完成"
                        record.risk_level = "失败"
                        record.result = result_summary
                        record.error = error_text
                        if item_id:
                            self.root.after(0, self.update_tree_item, item_id, record.status, "-", record.risk_level, record.result)
                        self.log_queue.put(f"[-] {record.target} {result_summary}: {error_text}")
                        self.schedule_progress_snapshot(reason="record-error")
        finally:
            executor.shutdown(wait=False, cancel_futures=True)

        if self.capture_var.get():
            capture_records = pending_records if self.is_scanning else [record for record in pending_records if record.status == "已完成"]
            if capture_records:
                if not self.is_scanning:
                    self.log_queue.put("[*] 已停止主扫描，正在为已完成结果执行截图阶段。")
                self.capture_stage(capture_records, screenshot_dir, allow_when_stopped=not self.is_scanning)

        if self.is_scanning:
            self.log_queue.put("[*] 全部目标分析完成。")
        self.save_progress_snapshot(reason="scan-finished")
        self.root.after(0, self.reset_scan_button)

    def inspect_target_threadsafe(self, record: AuditRecord, screenshot_dir: Path) -> AuditRecord:
        candidates = self.get_proxy_candidates_for_record(record)
        last_exc = None
        for index, proxy_addr in enumerate(candidates, start=1):
            session = self.build_requests_session(proxy_addr)
            record.proxy_used = proxy_addr or ""
            try:
                analyzed = self.inspect_target(session, record, screenshot_dir)
                self.mark_proxy_success(proxy_addr or "")
                return analyzed
            except Exception as exc:
                last_exc = exc
                summary, _detail = describe_request_exception(exc)
                self.mark_proxy_failure(proxy_addr or "", summary)
                if self.should_retry_with_next_proxy(exc, index, len(candidates)):
                    target_proxy = proxy_addr or "直连"
                    self.log_queue.put(
                        f"[~] {record.target} 代理出口失败，准备切换下一个代理: {target_proxy} -> {compact_exception_message(exc)}"
                    )
                    continue
                raise
        if last_exc:
            raise last_exc
        raise RuntimeError("未获取到可用代理出口")

    def analyze_record_from_html(self, record: AuditRecord, page_url: str, html_text: str) -> BeautifulSoup:
        soup = BeautifulSoup(html_text or "", "html.parser")
        record.final_url = page_url
        record.status = "已完成"
        record.title = self.extract_title(soup)

        text_blob = self.collect_text_blob(soup)
        record.password_field_count = self.estimate_password_field_count(soup)
        record.captcha_present = self.contains_any(text_blob, CAPTCHA_KEYWORDS) or self.has_captcha_widget(soup)
        record.mfa_present = self.contains_any(text_blob, MFA_KEYWORDS)
        record.lockout_hint = self.contains_any(text_blob, LOCKOUT_KEYWORDS)
        record.default_hint = self.contains_any(text_blob, DEFAULT_HINT_KEYWORDS)
        record.login_score = self.compute_login_score(soup, text_blob, record.title)
        record.login_form = record.password_field_count > 0 or record.login_score >= 3
        record.form_action, record.form_method, record.field_summary = self.extract_form_details(soup)
        actionable_login = self.is_actionable_login_record(record)

        findings = []
        if record.login_form:
            findings.append("疑似登录页")
        if record.default_hint:
            findings.append("存在默认账号/初始密码提示")
        if not record.captcha_present and actionable_login:
            findings.append("未见验证码")
        if not record.mfa_present and actionable_login:
            findings.append("未见多因素认证提示")
        if not record.lockout_hint and actionable_login:
            findings.append("未见锁定策略提示")

        record.risk_level = self.calculate_risk(record)
        record.result = " | ".join(findings[:5]) if findings else "未识别到登录相关特征"
        self.log_record_trace(
            record,
            "静态分析 | "
            f"url={record.final_url or record.target} | title={record.title or '-'} | score={record.login_score} | "
            f"password_count={record.password_field_count} | field_summary={record.field_summary or '-'} | "
            f"login_form={record.login_form} | actionable={actionable_login} | captcha={record.captcha_present} | "
            f"mfa={record.mfa_present} | lockout={record.lockout_hint} | default_hint={record.default_hint}",
        )
        return soup

    def should_use_browser_render_fallback(self, record: AuditRecord, html_text: str) -> bool:
        if not self.browser_render_var.get():
            return False
        if self.is_actionable_login_record(record) or record.captcha_present:
            return False

        normalized_html = (html_text or "").lower()
        visible_text = re.sub(r"\s+", " ", BeautifulSoup(html_text or "", "html.parser").get_text(" ", strip=True))
        visible_text_len = len(visible_text)
        script_count = normalized_html.count("<script")
        has_spa_hint = any(token in normalized_html for token in BROWSER_RENDER_HINTS)
        has_login_keyword = any(keyword.lower() in normalized_html for keyword in (LOGIN_KEYWORDS + INPUT_HINTS + CAPTCHA_KEYWORDS))

        if has_spa_hint and visible_text_len < 260:
            return True
        if has_login_keyword and visible_text_len < 200:
            return True
        if script_count >= 4 and visible_text_len < 120:
            return True
        return False

    def init_headless_driver(self, proxy_addr: str | None = None):
        try:
            from selenium import webdriver
            from selenium.webdriver.chrome.options import Options as ChromeOptions
            from selenium.webdriver.chrome.service import Service as ChromeService
            from selenium.webdriver.edge.options import Options as EdgeOptions
            from selenium.webdriver.edge.service import Service as EdgeService
        except Exception:
            return None

        proxy_value = self.normalize_proxy_address(proxy_addr or "")
        if not proxy_value and self.use_proxy_var.get() and self.proxy_mode_var.get() == PROXY_MODE_SINGLE:
            proxy_value = self.normalize_proxy_address(self.proxy_entry.get())

        common_args = [
            "--headless=new",
            "--ignore-certificate-errors",
            "--disable-gpu",
            "--disable-popup-blocking",
            "--disable-notifications",
            "--disable-infobars",
            "--disable-blink-features=AutomationControlled",
            "--disable-features=MediaRouter,OptimizationHints,NotificationTriggers,PermissionChip,BlockInsecurePrivateNetworkRequests,PrivateNetworkAccessSendPreflights,PrivateNetworkAccessPermissionPrompt",
            "--window-size=1440,1080",
            "--log-level=3",
        ]
        prefs = {
            "credentials_enable_service": False,
            "profile.password_manager_enabled": False,
            "profile.default_content_setting_values.notifications": 2,
            "profile.default_content_setting_values.geolocation": 2,
            "profile.default_content_setting_values.media_stream_mic": 2,
            "profile.default_content_setting_values.media_stream_camera": 2,
            "profile.default_content_setting_values.automatic_downloads": 2,
            "profile.default_content_setting_values.protocol_handlers": 2,
        }
        browser_env = self.build_browser_env()

        for browser in ("edge", "chrome"):
            driver = None
            try:
                with open(os.devnull, "w", encoding="utf-8", errors="ignore") as sink:
                    with contextlib.redirect_stdout(sink), contextlib.redirect_stderr(sink):
                        if browser == "edge":
                            options = EdgeOptions()
                            for arg in common_args:
                                options.add_argument(arg)
                            if proxy_value:
                                options.add_argument(f"--proxy-server={proxy_value}")
                            options.add_experimental_option("prefs", prefs)
                            options.add_experimental_option("excludeSwitches", ["enable-logging"])
                            service = EdgeService(log_output=os.devnull, env=browser_env)
                            driver = webdriver.Edge(options=options, service=service)
                        else:
                            options = ChromeOptions()
                            for arg in common_args:
                                options.add_argument(arg)
                            if proxy_value:
                                options.add_argument(f"--proxy-server={proxy_value}")
                            options.add_experimental_option("prefs", prefs)
                            options.add_experimental_option("excludeSwitches", ["enable-logging"])
                            service = ChromeService(log_output=os.devnull, env=browser_env)
                            driver = webdriver.Chrome(options=options, service=service)
                driver.set_page_load_timeout(20)
                return driver
            except Exception:
                if driver:
                    try:
                        driver.quit()
                    except Exception:
                        pass
        return None

    def wait_for_browser_render(self, driver):
        deadline = time.time() + self.get_render_wait()
        while time.time() < deadline:
            try:
                state = driver.execute_script("return document.readyState")
                if state == "complete":
                    break
            except Exception:
                break
            time.sleep(0.2)
        remaining = max(0.2, deadline - time.time())
        time.sleep(remaining)

    def fetch_rendered_page(self, url: str) -> tuple[str, str] | None:
        probed = self.probe_login_with_devtools(url)
        if probed and probed.get("html"):
            return probed.get("current_url") or url, probed.get("html") or ""
        driver = self.init_headless_driver()
        if not driver:
            return None
        try:
            driver.get(url)
            self.wait_for_browser_render(driver)
            self.dismiss_browser_obstructions(driver)
            return driver.current_url, driver.page_source or ""
        except Exception:
            return None
        finally:
            try:
                driver.quit()
            except Exception:
                pass

    def probe_login_with_browser(self, url: str, proxy_addr: str | None = None) -> dict | None:
        devtools_probe = self.probe_login_with_devtools(url, proxy_addr)
        if self.browser_probe_has_login_signal(devtools_probe):
            return devtools_probe
        probe_url = (devtools_probe or {}).get("current_url") or url
        driver = self.init_headless_driver(proxy_addr)
        if not driver:
            return devtools_probe
        try:
            active_record = getattr(self, "_browser_probe_record", None)
            driver.get(probe_url)
            self.wait_for_browser_render(driver)
            self.dismiss_browser_obstructions(driver)
            dom = BruteForceHandler(
                None,
                None,
                locator_rule_resolver=self.resolve_locator_rule_for_record,
            )
            dom._active_record = active_record
            dom = dom.locate_login_dom(driver)
            if not dom or dom.get("pass") is None:
                fallback_probe = {
                    "current_url": driver.current_url or url,
                    "html": driver.page_source or "",
                    "login_form": False,
                    "captcha_present": False,
                    "field_summary": "",
                    "form_method": "",
                    "probe_backend": "webdriver",
                }
                return devtools_probe or fallback_probe

            hints = []
            if dom.get("user") is not None:
                hints.append("text:browser_user")
            if dom.get("pass") is not None:
                hints.append("password:browser_password")
            if dom.get("captcha") is not None:
                hints.append("text:browser_captcha")

            return {
                "current_url": driver.current_url or probe_url,
                "html": driver.page_source or "",
                "login_form": True,
                "captcha_present": dom.get("captcha") is not None,
                "field_summary": " | ".join(hints),
                "form_method": "POST",
                "probe_backend": "webdriver",
                "probe_url": probe_url,
            }
        except Exception:
            return devtools_probe
        finally:
            try:
                driver.quit()
            except Exception:
                pass

    def inspect_target(self, session: requests.Session, record: AuditRecord, screenshot_dir: Path) -> AuditRecord:
        response = session.get(
            record.target,
            timeout=8,
            verify=False,
            allow_redirects=self.follow_redirect_var.get(),
        )
        response.encoding = response.apparent_encoding or response.encoding
        soup = self.analyze_record_from_html(record, response.url, response.text)
        if response.status_code >= 400:
            record.result = f"{record.result} | HTTP {response.status_code}".strip(" |")

        actionable_login = self.is_actionable_login_record(record)
        browser_probe_needed = (
            self.browser_render_var.get()
            and not actionable_login
            and (self.brute_var.get() or record.login_form or self.should_use_browser_render_fallback(record, response.text))
        )
        self.log_record_trace(
            record,
            "补扫判定 | "
            f"browser_probe_needed={browser_probe_needed} | actionable={actionable_login} | "
            f"login_form={record.login_form} | brute={'on' if self.brute_var.get() else 'off'}",
        )

        if browser_probe_needed:
            self._browser_probe_record = record
            probed = self.probe_login_with_browser(record.final_url or record.target, record.proxy_used or None)
            self._browser_probe_record = None
            if probed:
                self.log_record_trace(
                    record,
                    "浏览器补扫 | "
                    f"backend={probed.get('probe_backend', '-')} | probe_url={probed.get('probe_url') or '-'} | "
                    f"current_url={probed.get('current_url') or '-'} | input_count={probed.get('input_count', 0)} | "
                    f"field_summary={probed.get('field_summary') or '-'} | captcha={bool(probed.get('captcha_present'))} | "
                    f"route_attempts={', '.join(probed.get('route_attempts') or []) or '-'}",
                )
                probe_signal = self.browser_probe_has_login_signal(probed)
                rendered_url = probed.get("current_url") or record.final_url or record.target
                rendered_html = probed.get("html") or ""
                soup = self.analyze_record_from_html(record, rendered_url, rendered_html)
                if probed.get("login_form"):
                    record.login_form = True
                    record.password_field_count = max(1, record.password_field_count)
                    record.field_summary = record.field_summary or probed.get("field_summary", "")
                    record.form_method = record.form_method or probed.get("form_method", "POST")
                if probed.get("captcha_present"):
                    record.captcha_present = True
                    if "未见验证码" in record.result:
                        record.result = record.result.replace("未见验证码", "存在验证码")
                if response.status_code >= 400:
                    record.result = f"{record.result} | HTTP {response.status_code}".strip(" |")
                hit_label = "浏览器渲染补扫命中(CDP)" if probed.get("probe_backend") == "cdp" else "浏览器渲染补扫命中"
                feature_label = "浏览器渲染补扫发现登录页特征(CDP)" if probed.get("probe_backend") == "cdp" else "浏览器渲染补扫发现登录页特征"
                miss_label = "浏览器渲染补扫未命中(CDP)" if probed.get("probe_backend") == "cdp" else "浏览器渲染补扫未命中"
                actionable_after_render = self.is_actionable_login_record(record)
                if probe_signal or actionable_after_render:
                    record.result = f"{record.result} | {hit_label}".strip(" |")
                    record.risk_level = self.calculate_risk(record)
                elif record.login_form:
                    record.result = f"{record.result} | {feature_label}".strip(" |")
                    record.risk_level = self.calculate_risk(record)
                else:
                    record.result = f"{record.result} | {miss_label}".strip(" |")
            else:
                self.log_record_trace(record, "浏览器补扫失败 | 未获取到有效 probe 结果", level="WARN")
                record.result = f"{record.result} | 浏览器补扫失败".strip(" |")

        if self.brute_var.get() and not self.is_actionable_login_record(record):
            record.result = f"{record.result} | 未识别到可用登录框，未进入弱口令尝试".strip(" |")

        self.log_record_trace(
            record,
            f"最终结果 | risk={record.risk_level} | result={record.result} | error={record.error or '-'}",
            level="RESULT",
        )
        self.log_queue.put(
            f"[+] {record.target} 分析完成: 标题={record.title or '无标题'} 风险={record.risk_level}"
        )
        if self.brute_var.get() and self.is_actionable_login_record(record):
            self.log_queue.put(f"[*] 正在对 {record.target} 进行弱口令尝试...")
            brute = BruteForceHandler(
                session,
                self.log_queue,
                driver_factory=lambda proxy_addr=record.proxy_used: self.init_headless_driver(proxy_addr or None),
                render_wait=self.get_render_wait(),
                captcha_ocr_enabled=self.captcha_ocr_var.get(),
                ocr_endpoint=self.ocr_endpoint_entry.get().strip(),
                ocr_endpoint_resolver=self.resolve_ocr_endpoint_for_record,
                locator_rule_resolver=self.resolve_locator_rule_for_record,
                captcha_lock=self.captcha_brute_lock,
            )
            res = brute.run(
                record,
                dict_mode=self.dict_mode_var.get(),
                user_dict_path=self.user_dict_entry.get().strip(),
                pass_dict_path=self.pass_dict_entry.get().strip(),
                soup=soup,
            )
            record.result += f" | 弱口令检测: {res}"

            if res.startswith(BRUTE_FORCE_SUCCESS_PREFIX):
                record.risk_level = "高"
                self.log_queue.put(f"[!] 发现弱口令风险: {record.target} -> {res}")
        
        return record

    def capture_stage(self, records: list[AuditRecord], screenshot_dir: Path, allow_when_stopped: bool = False):
        capture_targets = [record for record in records if self.should_capture_record(record)]
        if not capture_targets:
            self.log_queue.put(f"[*] 截图阶段跳过，当前策略={self.capture_policy_var.get()}，没有符合条件的目标。")
            return

        capture_workers = 1 if self.get_worker_count() <= 2 else 2
        self.log_queue.put(f"[*] 进入截图阶段，共 {len(capture_targets)} 条，截图并发 {capture_workers}。")
        executor = concurrent.futures.ThreadPoolExecutor(
            max_workers=capture_workers,
            thread_name_prefix="capture",
        )
        future_map = {}
        try:
            for record in capture_targets:
                if not self.is_scanning and not allow_when_stopped:
                    break
                future = executor.submit(self.capture_record_screenshot, record, screenshot_dir)
                future_map[future] = record

            while future_map:
                done, _pending = concurrent.futures.wait(
                    future_map.keys(),
                    timeout=0.2,
                    return_when=concurrent.futures.FIRST_COMPLETED,
                )
                if not self.is_scanning and not allow_when_stopped:
                    for future in list(future_map):
                        future.cancel()
                    break
                for future in done:
                    record = future_map.pop(future)
                    try:
                        screenshot_path = future.result()
                        if screenshot_path:
                            record.screenshot_path = str(screenshot_path)
                            self.log_queue.put(f"[+] 已截图: {record.target}")
                        else:
                            self.log_queue.put(f"[~] 截图跳过或失败: {record.target}")
                        self.schedule_progress_snapshot(reason="screenshot-updated")
                        item_id = self.find_item_id_by_record(record)
                        if item_id:
                            self.root.after(
                                0,
                                self.update_tree_item,
                                item_id,
                                record.status,
                                record.title or "-",
                                record.risk_level,
                                record.result,
                            )
                    except Exception as exc:
                        self.log_queue.put(f"[~] 截图异常: {record.target} -> {exc}")
        finally:
            executor.shutdown(wait=False, cancel_futures=True)

    def record_has_login_signal(self, record: AuditRecord) -> bool:
        result_text = record.result or ""
        return bool(
            record.login_form
            or self.is_actionable_login_record(record)
            or record.password_field_count > 0
            or "疑似登录页" in result_text
            or "浏览器渲染补扫命中" in result_text
            or "浏览器渲染补扫发现登录页特征" in result_text
        )

    def should_capture_record(self, record: AuditRecord) -> bool:
        if record.screenshot_path:
            return False
        policy = self.capture_policy_var.get()
        if policy == CAPTURE_POLICY_ALL:
            return True
        if policy == CAPTURE_POLICY_HIGH:
            return record.risk_level == "高"
        if policy == CAPTURE_POLICY_LOGIN:
            return self.record_has_login_signal(record)
        return self.record_has_login_signal(record) or record.risk_level in {"高", "中"} or bool(record.field_summary)

    def capture_record_screenshot(self, record: AuditRecord, screenshot_dir: Path) -> Path | None:
        delay = self.get_capture_delay()
        if delay > 0:
            time.sleep(delay)
        return self.try_capture_screenshot(record.final_url or record.target, screenshot_dir, record.record_id)

    def estimate_password_field_count(self, soup: BeautifulSoup) -> int:
        strict_matches = soup.find_all("input", {"type": re.compile("password", re.I)})
        if strict_matches:
            return len(strict_matches)
        loose_count = 0
        for tag in soup.find_all(["input", "textarea"]):
            field_type = (tag.get("type") or tag.name or "").strip().lower()
            if field_type in {"hidden", "submit", "button", "checkbox", "radio", "file"}:
                continue
            hints = " ".join(
                filter(
                    None,
                    [
                        tag.get("placeholder"),
                        tag.get("name"),
                        tag.get("id"),
                        tag.get("aria-label"),
                        tag.get("title"),
                        tag.get("autocomplete"),
                        tag.get("data-placeholder"),
                        tag.get("data-label"),
                    ],
                )
            ).lower()
            if any(keyword in hints for keyword in PASSWORD_FIELD_KEYWORDS):
                loose_count += 1
        return loose_count

    def extract_form_details(self, soup: BeautifulSoup) -> tuple[str, str, str]:
        form = None
        for candidate in soup.find_all("form"):
            if candidate.find("input", {"type": re.compile("password", re.I)}):
                form = candidate
                break
        if form is None:
            form = soup.find("form")
        if form is not None:
            action = (form.get("action") or "").strip()
            method = (form.get("method") or "").strip().upper()
            if not method:
                method = "POST" if form.find("input", {"type": re.compile("password", re.I)}) else "GET"
            fields = []
            for tag in form.find_all(["input", "select", "textarea"]):
                input_type = (tag.get("type") or tag.name).strip().lower()
                name = (tag.get("name") or tag.get("id") or tag.get("placeholder") or "").strip()
                if not name and input_type == "hidden":
                    continue
                label = f"{input_type}:{name}" if name else input_type
                fields.append(label[:48])
            return action[:120], method[:12], " | ".join(fields[:8])

        loose_fields = []
        has_password_hint = False
        for tag in soup.find_all(["input", "select", "textarea"]):
            input_type = (tag.get("type") or tag.name).strip().lower()
            if input_type in {"hidden", "submit", "button", "checkbox", "radio", "file"}:
                continue
            name = (
                tag.get("name")
                or tag.get("id")
                or tag.get("placeholder")
                or tag.get("aria-label")
                or tag.get("title")
                or ""
            ).strip()
            hints = " ".join(
                filter(
                    None,
                    [
                        name,
                        tag.get("autocomplete"),
                        tag.get("data-placeholder"),
                        tag.get("data-label"),
                    ],
                )
            ).lower()
            if not any(keyword in hints for keyword in USER_FIELD_KEYWORDS + PASSWORD_FIELD_KEYWORDS + CAPTCHA_FIELD_KEYWORDS):
                continue
            if any(keyword in hints for keyword in PASSWORD_FIELD_KEYWORDS):
                has_password_hint = True
            label = f"{input_type}:{name}" if name else input_type
            loose_fields.append(label[:48])
        if loose_fields:
            return "", "POST" if has_password_hint else "", " | ".join(loose_fields[:8])
        return "", "", ""

    def extract_title(self, soup: BeautifulSoup) -> str:
        title_tag = soup.find("title")
        if title_tag and title_tag.text:
            return title_tag.text.strip()[:80]
        header = soup.find(["h1", "h2"])
        if header and header.text:
            return header.text.strip()[:80]
        return "无标题"

    def collect_text_blob(self, soup: BeautifulSoup) -> str:
        parts = [soup.get_text(" ", strip=True)]
        for tag in soup.find_all(["input", "button", "label", "a"]):
            for attr in ("placeholder", "value", "name", "id", "aria-label", "title", "data-placeholder", "data-label"):
                value = tag.get(attr)
                if value:
                    parts.append(str(value))
        return " ".join(parts).lower()

    def contains_any(self, text_blob: str, keywords: list[str]) -> bool:
        lower_keywords = [kw.lower() for kw in keywords]
        return any(keyword in text_blob for keyword in lower_keywords)

    def has_captcha_widget(self, soup: BeautifulSoup) -> bool:
        for tag in soup.find_all(["img", "input", "div", "span"]):
            raw = " ".join(
                filter(
                    None,
                    [
                        tag.get("src"),
                        tag.get("class") and " ".join(tag.get("class")),
                        tag.get("id"),
                        tag.get("name"),
                    ],
                )
            ).lower()
            if any(keyword.lower() in raw for keyword in CAPTCHA_KEYWORDS):
                return True
        return False

    def compute_login_score(self, soup: BeautifulSoup, text_blob: str, title: str) -> int:
        score = 0
        if self.contains_any(title.lower(), LOGIN_KEYWORDS):
            score += 2
        keyword_hits = sum(1 for keyword in LOGIN_KEYWORDS if keyword.lower() in text_blob)
        score += min(keyword_hits, 3)
        forms = soup.find_all("form")
        if forms:
            score += 1
        if self.estimate_password_field_count(soup):
            score += 3
        hint_hits = 0
        for tag in soup.find_all("input"):
            joined = " ".join(
                filter(
                    None,
                    [
                        tag.get("placeholder"),
                        tag.get("name"),
                        tag.get("id"),
                        tag.get("aria-label"),
                        tag.get("title"),
                        tag.get("data-placeholder"),
                        tag.get("data-label"),
                    ],
                )
            ).lower()
            if any(hint.lower() in joined for hint in INPUT_HINTS):
                hint_hits += 1
        score += min(hint_hits, 2)
        if self.mode_var.get() == "NLP模式":
            text_words = re.findall(r"[\u4e00-\u9fa5a-zA-Z0-9]+", text_blob)
            weight = sum(
                1
                for token in text_words
                if token in {"登录", "登陆", "用户名", "密码", "admin", "login", "account", "password"}
            )
            score += min(weight, 4)
        return score

    def calculate_risk(self, record: AuditRecord) -> str:
        actionable_login = self.is_actionable_login_record(record)
        if not actionable_login and not record.default_hint:
            return "低"
        risk = 0
        if record.default_hint:
            risk += 3
        if actionable_login:
            if not record.captcha_present:
                risk += 1
            if not record.mfa_present:
                risk += 1
            if not record.lockout_hint:
                risk += 1
            if record.password_field_count >= 1:
                risk += 1

        if risk >= 5:
            return "高"
        if risk >= 3:
            return "中"
        return "低"

    def try_capture_screenshot(self, url: str, screenshot_dir: Path, record_id: int) -> Path | None:
        filename = screenshot_dir / f"{record_id:04d}.png"
        cli_capture = self.capture_screenshot_with_browser_cli(url, filename)
        if cli_capture:
            return cli_capture
        driver = self.init_headless_driver()
        if not driver:
            return None
        try:
            driver.get(url)
            self.wait_for_browser_render(driver)
            self.dismiss_browser_obstructions(driver)
            driver.save_screenshot(str(filename))
            return filename
        except Exception:
            return None
        finally:
            try:
                driver.quit()
            except Exception:
                pass
        return None

    def render_html_screenshot(self, record: AuditRecord) -> str:
        if not record.screenshot_path:
            return "-"
        path = Path(record.screenshot_path)
        if not path.exists():
            return html.escape(record.screenshot_path)
        uri = path.resolve().as_uri()
        return f'<a href="{uri}" target="_blank"><img src="{uri}" alt="screenshot"><span>查看原图</span></a>'

    def get_selected_record(self) -> AuditRecord | None:
        selection = self.tree.selection()
        if not selection:
            messagebox.showinfo("提示", "请先选中一条记录。")
            return None
        return self.records_by_item.get(selection[0])

    def get_selected_record_silent(self) -> AuditRecord | None:
        selection = self.tree.selection()
        if not selection:
            return None
        return self.records_by_item.get(selection[0])

    def open_selected_screenshot(self):
        record = self.get_selected_record()
        if not record:
            return
        if not record.screenshot_path:
            messagebox.showinfo("提示", "这条记录还没有截图。")
            return
        path = Path(record.screenshot_path)
        if not path.exists():
            messagebox.showwarning("提示", f"截图文件不存在:\n{path}")
            return
        os.startfile(path)

    def open_preview_zoom(self):
        record = self.get_selected_record_silent()
        if not record or not record.screenshot_path:
            messagebox.showinfo("提示", "当前记录没有可放大的截图。")
            return
        path = Path(record.screenshot_path)
        if not path.exists():
            messagebox.showwarning("提示", f"截图文件不存在:\n{path}")
            return

        window = tk.Toplevel(self.root)
        window.title(f"截图放大查看 - {record.title or record.target}")
        window.geometry("1280x900")
        window.transient(self.root)

        top_bar = tk.Frame(window, pady=8, padx=12)
        top_bar.pack(fill=tk.X)
        tk.Label(
            top_bar,
            text=record.final_url or record.target,
            anchor="w",
            justify=tk.LEFT,
            wraplength=1080,
        ).pack(side=tk.LEFT, fill=tk.X, expand=True)
        tk.Button(top_bar, text="打开原图", command=lambda: os.startfile(path)).pack(side=tk.RIGHT)

        canvas = tk.Canvas(window, bg="#111827", highlightthickness=0)
        canvas.pack(fill=tk.BOTH, expand=True)

        image = Image.open(path)
        state = {"scale": 1.0, "photo": None, "image_id": None}

        def redraw():
            display = image.copy()
            scaled_size = (
                max(1, int(display.width * state["scale"])),
                max(1, int(display.height * state["scale"])),
            )
            display = display.resize(scaled_size)
            state["photo"] = ImageTk.PhotoImage(display)
            if state["image_id"] is None:
                state["image_id"] = canvas.create_image(20, 20, image=state["photo"], anchor="nw")
            else:
                canvas.itemconfig(state["image_id"], image=state["photo"])
            canvas.coords(state["image_id"], 20, 20)
            canvas.config(scrollregion=(0, 0, display.width + 40, display.height + 40))

        def zoom(step: float):
            state["scale"] = max(0.2, min(3.0, state["scale"] + step))
            redraw()

        control_bar = tk.Frame(window, pady=8, padx=12)
        control_bar.pack(fill=tk.X)
        tk.Button(control_bar, text="放大", command=lambda: zoom(0.2)).pack(side=tk.LEFT)
        tk.Button(control_bar, text="缩小", command=lambda: zoom(-0.2)).pack(side=tk.LEFT, padx=(6, 0))
        tk.Button(control_bar, text="重置", command=lambda: reset_zoom()).pack(side=tk.LEFT, padx=(6, 0))

        def reset_zoom():
            state["scale"] = 1.0
            redraw()

        def on_mouse_wheel(event):
            if event.delta > 0:
                zoom(0.1)
            elif event.delta < 0:
                zoom(-0.1)

        def start_pan(event):
            canvas.scan_mark(event.x, event.y)

        def drag_pan(event):
            canvas.scan_dragto(event.x, event.y, gain=1)

        canvas.bind("<Configure>", lambda _event: redraw())
        canvas.bind("<MouseWheel>", on_mouse_wheel)
        canvas.bind("<ButtonPress-1>", start_pan)
        canvas.bind("<B1-Motion>", drag_pan)
        window.bind("<Control-plus>", lambda _event: zoom(0.2))
        window.bind("<Control-minus>", lambda _event: zoom(-0.2))
        reset_zoom()

    def open_selected_target(self):
        record = self.get_selected_record()
        if not record:
            return
        target = record.final_url or record.target
        if not target:
            messagebox.showinfo("提示", "这条记录没有可打开的地址。")
            return
        webbrowser.open(target)

    def open_profile_dashboard(self):
        window = tk.Toplevel(self.root)
        window.title("画像统计")
        window.geometry("1120x720")
        window.transient(self.root)

        profile = self.get_profile_dashboard_data(self.get_records_snapshot())

        summary_frame = tk.LabelFrame(window, text="核心指标", padx=10, pady=10)
        summary_frame.pack(fill=tk.X, padx=12, pady=(10, 6))
        for index, item in enumerate(profile["summary_items"]):
            card = tk.Frame(summary_frame, bd=1, relief=tk.SOLID, padx=10, pady=8, bg="#f8fbff")
            card.grid(row=0, column=index, padx=4, sticky="nsew")
            tk.Label(card, text=item["label"], fg="#64748b", bg="#f8fbff").pack(anchor=tk.W)
            tk.Label(card, text=str(item["value"]), font=("Microsoft YaHei", 14, "bold"), bg="#f8fbff", fg="#0f172a").pack(anchor=tk.W, pady=(4, 0))
            summary_frame.columnconfigure(index, weight=1)

        notebook = ttk.Notebook(window)
        notebook.pack(fill=tk.BOTH, expand=True, padx=12, pady=(0, 10))

        tabs = [
            ("站点画像", profile["profile_tags"], "画像标签"),
            ("OCR规则", profile["ocr_rules"], "规则"),
            ("代理出口", profile["proxy_usage"], "代理出口"),
            ("失败原因", profile["failures"], "失败类型"),
        ]

        for title, rows, first_col in tabs:
            frame = tk.Frame(notebook)
            notebook.add(frame, text=title)
            self.populate_profile_tab(frame, rows, first_col)

        button_bar = tk.Frame(window, pady=8, padx=12)
        button_bar.pack(fill=tk.X)
        tk.Button(button_bar, text="关闭", command=window.destroy).pack(side=tk.RIGHT)

    def open_locator_rule_quick_editor(self):
        record = self.get_selected_record()
        if not record:
            return

        parsed = urlparse(record.final_url or record.target or "")
        host = parsed.netloc or parsed.path or record.target

        window = tk.Toplevel(self.root)
        window.title(f"快速录入定位规则 - {host}")
        window.geometry("980x640")
        window.minsize(920, 580)
        window.transient(self.root)
        window.configure(bg=self.palette["bg"])

        header = self.make_section(window, "当前目标", padx=12, pady=10)
        header.pack(fill=tk.X, padx=12, pady=(12, 8))
        self.make_label(
            header,
            text=(
                f"目标: {record.target}\n"
                f"最终URL: {record.final_url or '-'}\n"
                "录入 = 把这个站点的登录框定位规则写入规则文件，不是填账号密码。"
            ),
            anchor="w",
            justify=tk.LEFT,
            wraplength=900,
        ).pack(fill=tk.X)
        self.make_label(
            header,
            text=f"字段摘要: {record.field_summary or '当前记录还没有提取到字段摘要，可先扫一次再补规则。建议优先按 host 维度写规则，避免误伤其他站点。'}",
            anchor="w",
            justify=tk.LEFT,
            wraplength=900,
            muted=True,
        ).pack(fill=tk.X, pady=(6, 0))

        body = tk.Frame(window, bg=self.palette["bg"], padx=12)
        body.pack(fill=tk.BOTH, expand=True)
        body.columnconfigure(0, weight=3)
        body.columnconfigure(1, weight=2)
        body.rowconfigure(0, weight=1)

        form = self.make_section(body, "规则录入", padx=12, pady=10)
        form.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        form.columnconfigure(1, weight=1)

        preview_panel = self.make_section(body, "预览与提示", padx=12, pady=10)
        preview_panel.grid(row=0, column=1, sticky="nsew")
        preview_panel.columnconfigure(0, weight=1)
        preview_panel.rowconfigure(1, weight=1)

        scope_var = tk.StringVar(value="host")
        pattern_var = tk.StringVar(value=host)
        user_var = tk.StringVar(value="css:input[placeholder*='账号'], input[placeholder*='手机号'], input[name='username']")
        pass_var = tk.StringVar(value="css:input[type='password'], input[placeholder*='密码']")
        submit_var = tk.StringVar(value="xpath://button[contains(.,'登录')]")
        captcha_var = tk.StringVar(value="css:input[placeholder*='验证码'], input[name*='captcha']")
        frame_var = tk.StringVar(value="")

        field_defs = [
            ("作用域", "决定用 host / url / title / target 哪种维度匹配。", lambda parent: ttk.Combobox(parent, textvariable=scope_var, values=["host", "url", "title", "target", "default"], width=12, state="readonly")),
            ("匹配值", "default 作用域可留空，其他作用域建议填写尽量稳定的片段。", lambda parent: self.make_entry(parent, textvariable=pattern_var)),
            ("账号框", "支持多个候选，用英文逗号分隔。", lambda parent: self.make_entry(parent, textvariable=user_var)),
            ("密码框", "通常建议至少保留一个 type=password 的 css 选择器。", lambda parent: self.make_entry(parent, textvariable=pass_var)),
            ("登录按钮", "可选，按钮点击识别不稳定时建议显式写入。", lambda parent: self.make_entry(parent, textvariable=submit_var)),
            ("验证码框", "页面存在验证码时再填写。", lambda parent: self.make_entry(parent, textvariable=captcha_var)),
            ("Iframe", "登录框在 iframe 里时填写 frame 定位。", lambda parent: self.make_entry(parent, textvariable=frame_var)),
        ]

        pattern_entry = None
        for row, (label, hint, widget_factory) in enumerate(field_defs):
            self.make_label(form, text=label + ":", anchor="w").grid(row=row * 2, column=0, sticky="w", pady=(0, 4), padx=(0, 8))
            widget = widget_factory(form)
            widget.grid(row=row * 2, column=1, sticky="ew", pady=(0, 2))
            if label == "匹配值":
                pattern_entry = widget
            self.make_label(form, text=hint, anchor="w", justify=tk.LEFT, wraplength=520, muted=True).grid(
                row=row * 2 + 1,
                column=1,
                sticky="w",
                pady=(0, 8),
            )

        self.make_label(
            form,
            text="值支持 css: 或 xpath: 前缀。留空字段不会写入规则；账号框和密码框必须填写。",
            anchor="w",
            justify=tk.LEFT,
            wraplength=560,
            muted=True,
        ).grid(row=len(field_defs) * 2, column=0, columnspan=2, sticky="w", pady=(6, 0))

        self.make_label(preview_panel, text="即将写入的规则", anchor="w", font=self.font_section).grid(row=0, column=0, sticky="w")
        preview_text = tk.Text(
            preview_panel,
            height=10,
            state=tk.DISABLED,
            wrap=tk.WORD,
            bg="#f8fafc",
            fg=self.palette["text"],
            insertbackground=self.palette["text"],
            relief=tk.SOLID,
            bd=1,
            highlightthickness=1,
            highlightbackground=self.palette["border"],
            highlightcolor=self.palette["accent"],
            font=self.font_normal,
        )
        preview_text.grid(row=1, column=0, sticky="nsew", pady=(6, 10))

        help_text = (
            "建议:\n"
            "1. 优先写 host 规则，命中范围更稳。\n"
            "2. 定位器尽量先写 placeholder/name/id 等稳定属性，再补 xpath。\n"
            "3. 如果页面有多个登录入口，先放多个 css 候选，不要只赌一个。\n"
            "4. 保存后重新扫描当前目标即可验证规则是否生效。"
        )
        self.make_label(
            preview_panel,
            text=help_text,
            anchor="nw",
            justify=tk.LEFT,
            wraplength=320,
            muted=True,
        ).grid(row=2, column=0, sticky="ew")

        def build_rule_text(include_placeholder=False):
            scope = scope_var.get().strip() or "host"
            pattern = pattern_var.get().strip()
            selectors = []
            for key, value in [
                ("user", user_var.get().strip()),
                ("pass", pass_var.get().strip()),
                ("submit", submit_var.get().strip()),
                ("captcha", captcha_var.get().strip()),
                ("frame", frame_var.get().strip()),
            ]:
                if value:
                    selectors.append(f"{key}={value}")

            if scope == "default":
                head = "default"
            else:
                head_pattern = pattern or ("<匹配值>" if include_placeholder else "")
                if not head_pattern:
                    return ""
                head = f"{scope}:{head_pattern}"

            if not selectors and include_placeholder:
                selectors = ["user=css:input[name='username']", "pass=css:input[type='password']"]
            return f"{head} => " + "; ".join(selectors) if selectors else head

        def refresh_preview(*_args):
            if pattern_entry is not None:
                if scope_var.get().strip() == "default":
                    pattern_entry.config(state=tk.DISABLED)
                else:
                    pattern_entry.config(state=tk.NORMAL)

            preview_value = build_rule_text(include_placeholder=True)
            preview_text.config(state=tk.NORMAL)
            preview_text.delete("1.0", tk.END)
            preview_text.insert("1.0", preview_value or "填写作用域、匹配值和核心定位器后，这里会显示最终规则。")
            preview_text.config(state=tk.DISABLED)

        def save_rule():
            scope = scope_var.get().strip() or "host"
            pattern = pattern_var.get().strip()
            if scope != "default" and not pattern:
                messagebox.showwarning("提示", "匹配值不能为空。", parent=window)
                return
            if not user_var.get().strip() or not pass_var.get().strip():
                messagebox.showwarning("提示", "至少需要填写账号框和密码框定位器。", parent=window)
                return
            rule_text = build_rule_text()
            self.append_locator_rule(rule_text)
            messagebox.showinfo("成功", "定位规则已保存，重新扫描该目标即可生效。", parent=window)
            window.destroy()

        for variable in [scope_var, pattern_var, user_var, pass_var, submit_var, captcha_var, frame_var]:
            variable.trace_add("write", refresh_preview)
        refresh_preview()

        button_bar = tk.Frame(window, bg=self.palette["bg"], padx=12, pady=12)
        button_bar.pack(fill=tk.X)
        self.make_button(button_bar, text="写入规则", command=save_rule, tone="primary").pack(side=tk.LEFT)
        self.make_button(button_bar, text="打开规则文件", command=lambda: os.startfile(self.ensure_locator_rule_file())).pack(side=tk.LEFT, padx=(6, 0))
        self.make_button(button_bar, text="关闭", command=window.destroy).pack(side=tk.RIGHT)
        window.bind("<Control-s>", lambda _event: save_rule())

    def open_proxy_status_window(self):
        window = tk.Toplevel(self.root)
        window.title("代理状态")
        window.geometry("980x560")
        window.transient(self.root)

        top_bar = tk.Frame(window, pady=8, padx=12)
        top_bar.pack(fill=tk.X)
        mode_text = self.proxy_mode_var.get() if self.use_proxy_var.get() else "未启用代理"
        info_parts = [f"当前模式: {mode_text}"]
        if self.use_proxy_var.get() and self.proxy_mode_var.get() == PROXY_MODE_POOL:
            info_parts.append(f"代理池文件: {self.proxy_pool_entry.get().strip() or '-'}")
            info_parts.append(f"熔断阈值: {self.get_proxy_fail_threshold()} 次")
            info_parts.append(f"冷却: {self.get_proxy_cooldown_seconds()}s")
        elif self.use_proxy_var.get():
            info_parts.append(f"代理地址: {self.normalize_proxy_address(self.proxy_entry.get()) or '-'}")
        info_var = tk.StringVar(value=" | ".join(info_parts))
        tk.Label(top_bar, textvariable=info_var, anchor="w", justify=tk.LEFT, wraplength=820).pack(side=tk.LEFT, fill=tk.X, expand=True)

        table_frame = tk.Frame(window, padx=12)
        table_frame.pack(fill=tk.BOTH, expand=True)
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)
        columns = ("代理", "状态", "失败次数", "剩余冷却", "分配目标", "最后成功", "最近错误")
        tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=18)
        y_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        x_scroll = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

        self.configure_tree_columns(tree, [
            ("代理", 220, tk.W),
            ("状态", 76, tk.CENTER),
            ("失败次数", 80, tk.CENTER),
            ("剩余冷却", 90, tk.CENTER),
            ("分配目标", 80, tk.CENTER),
            ("最后成功", 90, tk.CENTER),
            ("最近错误", 360, tk.W),
        ])

        tree.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")

        status_var = tk.StringVar(value="")
        tk.Label(window, textvariable=status_var, anchor="w", padx=12, pady=6, fg="#475569").pack(fill=tk.X)

        def populate():
            for item in tree.get_children():
                tree.delete(item)
            rows = self.get_proxy_status_rows()
            if not rows:
                status_var.set("当前没有可展示的代理状态。")
                return
            for row in rows:
                tags = ("cooldown",) if row["status"] == "冷却中" else ("healthy",)
                tree.insert(
                    "",
                    tk.END,
                    values=(
                        row["proxy"],
                        row["status"],
                        row["fail_count"],
                        row["cooldown"],
                        row["assigned"],
                        row["last_ok"],
                        row["last_error"],
                    ),
                    tags=tags,
                )
            healthy_count = sum(1 for row in rows if row["status"] == "可用")
            cooldown_count = len(rows) - healthy_count
            status_var.set(f"代理总数 {len(rows)} | 可用 {healthy_count} | 冷却中 {cooldown_count}")

        tree.tag_configure("healthy", background="#edf9ed")
        tree.tag_configure("cooldown", background="#fff0cf")
        populate()

        button_bar = tk.Frame(window, pady=8, padx=12)
        button_bar.pack(fill=tk.X)
        tk.Button(button_bar, text="刷新", command=populate).pack(side=tk.LEFT)
        tk.Button(button_bar, text="重置健康状态", command=lambda: reset_proxy_health()).pack(side=tk.LEFT, padx=(6, 0))
        tk.Button(button_bar, text="关闭", command=window.destroy).pack(side=tk.RIGHT)

        def reset_proxy_health():
            self.proxy_health.clear()
            self.log_message("[*] 已清空代理健康状态与熔断计数。")
            populate()

    def populate_profile_tab(self, frame, rows: list[dict], first_col_name: str):
        table_frame = tk.Frame(frame, padx=12, pady=12)
        table_frame.pack(fill=tk.BOTH, expand=True)
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)
        columns = (first_col_name, "数量")
        tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=16)
        y_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        x_scroll = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        self.configure_tree_columns(tree, [(first_col_name, 720, tk.W), ("数量", 100, tk.CENTER)])
        tree.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")

        if not rows:
            tree.insert("", tk.END, values=("-", 0))
            return
        for row in rows:
            tree.insert("", tk.END, values=(row["name"], row["count"]))

    def open_output_dir(self):
        os.startfile(self.output_dir)

    def open_project_manager(self):
        window = tk.Toplevel(self.root)
        window.title("工程列表")
        window.geometry("860x520")
        window.transient(self.root)
        window.grab_set()

        search_frame = tk.Frame(window, pady=10, padx=12)
        search_frame.pack(fill=tk.X)
        tk.Label(search_frame, text="搜索工程:").pack(side=tk.LEFT)
        search_var = tk.StringVar()
        search_entry = tk.Entry(search_frame, textvariable=search_var, width=38)
        search_entry.pack(side=tk.LEFT, padx=(6, 8))

        table_frame = tk.Frame(window, padx=12)
        table_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 8))
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)

        columns = ("工程名", "源文件", "保存时间", "已完成", "待处理")
        tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=16)
        y_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        x_scroll = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        self.configure_tree_columns(tree, [
            ("工程名", 180, tk.W),
            ("源文件", 360, tk.W),
            ("保存时间", 150, tk.CENTER),
            ("已完成", 70, tk.CENTER),
            ("待处理", 70, tk.CENTER),
        ])
        tree.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")

        status_var = tk.StringVar(value="双击加载工程，删除前请确认当前未在扫描。")
        tk.Label(window, textvariable=status_var, anchor="w").pack(fill=tk.X, padx=12)

        button_bar = tk.Frame(window, pady=10)
        button_bar.pack(fill=tk.X, padx=12)

        projects = self.get_project_summaries()
        item_map = {}

        def populate_projects(query: str = ""):
            item_map.clear()
            for item in tree.get_children():
                tree.delete(item)
            keyword = query.strip().lower()
            filtered = [
                project
                for project in projects
                if not keyword
                or keyword in project["name"].lower()
                or keyword in project["source_file"].lower()
            ]
            for project in filtered:
                item_id = tree.insert(
                    "",
                    tk.END,
                    values=(
                        project["name"],
                        project["source_file"],
                        project["saved_at"],
                        project["completed"],
                        project["pending"],
                    ),
                )
                item_map[item_id] = project
            status_var.set(f"共 {len(filtered)} 个工程，双击可直接加载。")

        populate_projects()
        search_var.trace_add("write", lambda *_args: populate_projects(search_var.get()))
        search_entry.focus_set()

        def get_selected_project():
            selection = tree.selection()
            if not selection:
                messagebox.showinfo("提示", "请先选中一个工程。", parent=window)
                return None
            return item_map.get(selection[0])

        def load_selected():
            project = get_selected_project()
            if not project:
                return
            self.load_project_file(Path(project["path"]))
            self.write_latest_project_pointer(project["saved_at"])
            self.log_message(f"[*] 已切换工程: {project['name']}")
            window.destroy()

        def delete_selected():
            project = get_selected_project()
            if not project:
                return
            if self.is_scanning:
                messagebox.showwarning("提示", "请先停止当前扫描，再删除工程。", parent=window)
                return
            confirmed = messagebox.askyesno(
                "确认删除",
                f"确定删除工程吗？\n\n{project['name']}\n{project['source_file']}",
                parent=window,
            )
            if not confirmed:
                return
            try:
                Path(project["path"]).unlink(missing_ok=True)
                selected = tree.selection()[0]
                projects.remove(project)
                populate_projects(search_var.get())
                status_var.set(f"已删除工程: {project['name']}")
            except Exception as exc:
                messagebox.showerror("错误", f"删除工程失败: {exc}", parent=window)

        def open_projects_dir():
            os.startfile(self.projects_dir)

        tk.Button(button_bar, text="加载工程", command=load_selected).pack(side=tk.LEFT)
        tk.Button(button_bar, text="删除工程", command=delete_selected).pack(side=tk.LEFT, padx=(6, 0))
        tk.Button(button_bar, text="打开工程目录", command=open_projects_dir).pack(side=tk.LEFT, padx=(6, 0))
        tk.Button(button_bar, text="关闭", command=window.destroy).pack(side=tk.RIGHT)
        tree.bind("<Double-1>", lambda _event: load_selected())

    def get_project_summaries(self) -> list[dict]:
        summaries = []
        for path in sorted(self.projects_dir.glob("*.json"), key=lambda p: p.stat().st_mtime, reverse=True):
            try:
                with open(path, "r", encoding="utf-8") as handle:
                    payload = json.load(handle)
                records = payload.get("records") or []
                completed = sum(1 for record in records if record.get("status") == "已完成")
                total = len(records)
                source_file = payload.get("source_file", "")
                summaries.append(
                    {
                        "name": Path(source_file).stem if source_file else path.stem,
                        "source_file": source_file or str(path),
                        "saved_at": payload.get("saved_at", "-"),
                        "completed": completed,
                        "pending": max(0, total - completed),
                        "path": str(path),
                    }
                )
            except Exception:
                continue
        return summaries

    def on_tree_double_click(self, _event):
        record = self.get_selected_record()
        if not record:
            return
        if record.screenshot_path and Path(record.screenshot_path).exists():
            self.open_selected_screenshot()
        elif record.error or len(record.result or "") > 80:
            self.open_record_detail_window(record)
        else:
            self.open_selected_target()

    def on_tree_select(self, _event):
        record = self.get_selected_record_silent()
        self.update_preview(record)

    def update_preview(self, record: AuditRecord | None):
        if not record:
            self.preview_title_var.set("未选择记录")
            self.preview_meta_var.set("目标、风险、最终URL 会显示在这里")
            self.preview_image = None
            self.preview_image_label.config(image="", text="暂无截图")
            self.set_preview_text("")
            return

        self.preview_title_var.set(record.title or record.target)
        meta_lines = [
            f"目标: {record.target}",
            f"最终URL: {record.final_url or '-'}",
            f"代理: {record.proxy_used or '直连'}",
            f"OCR: {record.ocr_endpoint_used or '-'}",
            f"状态: {record.status}    风险: {record.risk_level}",
        ]
        self.preview_meta_var.set("\n".join(meta_lines))
        self.load_preview_image(record.screenshot_path)
        self.set_preview_text(self.build_record_detail_text(record))

    def load_preview_image(self, screenshot_path: str):
        if not screenshot_path or not Path(screenshot_path).exists():
            self.preview_image = None
            self.preview_image_label.config(image="", text="暂无截图")
            return
        try:
            image = Image.open(screenshot_path)
            self.preview_image_label.update_idletasks()
            max_width = max(420, self.preview_image_label.winfo_width() - 20)
            max_height = max(300, self.preview_image_label.winfo_height() - 20)
            image.thumbnail((max_width, max_height))
            self.preview_image = ImageTk.PhotoImage(image)
            self.preview_image_label.config(image=self.preview_image, text="")
        except Exception:
            self.preview_image = None
            self.preview_image_label.config(image="", text="截图加载失败")

    def set_preview_text(self, text: str):
        self.preview_text.config(state=tk.NORMAL)
        self.preview_text.delete("1.0", tk.END)
        self.preview_text.insert(tk.END, text)
        self.preview_text.xview_moveto(0)
        self.preview_text.yview_moveto(0)
        self.preview_text.config(state=tk.DISABLED)

    def split_pipe_items(self, text: str) -> list[str]:
        items = [part.strip() for part in (text or "").split(" | ") if part.strip()]
        return items or ["-"]

    def build_record_detail_text(self, record: AuditRecord) -> str:
        lines = [
            f"表单方法: {record.form_method or '-'}",
            f"表单Action: {record.form_action or '-'}",
            f"代理出口: {record.proxy_used or '直连'}",
            f"OCR接口: {record.ocr_endpoint_used or '-'}",
            f"OCR规则: {record.ocr_route_rule or '-'}",
            f"定位规则: {record.locator_rule_used or '-'}",
            f"画像标签: {', '.join(self.get_record_profile_tags(record)) or '-'}",
            f"字段摘要: {record.field_summary or '-'}",
            f"可用登录框: {'是' if self.is_actionable_login_record(record) else '否'}",
            f"登录评分: {record.login_score}",
            f"密码框数量: {record.password_field_count}",
            f"验证码: {'是' if record.captcha_present else '否'}",
            f"MFA: {'是' if record.mfa_present else '否'}",
            f"锁定提示: {'是' if record.lockout_hint else '否'}",
            f"默认账号提示: {'是' if record.default_hint else '否'}",
            "",
            "审计结果:",
        ]
        for index, item in enumerate(self.split_pipe_items(record.result), start=1):
            lines.append(f"{index}. {item}")
        lines.extend(["", "错误信息:"])
        for item in self.split_pipe_items(record.error or "-"):
            lines.append(f"- {item}")
        return "\n".join(lines)

    def open_selected_detail(self):
        record = self.get_selected_record()
        if not record:
            return
        self.open_record_detail_window(record)

    def open_record_detail_window(self, record: AuditRecord):
        window = tk.Toplevel(self.root)
        window.title(f"结果详情 - {record.title or record.target}")
        window.geometry("1100x760")
        window.transient(self.root)

        top_bar = tk.Frame(window, pady=8, padx=12)
        top_bar.pack(fill=tk.X)
        tk.Label(
            top_bar,
            text=f"目标: {record.target}\n最终URL: {record.final_url or '-'}\n状态: {record.status}    风险: {record.risk_level}",
            anchor="w",
            justify=tk.LEFT,
            wraplength=920,
        ).pack(side=tk.LEFT, fill=tk.X, expand=True)
        tk.Button(top_bar, text="打开目标", command=lambda: webbrowser.open(record.final_url or record.target)).pack(side=tk.RIGHT)

        body = tk.Frame(window, padx=12, pady=8)
        body.pack(fill=tk.BOTH, expand=True)
        body.columnconfigure(0, weight=1)
        body.rowconfigure(0, weight=1)

        y_scroll = ttk.Scrollbar(body, orient="vertical")
        x_scroll = ttk.Scrollbar(body, orient="horizontal")
        text_widget = tk.Text(
            body,
            wrap=tk.NONE,
            yscrollcommand=y_scroll.set,
            xscrollcommand=x_scroll.set,
        )
        y_scroll.config(command=text_widget.yview)
        x_scroll.config(command=text_widget.xview)
        text_widget.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")
        text_widget.insert(tk.END, self.build_record_detail_text(record))
        text_widget.config(state=tk.DISABLED)

        button_bar = tk.Frame(window, pady=8, padx=12)
        button_bar.pack(fill=tk.X)
        tk.Button(button_bar, text="关闭", command=window.destroy).pack(side=tk.RIGHT)

    def update_summary(self):
        total = len(self.all_records)
        completed = sum(1 for record in self.all_records if record.status == "已完成")
        pending = sum(1 for record in self.all_records if record.status != "已完成")
        high = sum(1 for record in self.all_records if record.risk_level == "高")
        medium = sum(1 for record in self.all_records if record.risk_level == "中")
        low = sum(1 for record in self.all_records if record.risk_level == "低")
        screenshots = sum(1 for record in self.all_records if record.screenshot_path)
        self.summary_var.set(
            f"摘要: 总计 {total} | 已完成 {completed} | 待处理 {pending} | 高 {high} | 中 {medium} | 低 {low} | 有截图 {screenshots}"
        )

    def ensure_selection(self):
        children = self.tree.get_children()
        if not children:
            self.update_preview(None)
            return
        selected = self.tree.selection()
        if not selected or selected[0] not in children:
            self.tree.selection_set(children[0])
            self.tree.focus(children[0])
            self.update_preview(self.records_by_item.get(children[0]))

    def select_relative_record(self, offset: int):
        children = list(self.tree.get_children())
        if not children:
            return
        selected = self.tree.selection()
        if selected and selected[0] in children:
            current_index = children.index(selected[0])
        else:
            current_index = 0
        next_index = max(0, min(len(children) - 1, current_index + offset))
        next_item = children[next_index]
        self.tree.selection_set(next_item)
        self.tree.focus(next_item)
        self.tree.see(next_item)
        self.update_preview(self.records_by_item.get(next_item))

    def find_item_id_by_record(self, target_record: AuditRecord):
        return self.item_id_by_record_key.get(target_record.record_id)

    def tag_for_record(self, record: AuditRecord) -> str:
        if record.risk_level == "高":
            return "risk_high"
        if record.risk_level == "中":
            return "risk_medium"
        if record.risk_level == "失败":
            return "status_error"
        return "risk_low"

    def update_tree_item(self, item_id, status: str, title: str, risk_level: str, result: str):
        current = self.tree.item(item_id)["values"]
        record = self.records_by_item.get(item_id)
        tags = (self.tag_for_record(record),) if record else ()
        self.tree.item(item_id, values=(current[0], current[1], status, title, risk_level, result), tags=tags)
        self.update_summary()
        if item_id in self.tree.selection():
            self.update_preview(record)

    def reset_scan_button(self):
        self.is_scanning = False
        self.btn_start.config(text="开始审计", bg="#9ed0ff")
        self.btn_retry_failed.config(state=tk.NORMAL)
        self.btn_rescan.config(state=tk.NORMAL)
        self.toggle_proxy_state()
        self.schedule_progress_snapshot(reason="ui-reset")

    def get_records_snapshot(self) -> list[AuditRecord]:
        with self.result_lock:
            return list(self.filtered_records())

    def save_progress_snapshot(self, reason: str = "autosave"):
        if not self.current_project_path and self.last_loaded_path:
            self.current_project_path = self.project_path_for_source(self.last_loaded_path)
        if not self.current_project_path:
            return
        try:
            payload = {
                "saved_at": datetime.now().isoformat(timespec="seconds"),
                "reason": reason,
                "source_file": self.last_loaded_path,
                "mode": self.mode_var.get(),
                "use_proxy": self.use_proxy_var.get(),
                "proxy": self.proxy_entry.get().strip(),
                "proxy_mode": self.proxy_mode_var.get(),
                "proxy_pool_path": self.proxy_pool_entry.get().strip(),
                "proxy_retry": self.proxy_retry_var.get(),
                "proxy_fail_threshold": self.proxy_fail_threshold_var.get(),
                "proxy_cooldown": self.proxy_cooldown_var.get(),
                "capture": self.capture_var.get(),
                "capture_policy": self.capture_policy_var.get(),
                "capture_delay": self.capture_delay_var.get(),
                "browser_render": self.browser_render_var.get(),
                "render_wait": self.render_wait_var.get(),
                "captcha_ocr": self.captcha_ocr_var.get(),
                "ocr_endpoint": self.ocr_endpoint_entry.get().strip(),
                "ocr_route_file": self.ocr_route_entry.get().strip(),
                "locator_rule_file": self.locator_rule_entry.get().strip(),
                "follow_redirect": self.follow_redirect_var.get(),
                "workers": self.worker_var.get(),
                "filter": self.filter_var.get(),
                "brute_enabled": self.brute_var.get(),
                "dict_mode": self.dict_mode_var.get(),
                "user_dict_path": self.user_dict_entry.get().strip(),
                "pass_dict_path": self.pass_dict_entry.get().strip(),
                "is_scanning": self.is_scanning,
                "records": [asdict(record) for record in self.all_records],
            }
            self.current_project_path.parent.mkdir(parents=True, exist_ok=True)
            temp_path = self.current_project_path.with_suffix(".tmp")
            with open(temp_path, "w", encoding="utf-8") as handle:
                json.dump(payload, handle, ensure_ascii=False, indent=2)
            os.replace(temp_path, self.current_project_path)
            self.write_latest_project_pointer(payload.get("saved_at", ""))
        except Exception as exc:
            self.log_queue.put(f"[~] 自动保存失败: {exc}")

    def load_autosave_if_exists(self):
        if not self.latest_project_path.exists():
            return
        try:
            with open(self.latest_project_path, "r", encoding="utf-8") as handle:
                latest = json.load(handle)
            project_file = latest.get("project_file", "")
            if not project_file:
                return
            self.load_project_file(Path(project_file))
        except Exception as exc:
            self.log_message(f"[~] 读取自动保存进度失败: {exc}")

    def on_close(self):
        self.is_scanning = False
        self.save_progress_snapshot(reason="window-close")
        self.root.destroy()

    def project_path_for_source(self, source_file: str) -> Path:
        source = str(Path(source_file).resolve())
        digest = hashlib.sha1(source.encode("utf-8")).hexdigest()[:12]
        safe_name = re.sub(r"[^0-9A-Za-z_\u4e00-\u9fa5-]+", "_", Path(source).stem)[:48] or "project"
        return self.projects_dir / f"{safe_name}_{digest}.json"

    def write_latest_project_pointer(self, saved_at: str):
        payload = {
            "project_file": str(self.current_project_path),
            "source_file": self.last_loaded_path,
            "saved_at": saved_at,
        }
        with open(self.latest_project_path, "w", encoding="utf-8") as handle:
            json.dump(payload, handle, ensure_ascii=False, indent=2)

    def load_project_file(self, project_path: Path, sync_source: bool = True):
        with open(project_path, "r", encoding="utf-8") as handle:
            payload = json.load(handle)
        records = payload.get("records") or []
        if not records:
            return
        self.proxy_assignment.clear()
        self.proxy_round_robin_index = 0
        self.proxy_health.clear()
        self.current_project_path = project_path
        self.set_detail_log_path(self.detail_log_path_for_project(project_path))
        self.all_records = [AuditRecord(**record) for record in records]
        for record in self.all_records:
            if record.status == "扫描中...":
                record.status = "等待中"
        self.last_loaded_path = payload.get("source_file", "")
        if self.last_loaded_path:
            self.file_entry.delete(0, tk.END)
            self.file_entry.insert(0, self.last_loaded_path)
        self.mode_var.set(payload.get("mode", "规则模式"))
        self.use_proxy_var.set(bool(payload.get("use_proxy", False)))
        self.proxy_entry.config(state=tk.NORMAL)
        self.proxy_entry.delete(0, tk.END)
        self.proxy_entry.insert(0, payload.get("proxy", "127.0.0.1:8080"))
        self.proxy_mode_var.set(payload.get("proxy_mode", PROXY_MODE_SINGLE))
        self.proxy_pool_entry.config(state=tk.NORMAL)
        self.proxy_pool_entry.delete(0, tk.END)
        self.proxy_pool_entry.insert(0, payload.get("proxy_pool_path", ""))
        self.proxy_retry_var.set(str(payload.get("proxy_retry", "2")))
        self.proxy_fail_threshold_var.set(str(payload.get("proxy_fail_threshold", "2")))
        self.proxy_cooldown_var.set(str(payload.get("proxy_cooldown", "120")))
        self.capture_var.set(bool(payload.get("capture", True)))
        saved_capture_policy = payload.get("capture_policy", CAPTURE_POLICY_HIT)
        self.capture_policy_var.set(saved_capture_policy if saved_capture_policy in CAPTURE_POLICY_OPTIONS else CAPTURE_POLICY_HIT)
        self.capture_delay_var.set(str(payload.get("capture_delay", "0.4")))
        self.browser_render_var.set(bool(payload.get("browser_render", True)))
        self.render_wait_var.set(str(payload.get("render_wait", "2.5")))
        self.captcha_ocr_var.set(bool(payload.get("captcha_ocr", True)))
        self.ocr_endpoint_entry.delete(0, tk.END)
        self.ocr_endpoint_entry.insert(0, payload.get("ocr_endpoint", DEFAULT_OCR_ENDPOINT))
        self.ocr_route_entry.delete(0, tk.END)
        self.ocr_route_entry.insert(0, payload.get("ocr_route_file", ""))
        self.locator_rule_entry.delete(0, tk.END)
        self.locator_rule_entry.insert(0, payload.get("locator_rule_file", ""))
        self.follow_redirect_var.set(bool(payload.get("follow_redirect", True)))
        self.worker_var.set(str(payload.get("workers", "4")))
        self.filter_var.set(payload.get("filter", "全部"))
        self.brute_var.set(bool(payload.get("brute_enabled", False)))
        self.dict_mode_var.set(payload.get("dict_mode", DEFAULT_BRUTE_DICT_MODE))
        self.user_dict_entry.config(state=tk.NORMAL)
        self.user_dict_entry.delete(0, tk.END)
        self.user_dict_entry.insert(0, payload.get("user_dict_path", ""))
        self.pass_dict_entry.config(state=tk.NORMAL)
        self.pass_dict_entry.delete(0, tk.END)
        self.pass_dict_entry.insert(0, payload.get("pass_dict_path", ""))
        self.toggle_proxy_state()
        self.toggle_brute_dict_state()
        sync_stats = None
        if sync_source and self.last_loaded_path:
            source_path = Path(self.last_loaded_path)
            if source_path.exists():
                sync_stats = self.sync_records_with_source(self.last_loaded_path)
            else:
                self.log_message(f"[~] 源文件不存在，沿用工程缓存列表: {source_path}")
        self.rebuild_tree()
        self.ensure_selection()
        self.mark_source_watch_baseline(self.last_loaded_path)
        completed = sum(1 for record in self.all_records if record.status == "已完成")
        pending = len(self.all_records) - completed
        self.log_message(
            f"[*] 已恢复工程: {Path(self.last_loaded_path).name if self.last_loaded_path else project_path.name} "
            f"| 已完成 {completed} 条 | 待继续 {pending} 条 | 保存时间 {payload.get('saved_at', '-')}"
        )
        if sync_stats is not None and (sync_stats["added"] or sync_stats["removed"]):
            self.log_message(
                f"[*] 已按源文件同步目标列表: 总计 {sync_stats['total']} | 新增 {sync_stats['added']} | 删除 {sync_stats['removed']} | 保留 {sync_stats['kept']}"
            )
        if not self.browser_render_var.get():
            self.log_message("[~] 当前工程恢复为“关闭浏览器补扫识别”，纯前端登录页可能识别不到登录框。")
        if self.capture_var.get() and self.capture_policy_var.get() == CAPTURE_POLICY_HIGH:
            self.log_message("[~] 当前工程截图策略是“仅高风险”，只有高风险结果才会截图。")

    def log_message(self, message: str):
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)
        self.write_detail_log(message, level="UI")

    def check_log_queue(self):
        try:
            while True:
                self.log_message(self.log_queue.get_nowait())
        except queue.Empty:
            pass
        finally:
            self.root.after(100, self.check_log_queue)


if __name__ == "__main__":
    root = tk.Tk()
    app = SecurityAuditGUI(root)
    root.mainloop()
